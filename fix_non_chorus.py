#!/usr/bin/env python3
"""
Script pour améliorer la détection des factures non-Chorus
"""

import os
import shutil
import re
from datetime import datetime
import logging

def fix_non_chorus_detection():
    """
    Améliore la détection des factures non-Chorus et ajoute des logs de diagnostic
    """
    main_file = "main.py"
    backup_file = f"main.py.backup_non_chorus_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    
    if not os.path.exists(main_file):
        print("Erreur: main.py non trouvé")
        return False
    
    print(f"Sauvegarde: {backup_file}")
    shutil.copy2(main_file, backup_file)
    
    try:
        with open(main_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 1. Améliorer la méthode est_facture_correspondante pour être plus flexible
        old_method_pattern = r'def est_facture_correspondante\(self, cell_text, facture_num\):.*?return False'
        new_method = '''def est_facture_correspondante(self, cell_text, facture_num):
        """
        Vérifie si une cellule contient le numéro de facture donné,
        avec une approche équilibrée entre stricte et souple
        
        Args:
            cell_text (str): Texte de la cellule à vérifier
            facture_num (str): Numéro de facture à rechercher
        
        Returns:
            bool: True si la cellule contient le numéro de facture, False sinon
        """
        if cell_text is None or facture_num is None:
            return False
            
        # Convertir en texte
        cell_text = str(cell_text).strip()
        facture_num = str(facture_num).strip()
        
        # Log pour diagnostic
        logger.debug(f"Comparaison: '{cell_text}' avec '{facture_num}'")
        
        # Vérification stricte: les textes sont identiques
        if cell_text.lower() == facture_num.lower():
            logger.debug("Correspondance exacte trouvée")
            return True
            
        # Extraire le numéro pur de la facture (après "Facture N°" ou similaire)
        facture_num_pur = facture_num.lower()
        for prefix in ["facture n°", "facture n ", "facture ", "fact ", "fact. ", "n°", "n "]:
            if facture_num_pur.startswith(prefix):
                facture_num_pur = facture_num_pur.replace(prefix, "", 1)
        facture_num_pur = facture_num_pur.strip()
        
        # Vérifier si la cellule contient exactement le numéro pur
        if cell_text.lower() == facture_num_pur:
            logger.debug(f"Correspondance avec numéro pur: '{facture_num_pur}'")
            return True
            
        # Vérifier si la cellule contient le numéro pur avec "Facture" ou autre préfixe
        cell_text_lower = cell_text.lower()
        if any(cell_text_lower == f"{prefix}{facture_num_pur}" for prefix in ["facture n°", "facture n ", "facture ", "fact ", "fact. ", "n°", "n "]):
            logger.debug(f"Correspondance avec préfixe et numéro pur")
            return True
            
        # Vérifier si c'est un nombre qui correspond
        try:
            # Si le numéro pur est un nombre et correspond à la cellule convertie en nombre
            if facture_num_pur.isdigit() and cell_text.replace(".", "").isdigit():
                if int(facture_num_pur) == int(float(cell_text)):
                    # Vérifier que ce n'est pas un nombre beaucoup plus grand contenant le numéro
                    if len(str(int(float(cell_text)))) <= len(facture_num_pur) + 1:
                        logger.debug(f"Correspondance numérique: {facture_num_pur} == {int(float(cell_text))}")
                        return True
        except (ValueError, TypeError):
            pass
            
        # Pas de correspondance trouvée
        return False'''
        
        # Utiliser une expression régulière pour remplacer la méthode
        import re
        content = re.sub(old_method_pattern, new_method, content, flags=re.DOTALL)
        
        # 2. Ajouter une fonction de diagnostic pour afficher toutes les cellules d'une feuille
        diagnostic_method = '''
    def afficher_contenu_feuille(self, sheet_name, facture_num, max_rows=30, max_cols=10):
        """
        Affiche le contenu des premières cellules d'une feuille pour diagnostic
        
        Args:
            sheet_name (str): Nom de la feuille à analyser
            facture_num (str): Numéro de facture recherché
            max_rows (int): Nombre maximum de lignes à afficher
            max_cols (int): Nombre maximum de colonnes à afficher
        """
        try:
            # Créer une copie temporaire du fichier Excel
            temp_file = os.path.join(os.path.dirname(self.current_file), f"temp_diagnostic_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            shutil.copy2(self.current_file, temp_file)
            
            logger.info(f"=== DIAGNOSTIC: Contenu de la feuille '{sheet_name}' pour recherche de '{facture_num}' ===")
            
            try:
                # Essayer d'abord avec openpyxl qui est plus fiable pour la lecture
                import openpyxl
                wb = openpyxl.load_workbook(temp_file)
                
                # Trouver la feuille correspondante
                sheet = None
                for ws in wb.worksheets:
                    if sheet_name.lower() in ws.title.lower():
                        sheet = ws
                        break
                
                if sheet:
                    for row in range(1, min(max_rows + 1, sheet.max_row + 1)):
                        row_values = []
                        for col in range(1, min(max_cols + 1, sheet.max_column + 1)):
                            cell_value = sheet.cell(row=row, column=col).value
                            row_values.append(str(cell_value) if cell_value is not None else "")
                        logger.info(f"Ligne {row}: {' | '.join(row_values)}")
                else:
                    logger.info(f"Feuille '{sheet_name}' non trouvée")
                    
            except Exception as e:
                logger.warning(f"Erreur lors de la lecture avec openpyxl: {e}")
                # Fallback à win32com
                try:
                    import win32com.client
                    excel = win32com.client.Dispatch("Excel.Application")
                    excel.Visible = False
                    excel.DisplayAlerts = False
                    workbook = excel.Workbooks.Open(temp_file)
                    
                    # Trouver la feuille correspondante
                    sheet = None
                    for i in range(1, workbook.Sheets.Count + 1):
                        if sheet_name.lower() in workbook.Sheets(i).Name.lower():
                            sheet = workbook.Sheets(i)
                            break
                    
                    if sheet:
                        for row in range(1, max_rows + 1):
                            row_values = []
                            for col in range(1, max_cols + 1):
                                cell_value = sheet.Cells(row, col).Value
                                row_values.append(str(cell_value) if cell_value is not None else "")
                            logger.info(f"Ligne {row}: {' | '.join(row_values)}")
                    else:
                        logger.info(f"Feuille '{sheet_name}' non trouvée")
                        
                    workbook.Close(False)
                    excel.Quit()
                except Exception as e:
                    logger.warning(f"Erreur lors de la lecture avec win32com: {e}")
            
            # Supprimer le fichier temporaire
            if os.path.exists(temp_file):
                os.remove(temp_file)
                
        except Exception as e:
            logger.error(f"Erreur lors du diagnostic: {e}")
        
        logger.info("=== FIN DIAGNOSTIC ===")'''
        
        # Ajouter la méthode de diagnostic à la classe MainWindow
        # Chercher la fin de la classe MainWindow
        class_end_pattern = r'if __name__ == "__main__":'
        # Insérer la méthode juste avant la fin de la classe
        content = content.replace(class_end_pattern, diagnostic_method + "\n\n" + class_end_pattern)
        
        # 3. Modifier la méthode save_invoice_file pour appeler le diagnostic en cas d'échec
        # Trouver les endroits où on log un échec de recherche de facture
        warning_pattern = r'logger\.warning\(f"Numéro de facture {facture_num}.*?non trouvé dans la feuille correspondant à l\'UH {uh}"\)'
        diagnostic_call = r'logger.warning(f"Numéro de facture {facture_num} (nettoyé: {facture_num_pur}) non trouvé dans la feuille correspondant à l\'UH {uh}")\n                    # Lancer le diagnostic pour voir le contenu de la feuille\n                    self.afficher_contenu_feuille(sheet.Name, facture_num)'
        
        # Remplacer pour la partie win32com
        content = content.replace(warning_pattern, diagnostic_call)
        
        # Faire de même pour la partie openpyxl
        warning_pattern_openpyxl = r'logger\.warning\(f"Numéro de facture {facture_num}.*?non trouvé dans la feuille correspondant à l\'UH {uh}"\)'
        diagnostic_call_openpyxl = r'logger.warning(f"Numéro de facture {facture_num} (nettoyé: {facture_num_pur}) non trouvé dans la feuille correspondant à l\'UH {uh}")\n                    # Lancer le diagnostic pour voir le contenu de la feuille\n                    self.afficher_contenu_feuille(sheet.title, facture_num)'
        
        # Remplacer pour la partie openpyxl (deuxième occurrence)
        content = content.replace(warning_pattern_openpyxl, diagnostic_call_openpyxl, 1)
        
        # 4. Augmenter le niveau de log pour voir plus de détails
        content = content.replace("logging.basicConfig(level=logging.INFO,", "logging.basicConfig(level=logging.DEBUG,")
        
        with open(main_file, 'w', encoding='utf-8') as f:
            f.write(content)
        
        print("[OK] Amélioration de la détection des factures non-Chorus et ajout de diagnostics!")
        return True
        
    except Exception as e:
        print(f"Erreur: {e}")
        if os.path.exists(backup_file):
            shutil.copy2(backup_file, main_file)
        return False

if __name__ == "__main__":
    print("=== Amélioration de la détection des factures non-Chorus ===")
    print()
    print("Problème identifié:")
    print("Les factures de clients non-Chorus ne sont pas trouvées dans les feuilles Excel")
    print()
    
    if fix_non_chorus_detection():
        print()
        print("=== Succès ===")
        print("Les améliorations suivantes ont été apportées:")
        print("1. Méthode de comparaison plus équilibrée pour les numéros de facture")
        print("2. Ajout d'une fonction de diagnostic qui affiche le contenu des feuilles")
        print("3. Activation automatique du diagnostic en cas d'échec de recherche")
        print("4. Augmentation du niveau de log pour plus de détails")
        print()
        print("Relancez l'application pour tester!")
        print("Consultez les logs pour voir le contenu des feuilles et comprendre pourquoi")
        print("certaines factures ne sont pas trouvées.")
    else:
        print("Échec de l'amélioration.")
