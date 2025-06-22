import os
import sys
import json
import pandas as pd
from PyQt5 import sip
from datetime import datetime
from openpyxl import load_workbook
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QLineEdit, QTextEdit,
    QDockWidget, QMenuBar, QAction, QFileDialog, QMessageBox, QProgressBar,
    QTabWidget, QMenu, QStatusBar, QFrame, QGroupBox, QComboBox, QCheckBox,
    QSpinBox, QStyle, QStyleFactory, QFormLayout, QSplitter, QDialog, 
    QDialogButtonBox, QScrollArea, QToolBar, QToolButton, QInputDialog, 
    QProgressDialog, QSystemTrayIcon, QSplashScreen
)
from floating_window import FloatingWindow

from PyQt5.QtCore import Qt, QTimer, QByteArray, QSize, QThread, pyqtSignal, QPoint, QEvent, QRect
from PyQt5.QtGui import QFont, QIcon, QColor, QStandardItemModel, QStandardItem, QPalette, QPixmap, QCursor, QPainter, QPen

# Configuration du logging
import logging
from logging.handlers import RotatingFileHandler

def setup_logging():
    """Configure le système de logging"""
    log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logs')
    os.makedirs(log_dir, exist_ok=True)
    
    log_file = os.path.join(log_dir, 'app.log')
    
    # Créer un formateur pour les messages de log
    formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    
    # Configurer le logger racine
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)
    
    # Ajouter un gestionnaire pour la sortie console
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)
    
    # Ajouter un gestionnaire pour le fichier de log
    file_handler = RotatingFileHandler(
        log_file, maxBytes=5*1024*1024, backupCount=5, encoding='utf-8'
    )
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    
    return logger

logger = setup_logging()

# Gestionnaire global d'exceptions
def global_exception_handler(exctype, value, tb):
    """Gestionnaire global des exceptions non attrapées"""
    logger.critical("Une erreur non gérée s'est produite", exc_info=(exctype, value, tb))
    sys.__excepthook__(exctype, value, tb)

sys.excepthook = global_exception_handler

# Palette de couleurs APHP
CYBERPUNK_COLORS = {
    'background': '#0a0e14',
    'secondary_bg': '#0f151d',
    'text': '#e6e6e6',
    'accent1': '#ff4d4d',  # Rouge APHP
    'accent2': '#ff9999',  # Rouge clair
    'accent3': '#cc0000',  # Rouge foncé
    'border': '#1a2634',
    'highlight': '#2a3b4d',
    'success': '#2ecc71',
    'warning': '#f39c12',
    'error': '#e74c3c',
    'drop_zone': 'rgba(0, 120, 255, 100)',
    'drop_zone_highlight': 'rgba(0, 120, 255, 200)'
}





class DraggableDockWidget(QDockWidget):
    """Widget dock avec fonctionnalités de drag & drop améliorées"""
    
    def __init__(self, title, parent=None):
        super().__init__(title, parent)
        self.setObjectName(f"{title.lower().replace(' ', '_')}Dock")
        self.setFeatures(QDockWidget.DockWidgetFeature.DockWidgetMovable | 
                         QDockWidget.DockWidgetFeature.DockWidgetFloatable |
                         QDockWidget.DockWidgetFeature.DockWidgetClosable)
        
        # Référence à la fenêtre principale
        self.main_window = parent
        
        # Variables pour le drag & drop
        self.drag_position = None
        
        # Personnaliser l'apparence
        self.setStyleSheet(f"""
            QDockWidget::title {{
                background-color: {CYBERPUNK_COLORS['accent1']};
                color: white;
                padding: 3px;
                border-radius: 2px;
            }}
            QDockWidget::title:hover {{
                background-color: {CYBERPUNK_COLORS['accent2']};
            }}
        """)
    
    def mousePressEvent(self, event):
        """Gérer le début d'un glisser-déposer"""
        # Ne capturer que les clics sur l'entête (la barre de titre)
        title_bar = self.titleBarWidget()
        title_height = 30  # Hauteur estimée de la barre de titre
        
        # Vérifier si le clic est dans la zone de la barre de titre
        if event.button() == Qt.MouseButton.LeftButton and event.y() < title_height:
            # Enregistrer la position relative pour le déplacement
            self.drag_position = event.pos()
        
        # Laisser Qt gérer normalement l'événement
        super().mousePressEvent(event)
    
    def mouseMoveEvent(self, event):
        """Gérer le déplacement pendant un glisser-déposer"""
        # Traitement standard par Qt
        super().mouseMoveEvent(event)
        
        # Si une fenêtre flottante est déplacée, activer les indicateurs visuels de la fenêtre principale
        if self.isFloating() and self.drag_position and self.main_window:
            # Activer le mode "accepter le dock" dans la fenêtre principale
            self.main_window.setDockOptions(
                QMainWindow.DockOption.AllowNestedDocks |
                QMainWindow.DockOption.AnimatedDocks |
                QMainWindow.DockOption.AllowTabbedDocks
            )
    
    def mouseReleaseEvent(self, event):
        """Gérer la fin d'un glisser-déposer"""
        # Réinitialiser la position de départ
        self.drag_position = None
        
        # Laisser Qt gérer normalement l'événement
        super().mouseReleaseEvent(event)

class Database:
    def __init__(self):
        self.data = {}
        self.db_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'database.json')
        self._loaded = False
        self._loading = False
        self._loader_thread = None
        self._on_loaded_callback = None
    
    def load_file(self, file_path):
        """Charge les données depuis un fichier JSON"""
        try:
            logger.info(f"Tentative de chargement du fichier JSON: {file_path}")
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                logger.info(f"Fichier JSON chargé avec succès: {len(data) if isinstance(data, dict) else 'format non valide'}")
                
                # Vérifier que les données sont au bon format
                if not isinstance(data, dict):
                    logger.error(f"Format de données JSON invalide: {type(data)}")
                    return False
                
                # Mettre à jour self.data avec les données chargées
                self.data = data
                logger.info(f"Base de données mise à jour avec {len(self.data)} entrées")
                return True
        except Exception as e:
            logger.error(f"Erreur lors du chargement du fichier {file_path}: {e}")
            return False
    
    def ensure_loaded(self, callback=None):
        """S'assure que la base de données est chargée"""
        if self._loaded:
            if callback:
                callback()
            return
        
        if self._loading:
            self._on_loaded_callback = callback
            return
        
        self._loading = True
        self._on_loaded_callback = callback
        self.load_database_async()
    
    def load_database(self):
        """Charge la base de données depuis le fichier JSON (synchrone)"""
        if os.path.exists(self.db_file):
            self.data = self.load_file(self.db_file)
        self._loaded = True
        self._loading = False
        return self.data
    
    def load_database_async(self):
        """Charge la base de données depuis le fichier JSON (asynchrone)"""
        if not os.path.exists(self.db_file):
            self._loaded = True
            self._loading = False
            if self._on_loaded_callback:
                self._on_loaded_callback()
            return
        
        self._loader_thread = DatabaseLoaderThread(self.db_file)
        self._loader_thread.finished.connect(self._on_database_loaded)
        self._loader_thread.error.connect(self._on_database_error)
        self._loader_thread.start()
    
    def _on_database_loaded(self, data):
        """Callback appelé lorsque la base de données est chargée"""
        self.data = data
        self._loaded = True
        self._loading = False
        
        if self._on_loaded_callback:
            self._on_loaded_callback()
    
    def _on_database_error(self, error_msg):
        """Callback appelé en cas d'erreur lors du chargement"""
        logger.error(error_msg)
        self._loading = False
        if self._on_loaded_callback:
            self._on_loaded_callback()
    
    def save_database(self):
        """Sauvegarde la base de données dans le fichier JSON"""
        try:
            with open(self.db_file, 'w', encoding='utf-8') as f:
                json.dump(self.data, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            logger.error(f"Erreur lors de la sauvegarde de la base de données: {e}")
            return False
    
    def load_from_dataframe(self, df, mapping):
        """Charge des données depuis un DataFrame pandas
        
        Args:
            df (pandas.DataFrame): DataFrame contenant les données à importer
            mapping (dict): Dictionnaire de mapping des colonnes {champ_cible: nom_colonne}
                - name: Nom du client (obligatoire)
                - client_code: Code client (optionnel)
                - chorus_code: Code Chorus (optionnel)
                - address: Adresse (optionnel)
                
        Returns:
            int: Nombre d'entrées ajoutées
        """
        logger.info(f"Début du chargement depuis DataFrame: {len(df)} lignes, mapping: {mapping}")
        
        # Vérifier que le mapping contient au moins le nom
        if not mapping.get('name'):
            logger.error("Mapping invalide: le champ 'name' est obligatoire")
            raise ValueError("Le mapping doit contenir au moins le champ 'name'")
        
        # Compteur d'entrées ajoutées
        entries_added = 0
        errors = 0
        
        # Parcourir les lignes du DataFrame
        for index, row in df.iterrows():
            try:
                # Récupérer les valeurs en fonction du mapping
                name = row.get(mapping['name'])
                
                # Vérifier que le nom n'est pas vide
                if pd.isna(name) or not str(name).strip():
                    logger.warning(f"Ligne {index+1}: Nom vide, ignorée")
                    continue
                
                # Convertir en chaîne de caractères et nettoyer
                name = str(name).strip()
                
                # Récupérer les autres valeurs (optionnelles)
                client_code = row.get(mapping.get('client_code')) if mapping.get('client_code') else None
                chorus_code = row.get(mapping.get('chorus_code')) if mapping.get('chorus_code') else None
                address = row.get(mapping.get('address')) if mapping.get('address') else None
                
                # Convertir les valeurs en chaînes de caractères si elles ne sont pas None
                if client_code is not None and not pd.isna(client_code):
                    client_code = str(client_code).strip()
                else:
                    client_code = None
                    
                if chorus_code is not None and not pd.isna(chorus_code):
                    chorus_code = str(chorus_code).strip()
                else:
                    chorus_code = None
                    
                if address is not None and not pd.isna(address):
                    address = str(address).strip()
                else:
                    address = None
                
                # Ajouter l'entrée à la base de données
                self.add_entry(name, client_code, chorus_code, address)
                entries_added += 1
                
                # Log périodique pour les gros fichiers
                if entries_added % 100 == 0:
                    logger.info(f"Progression: {entries_added} entrées traitées")
                    
            except Exception as e:
                errors += 1
                logger.error(f"Erreur lors du traitement de la ligne {index+1}: {str(e)}")
                # Continuer malgré l'erreur
        
        logger.info(f"Chargement terminé: {entries_added} entrées ajoutées, {errors} erreurs")
        return entries_added
        
    def add_entry(self, name, client_code=None, chorus_code=None, address=None):
        """Ajoute ou met à jour une entrée dans la base de données"""
        if not name:
            return False
        
        entry = {}
        if client_code:
            entry['client_code'] = client_code
        if chorus_code:
            entry['chorus_code'] = chorus_code
        if address:
            entry['address'] = address
        
        self.data[name] = entry
        return self.save_database()
    
    def load_from_dataframe(self, df, mapping):
        """Charge les données depuis un DataFrame pandas."""
        try:
            for _, row in df.iterrows():
                name = str(row[mapping['name']]).strip()
                if not name:
                    continue
                    
                entry = {}
                if 'client_code' in mapping and mapping['client_code'] in df.columns:
                    entry['client_code'] = str(row[mapping['client_code']]).strip()
                if 'chorus_code' in mapping and mapping['chorus_code'] in df.columns:
                    entry['chorus_code'] = str(row[mapping['chorus_code']]).strip()
                if 'address' in mapping and mapping['address'] in df.columns:
                    entry['address'] = str(row[mapping['address']]).strip()
                
                if entry:  # Ne pas ajouter d'entrées vides
                    self.data[name] = entry
            
            return self.save_database()
        except Exception as e:
            logger.error(f"Erreur lors du chargement des données depuis le DataFrame: {e}")
            return False
    
    def search_entries(self, query, category=None, exact_match=False):
        """Recherche dans la base de données
        
        Args:
            query (str): Texte à rechercher
            category (str, optional): Catégorie spécifique à rechercher (nom, client_code, chorus_code, address)
            exact_match (bool, optional): Si True, recherche une correspondance exacte au lieu d'une correspondance partielle
            
        Returns:
            dict: Dictionnaire des entrées correspondantes
        """
        if not query:
            return {}
        
        query = query.lower()
        results = {}
        
        for name, data in self.data.items():
            match_found = False
            
            # Recherche dans le nom
            if category is None or category == 'nom':
                if (exact_match and name.lower() == query) or \
                   (not exact_match and query in name.lower()):
                    results[name] = data
                    continue
            
            # Recherche dans les autres champs
            if category is None or category in ['client_code', 'chorus_code', 'address']:
                fields_to_search = [category] if category else ['client_code', 'chorus_code', 'address']
                
                for key in fields_to_search:
                    if key in data and data[key]:
                        value_str = str(data[key]).lower()
                        if (exact_match and value_str == query) or \
                           (not exact_match and query in value_str):
                            results[name] = data
                            match_found = True
                            break
                
                if match_found:
                    continue
        
        return results

class DatabaseLoaderThread(QThread):
    """Thread pour charger la base de données en arrière-plan"""
    finished = pyqtSignal(dict)
    progress = pyqtSignal(int)
    error = pyqtSignal(str)
    
    def __init__(self, db_file):
        super().__init__()
        self.db_file = db_file
    
    def run(self):
        try:
            with open(self.db_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            self.finished.emit(data)
        except Exception as e:
            self.error.emit(f"Erreur lors du chargement de la base de données: {e}")

class ConfirmationDialog(QDialog):
    def __init__(self, db_info, invoice_info, parent=None):
        super().__init__(parent)
        self.db_info = db_info
        self.invoice_info = invoice_info
        self.parent = parent
        self.setWindowTitle("Confirmation de correspondance")
        self.setMinimumWidth(800)
        self.setStyleSheet(f"""
        QDialog {{
            background-color: {CYBERPUNK_COLORS['background']};
        }}
        QLabel {{
            color: {CYBERPUNK_COLORS['accent3']};
            font-size: 12px;
        }}
        QLabel[title="true"] {{
            font-size: 14px;
            font-weight: bold;
            color: {CYBERPUNK_COLORS['accent1']};
        }}
        """)
        
        layout = QVBoxLayout()
        
        # Création de deux frames côte à côte
        comparison_layout = QHBoxLayout()
        
        # Frame gauche (base de données)
        db_frame = CustomFrame()
        db_layout = QVBoxLayout(db_frame)
        db_title = QLabel("Informations en base de données")
        db_title.setProperty("title", True)
        db_layout.addWidget(db_title)
        for key, value in self.db_info.items():
            if value:
                db_layout.addWidget(QLabel(f"{key}: {value}"))
        comparison_layout.addWidget(db_frame)
        
        # Frame droite (facture)
        invoice_frame = CustomFrame()
        invoice_layout = QVBoxLayout(invoice_frame)
        invoice_title = QLabel("Informations de la facture")
        invoice_title.setProperty("title", True)
        invoice_layout.addWidget(invoice_title)
        for key, value in self.invoice_info.items():
            if value:
                invoice_layout.addWidget(QLabel(f"{key}: {value}"))
        comparison_layout.addWidget(invoice_frame)
        
        layout.addLayout(comparison_layout)
        
        # Boutons
        button_layout = QHBoxLayout()
        accept_btn = CustomButton("Accepter")
        reject_btn = CustomButton("Refuser")
        
        accept_btn.clicked.connect(self.accept)
        reject_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(accept_btn)
        button_layout.addWidget(reject_btn)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)

class CustomButton(QPushButton):
    def __init__(self, text, parent=None):
        super().__init__(text, parent)
        self.setMinimumHeight(40)
        self.setFont(QFont('Segoe UI', 10))
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setStyleSheet(f"""
            QPushButton {{
                background-color: {CYBERPUNK_COLORS['accent1']};
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: 500;
                box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
            }}
            QPushButton:hover {{
                background-color: {CYBERPUNK_COLORS['accent2']};
            }}
            QPushButton:pressed {{
                background-color: {CYBERPUNK_COLORS['accent3']};
            }}
        """)

class CustomFrame(QFrame):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet(f"""
            QFrame {{
                background-color: {CYBERPUNK_COLORS['secondary_bg']};
                border: 1px solid {CYBERPUNK_COLORS['border']};
                border-radius: 8px;
                padding: 15px;
                box-shadow: 0 2px 6px rgba(0, 0, 0, 0.05);
            }}
        """)

# La classe StatsWidget a été supprimée car elle n'est plus utilisée

from PyQt5.QtCore import Qt, QSize, QThread, pyqtSignal, QTimer, QByteArray
from PyQt5.QtGui import QFont, QColor, QIcon, QPixmap
from openpyxl import load_workbook, Workbook
import logging
import traceback
from datetime import datetime
import shutil
import re
from unidecode import unidecode

def normalize_text(text):
    """Normalise le texte pour la recherche"""
    if text is None:
        return ""
    return str(text).lower().strip()

# Configuration du logging
def setup_logging():
    log_dir = "logs"
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)

    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = os.path.join(log_dir, f"factures_manager_{current_time}.log")

    logging.basicConfig(
        level=logging.DEBUG,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler(sys.stdout)
        ]
    )
    return logging.getLogger("FacturesManager")

logger = setup_logging()

# Gestionnaire global d'exceptions
def global_exception_handler(exctype, value, tb):
    error_msg = ''.join(traceback.format_exception(exctype, value, tb))
    logger.critical(f"Exception non gérée: {error_msg}")
    sys.__excepthook__(exctype, value, tb)  # Appel du gestionnaire par défaut

sys.excepthook = global_exception_handler

# Palette de couleurs APHP
CYBERPUNK_COLORS = {
    'background': '#FFFFFF',      # Fond blanc
    'secondary_bg': '#F5F5F5',    # Fond secondaire gris très clair
    'accent1': '#0099CC',         # Bleu principal APHP
    'accent2': '#005EB8',         # Bleu foncé APHP
    'accent3': '#003D8F',         # Bleu très foncé APHP
    'text': '#333333',            # Texte foncé
    'highlight': '#E6F7FF',       # Surbrillance bleu clair
    'border': '#CCCCCC',          # Bordures grises
    'success': '#2ecc71',         # Vert pour les succès
    'warning': '#f39c12',         # Orange pour les avertissements
    'error': '#e74c3c',           # Rouge pour les erreurs
    'drop_zone': 'rgba(0, 120, 255, 100)',
    'drop_zone_highlight': 'rgba(0, 120, 255, 200)'
}

# Constantes globales
EMPTY_CELL_SYMBOL = ""  # Cellules vides sans symbole

class Database:
    def __init__(self):
        self.data = {}
        self.db_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'database.json')
        self._loaded = False
        self._loading = False
        self._loader_thread = None
        self._on_loaded_callback = None

    def load_file(self, file_path):
        """Charge les données depuis un fichier JSON"""
        try:
            logger.info(f"Chargement du fichier de base de données: {file_path}")
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            # Normaliser les noms pour la recherche
            normalized_data = {}
            for key, value in data.items():
                normalized_key = normalize_text(key)
                normalized_data[normalized_key] = value

            self.data = normalized_data
            self._loaded = True
            logger.info(f"Base de données chargée avec succès: {len(self.data)} entrées")
            return True
        except Exception as e:
            logger.error(f"Erreur lors du chargement du fichier: {str(e)}")
            return False

    def ensure_loaded(self, callback=None):
        """S'assure que la base de données est chargée"""
        if self._loaded:
            # Déjà chargée, exécuter le callback immédiatement
            if callback:
                callback()
            return True
        elif self._loading:
            # Déjà en cours de chargement, enregistrer le callback
            if callback:
                self._on_loaded_callback = callback
            return False
        else:
            # Commencer le chargement
            self._loading = True
            self._on_loaded_callback = callback
            self.load_database_async()
            return False

    def load_database(self):
        """Charge la base de données depuis le fichier JSON (synchrone)"""
        if os.path.exists(self.db_file):
            try:
                with open(self.db_file, 'r', encoding='utf-8') as f:
                    self.data = json.load(f)
                    self._loaded = True
                    logger.info(f"Base de données chargée avec succès: {len(self.data)} entrées")
            except Exception as e:
                logger.error(f"Erreur lors du chargement de la base de données: {str(e)}")
        else:
            logger.warning("Fichier de base de données non trouvé, création d'une nouvelle base")
            self.data = {}
            self._loaded = True
            self.save_database()
    
    def load_database_async(self):
        """Charge la base de données depuis le fichier JSON (asynchrone)"""
        self._loader_thread = DatabaseLoaderThread(self.db_file)
        self._loader_thread.finished.connect(self._on_database_loaded)
        self._loader_thread.error.connect(self._on_database_error)
        self._loader_thread.start()
    
    def _on_database_loaded(self, data):
        """Callback appelé lorsque la base de données est chargée"""
        self.data = data
        self._loaded = True
        self._loading = False
        logger.info(f"Base de données chargée avec succès en arrière-plan: {len(self.data)} entrées")
        
        # Exécuter le callback si présent
        if self._on_loaded_callback:
            self._on_loaded_callback()
            self._on_loaded_callback = None
    
    def _on_database_error(self, error_msg):
        """Callback appelé en cas d'erreur lors du chargement"""
        logger.error(f"Erreur lors du chargement asynchrone de la base de données: {error_msg}")
        self.data = {}
        self._loaded = True
        self._loading = False
        
        # Exécuter le callback si présent
        if self._on_loaded_callback:
            self._on_loaded_callback()
            self._on_loaded_callback = None

    def save_database(self):
        """Sauvegarde la base de données dans le fichier JSON"""
        try:
            with open(self.db_file, 'w', encoding='utf-8') as f:
                json.dump(self.data, f, ensure_ascii=False, indent=4)
                logger.info("Base de données sauvegardée avec succès")
        except Exception as e:
            logger.error(f"Erreur lors de la sauvegarde de la base de données: {str(e)}")

    def add_entry(self, name, client_code=None, chorus_code=None, address=None):
        """Ajoute ou met à jour une entrée dans la base de données"""
        base_name = normalize_text(name)

        # Gérer les duplicatas en ajoutant un suffixe numérique si nécessaire
        final_name = base_name
        counter = 1
        while final_name in self.data:
            final_name = f"{base_name}_{counter}"
            counter += 1

        self.data[final_name] = {
            'name': name,
            'client_code': client_code if client_code is not None else "",
            'chorus_code': chorus_code if chorus_code is not None else "",
            'address': address if address is not None else ""
        }
        return final_name

    def load_from_dataframe(self, df, mapping):
        """Charge les données depuis un DataFrame pandas."""
        try:
            logger.info("Début du chargement depuis DataFrame")
            # Vérifier que toutes les colonnes mappées existent
            all_columns = []
            if isinstance(mapping['name'], list):
                all_columns.extend(mapping['name'])
            else:
                all_columns.append(mapping['name'])
            
            # Ajouter les autres colonnes mappées
            for key in ['client_code', 'chorus_code', 'address']:
                if key in mapping and mapping[key]:
                    if isinstance(mapping[key], list):
                        all_columns.extend(mapping[key])
                    else:
                        all_columns.append(mapping[key])

            logger.info(f"Colonnes requises: {all_columns}")
            logger.info(f"Colonnes disponibles: {df.columns.tolist()}")

            # Vérifier les colonnes manquantes
            missing_columns = [col for col in all_columns if col not in df.columns]
            if missing_columns:
                error_msg = f"Colonnes manquantes dans le fichier: {', '.join(missing_columns)}"
                logger.error(error_msg)
                raise ValueError(error_msg)

            logger.info("Toutes les colonnes requises sont présentes")
            logger.info("Mapping reçu: %s", mapping)

            entries_added = 0
            name_columns = mapping['name'] if isinstance(mapping['name'], list) else [mapping['name']]
            logger.info(f"Colonnes de nom à traiter: {name_columns}")

            # Pour chaque colonne de nom
            for name_column in name_columns:
                logger.info(f"Traitement de la colonne de nom: {name_column}")
                # Traiter chaque ligne
                for idx, row in df.iterrows():
                    try:
                        # Vérifier si la colonne existe dans la ligne
                        if name_column not in row:
                            logger.warning(f"Colonne {name_column} non trouvée dans la ligne {idx}")
                            continue
                            
                        name = row[name_column]
                        if pd.isna(name) or not str(name).strip():
                            logger.debug(f"Ligne {idx}: Nom vide ou null, ignorée")
                            continue

                        name = str(name).strip()
                        if not name:
                            logger.debug(f"Ligne {idx}: Nom vide après nettoyage, ignorée")
                            continue

                        # Extraire les autres champs avec gestion d'erreurs
                        client_code = None
                        chorus_code = None
                        address = None
                        
                        # Récupérer le code client s'il est mappé
                        if 'client_code' in mapping and mapping['client_code']:
                            client_code_col = mapping['client_code']
                            if client_code_col in row:
                                client_code = row[client_code_col]
                            else:
                                logger.warning(f"Colonne {client_code_col} non trouvée pour le code client dans la ligne {idx}")
                        
                        # Récupérer le code chorus s'il est mappé
                        if 'chorus_code' in mapping and mapping['chorus_code']:
                            chorus_code_col = mapping['chorus_code']
                            if chorus_code_col in row:
                                chorus_code = row[chorus_code_col]
                            else:
                                logger.warning(f"Colonne {chorus_code_col} non trouvée pour le code chorus dans la ligne {idx}")
                        
                        # Récupérer l'adresse si elle est mappée
                        if 'address' in mapping and mapping['address']:
                            address_col = mapping['address']
                            if address_col in row:
                                address = row[address_col]
                            else:
                                logger.warning(f"Colonne {address_col} non trouvée pour l'adresse dans la ligne {idx}")

                        # Convertir en string et nettoyer les valeurs
                        client_code = str(client_code).strip() if not pd.isna(client_code) and client_code is not None else ""
                        chorus_code = str(chorus_code).strip() if not pd.isna(chorus_code) and chorus_code is not None else ""
                        address = str(address).strip() if not pd.isna(address) and address is not None else ""

                        logger.info(f"Ligne {idx} - Traitement de l'entrée - Nom: {name}, Client: {client_code}, Chorus: {chorus_code}, Adresse: {address}")

                        # Ajouter l'entrée (permet les duplicatas car basé sur le nom normalisé)
                        self.add_entry(name, client_code, chorus_code, address)
                        entries_added += 1
                    except Exception as row_error:
                        logger.error(f"Erreur lors du traitement de la ligne {idx}: {str(row_error)}")
                        # Continuer avec la ligne suivante
                        continue

            logger.info(f"Sauvegarde de la base de données après ajout de {entries_added} entrées")
            self.save_database()
            logger.info(f"Chargement depuis DataFrame terminé avec succès: {entries_added} entrées ajoutées")
            return entries_added

        except Exception as e:
            logger.error(f"Erreur lors du chargement des données depuis DataFrame: {str(e)}")
            raise

    def search_entries(self, query, category=None, exact_match=False):
        """Recherche dans la base de données
        
        Args:
            query (str): Texte à rechercher
            category (str, optional): Catégorie spécifique à rechercher (nom, client_code, chorus_code, address)
            exact_match (bool, optional): Si True, recherche une correspondance exacte au lieu d'une correspondance partielle
            
        Returns:
            dict: Dictionnaire des entrées correspondantes
        """
        if not query:
            return {}
        
        query = query.lower()
        results = {}
        
        for name, entry in self.data.items():
            # Recherche dans tous les champs
            searchable_text = ' '.join(str(value).lower() for value in entry.values() if value)
            if query in searchable_text:
                results[name] = entry
        return results


class CustomButton(QPushButton):
    def __init__(self, text, parent=None):
        super().__init__(text, parent)
        self.setMinimumHeight(40)
        self.setFont(QFont('Segoe UI', 10))
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setStyleSheet(f"""
            QPushButton {{
                background-color: {CYBERPUNK_COLORS['accent1']};
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: 500;
                box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
            }}
            QPushButton:hover {{
                background-color: {CYBERPUNK_COLORS['accent2']};
            }}
            QPushButton:pressed {{
                background-color: {CYBERPUNK_COLORS['accent3']};
            }}
        """)


class CustomFrame(QFrame):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet(f"""
            QFrame {{
                background-color: {CYBERPUNK_COLORS['secondary_bg']};
                border: 1px solid {CYBERPUNK_COLORS['border']};
                border-radius: 8px;
                padding: 15px;
                box-shadow: 0 2px 6px rgba(0, 0, 0, 0.05);
            }}
        """)


# Pas besoin de classe personnalisée pour les fenêtres détachées
# Nous utiliserons directement les QDockWidget avec des configurations spéciales


# La classe StatsWidget a été supprimée car elle n'est plus utilisée


class DatabaseLoaderThread(QThread):
    """Thread pour charger la base de données en arrière-plan"""
    finished = pyqtSignal(dict)
    progress = pyqtSignal(int)
    error = pyqtSignal(str)
    
    def __init__(self, db_file):
        super().__init__()
        self.db_file = db_file
    
    def run(self):
        try:
            # Simuler un chargement progressif pour les grandes bases de données
            if os.path.exists(self.db_file):
                with open(self.db_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    
                    # Émettre des signaux de progression
                    total_items = len(data)
                    loaded_items = 0
                    
                    # Charger les données par lots pour éviter de bloquer l'interface
                    result = {}
                    for key, value in data.items():
                        result[key] = value
                        loaded_items += 1
                        
                        # Émettre la progression tous les 100 éléments
                        if loaded_items % 100 == 0 or loaded_items == total_items:
                            progress_percent = int(loaded_items * 100 / max(total_items, 1))
                            self.progress.emit(progress_percent)
                    
                    # Émettre le signal de fin avec les données chargées
                    self.finished.emit(result)
            else:
                # Base de données vide
                self.finished.emit({})
        
        except Exception as e:
            self.error.emit(str(e))
            logger.error(f"Erreur dans le thread de chargement: {str(e)}")
            self.finished.emit({})


class ConfirmationDialog(QDialog):
    def __init__(self, db_info, invoice_info, parent=None):
        super().__init__(parent)
        self.db_info = db_info
        self.invoice_info = invoice_info
        self.parent = parent
        self.setWindowTitle("Confirmation de correspondance")
        self.setMinimumWidth(800)
        self.setStyleSheet(f"""
        QDialog {{
            background-color: {CYBERPUNK_COLORS['background']};
        }}
        QLabel {{
            color: {CYBERPUNK_COLORS['accent3']};
            font-size: 12px;
        }}
        QLabel[title="true"] {{
            font-size: 14px;
            font-weight: bold;
            color: {CYBERPUNK_COLORS['accent1']};
        }}
        """)

        layout = QVBoxLayout()

        # Création de deux frames côte à côte
        comparison_layout = QHBoxLayout()

        # Frame gauche (base de données)
        db_frame = CustomFrame()
        db_layout = QVBoxLayout(db_frame)
        db_title = QLabel("Informations en base de données")
        db_title.setProperty("title", True)
        db_layout.addWidget(db_title)
        for key, value in self.db_info.items():
            if value:
                db_layout.addWidget(QLabel(f"{key}: {value}"))
        comparison_layout.addWidget(db_frame)

        # Frame droite (facture)
        invoice_frame = CustomFrame()
        invoice_layout = QVBoxLayout(invoice_frame)
        invoice_title = QLabel("Informations de la facture")
        invoice_title.setProperty("title", True)
        invoice_layout.addWidget(invoice_title)
        for key, value in self.invoice_info.items():
            if value:
                invoice_layout.addWidget(QLabel(f"{key}: {value}"))
        comparison_layout.addWidget(invoice_frame)

        layout.addLayout(comparison_layout)

        # Boutons
        button_layout = QHBoxLayout()
        accept_btn = CustomButton("Accepter")
        reject_btn = CustomButton("Refuser")

        accept_btn.clicked.connect(self.accept)
        reject_btn.clicked.connect(self.reject)

        button_layout.addWidget(accept_btn)
        button_layout.addWidget(reject_btn)
        layout.addLayout(button_layout)

        self.setLayout(layout)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Gestionnaire de Factures")
        self.resize(1200, 800)
        self.database = Database()
        self.current_excel_file = None
        self.manual_matches = {}
        self.current_tables = []
        
        # Fichier pour sauvegarder l'état de l'application
        self.state_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'app_state.json')
        
        # Indicateur de chargement
        self.loading_indicator = QProgressBar()
        self.loading_indicator.setRange(0, 0)  # Indicateur indéterminé
        self.loading_indicator.setMaximumWidth(200)
        self.loading_indicator.setTextVisible(False)
        
        # Gestion des modifications en attente
        self.pending_changes = {}
        self._updating_table = False  # Drapeau pour éviter les boucles de mise à jour
        
        # Activer les boutons de minimisation et plein écran pour la fenêtre principale
        self.setWindowFlags(self.windowFlags() | 
                            Qt.WindowMinimizeButtonHint | 
                            Qt.WindowMaximizeButtonHint)
        
        # Configurer l'interface utilisateur
        self.setup_ui()
        
        # S'assurer que les widgets dock sont visibles après un court délai
        QTimer.singleShot(500, self.restore_dock_widgets)
        
        # Charger l'état précédent de manière asynchrone
        QTimer.singleShot(100, self.load_state_async)
    
    def setup_dock_context_menu(self, dock_widget):
        """Configure le menu contextuel pour une fenêtre dockable"""
        dock_widget.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        dock_widget.customContextMenuRequested.connect(
            lambda pos: self.show_dock_context_menu(dock_widget, dock_widget.mapToGlobal(pos))
        )
    
    def show_dock_context_menu(self, dock_widget, pos):
        """Affiche le menu contextuel pour une fenêtre dockable"""
        menu = QMenu(self)
        
        # Action pour détacher/rattacher
        if dock_widget.isFloating():
            attach_action = QAction("Attacher la fenêtre", self)
            attach_action.triggered.connect(lambda: dock_widget.setFloating(False))
            menu.addAction(attach_action)
        else:
            detach_action = QAction("Détacher la fenêtre", self)
            detach_action.triggered.connect(lambda: dock_widget.setFloating(True))
            menu.addAction(detach_action)
        
        # Action pour fermer la fenêtre
        close_action = QAction("Fermer la fenêtre", self)
        close_action.triggered.connect(dock_widget.close)
        menu.addAction(close_action)
        
        # Afficher le menu
        menu.exec(pos)
    
    def setup_ui(self):
        """Configure l'interface utilisateur avec deux fenêtres principales"""
        # Créer le widget central et le layout principal
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Utiliser un QHBoxLayout comme layout principal
        main_layout = QHBoxLayout(central_widget)
        main_layout.setContentsMargins(5, 5, 5, 5)
        main_layout.setSpacing(5)
        
        # Créer un widget pour contenir les deux fenêtres
        self.container = QWidget()
        container_layout = QHBoxLayout(self.container)
        container_layout.setContentsMargins(0, 0, 0, 0)
        container_layout.setSpacing(5)
        
        # Créer les deux sections principales
        self.create_database_window()
        self.create_invoice_window()
        
        # Configurer les menus contextuels pour les fenêtres dockables
        for dock in [self.database_dock, self.invoice_dock]:
            self.setup_dock_context_menu(dock)
        
        # Ajouter les fenêtres au conteneur
        container_layout.addWidget(self.database_dock, 50)  # 50% de largeur
        container_layout.addWidget(self.invoice_dock, 50)   # 50% de largeur
        
        # Ajouter le conteneur au layout principal
        main_layout.addWidget(self.container)
        
        # Configurer la barre d'outils
        self.setup_toolbar()
        
        # Configurer la barre de menu
        self.setup_menu()
        
        # Afficher un indicateur de chargement
        self.statusBar().addPermanentWidget(self.loading_indicator)
        self.loading_indicator.hide()
        
        # Restaurer la géométrie et l'état de la fenêtre
        self.load_state()
        
        # Le comportement de fermeture est déjà géré par la méthode closeEvent
    
    def create_database_window(self):
        """Crée la fenêtre de la base de données"""
        self.database_dock = DraggableDockWidget("Base de données", self)
        self.database_widget = QWidget()
        self.database_layout = QVBoxLayout(self.database_widget)
        self.database_layout.setContentsMargins(5, 5, 5, 5)
        
        # Configurer l'interface de la base de données
        self.setup_database_interface()
        
        self.database_dock.setWidget(self.database_widget)
        
        # Connecter les signaux pour la gestion des fenêtres flottantes
        self.database_dock.topLevelChanged.connect(
            lambda: self.configure_floating_window(self.database_dock, self.database_dock.isFloating()))
    
    def create_invoice_window(self):
        """Crée la fenêtre des factures"""
        self.invoice_dock = DraggableDockWidget("Factures", self)
        self.invoice_widget = QWidget()
        self.invoice_layout = QVBoxLayout(self.invoice_widget)
        self.invoice_layout.setContentsMargins(5, 5, 5, 5)
        
        # Configurer l'interface des factures
        self.setup_invoice_interface()
        
        self.invoice_dock.setWidget(self.invoice_widget)
        
        # Connecter les signaux pour la gestion des fenêtres flottantes
        self.invoice_dock.topLevelChanged.connect(
            lambda: self.configure_floating_window(self.invoice_dock, self.invoice_dock.isFloating()))
    
    def add_database_entry(self):
        """Ajoute une nouvelle entrée vide à la base de données"""
        try:
            # Créer une nouvelle ligne vide dans le tableau
            row_position = self.db_table.rowCount()
            self.db_table.insertRow(row_position)
            
            # Ajouter des cellules vides
            self.db_table.setItem(row_position, 0, QTableWidgetItem(""))  # Nom
            self.db_table.setItem(row_position, 1, QTableWidgetItem(""))  # Code Client
            self.db_table.setItem(row_position, 2, QTableWidgetItem(""))  # Code Chorus
            self.db_table.setItem(row_position, 3, QTableWidgetItem(""))  # Adresse
            
            # Sélectionner la nouvelle ligne et la rendre éditable
            self.db_table.setCurrentCell(row_position, 0)
            self.db_table.editItem(self.db_table.item(row_position, 0))
            
            # Mettre à jour la base de données
            self.update_database_from_table()
            
            logger.info("Nouvelle entrée ajoutée à la base de données")
            
        except Exception as e:
            logger.error(f"Erreur lors de l'ajout d'une entrée à la base de données: {e}")
            QMessageBox.critical(self, "Erreur", f"Impossible d'ajouter une nouvelle entrée: {str(e)}")
    
    def filter_database(self):
        """Filtre les entrées de la base de données en fonction du texte de recherche et des options de filtrage"""
        search_text = self.db_search_edit.text().lower()
        category_index = self.filter_category.currentIndex()
        exact_match = self.exact_match.isChecked()
        
        visible_count = 0
        total_rows = self.db_table.rowCount()
        
        # Réinitialiser la mise en forme de toutes les cellules
        for row in range(total_rows):
            for col in range(self.db_table.columnCount()):
                item = self.db_table.item(row, col)
                if item:
                    item.setBackground(QColor(CYBERPUNK_COLORS['secondary_bg']))
                    font = item.font()
                    font.setBold(False)
                    item.setFont(font)
        
        # Si le champ de recherche est vide, tout afficher
        if not search_text:
            for row in range(total_rows):
                self.db_table.setRowHidden(row, False)
                visible_count += 1
            self.results_count.setText(f"{visible_count} résultats")
            return
        
        # Sinon, filtrer les lignes selon les critères
        for row in range(total_rows):
            match_found = False
            
            # Déterminer les colonnes à rechercher en fonction de la catégorie sélectionnée
            columns_to_search = []
            if category_index == 0:  # Tous les champs
                columns_to_search = range(self.db_table.columnCount())
            elif category_index == 1:  # Nom
                columns_to_search = [0]  # Colonne du nom
            elif category_index == 2:  # Code client
                columns_to_search = [1]  # Colonne du code client
            elif category_index == 3:  # Code chorus
                columns_to_search = [2]  # Colonne du code chorus
            elif category_index == 4:  # Adresse
                columns_to_search = [3]  # Colonne de l'adresse
            
            # Rechercher dans les colonnes spécifiées
            for col in columns_to_search:
                item = self.db_table.item(row, col)
                if not item:
                    continue
                    
                cell_text = item.text().lower()
                
                # Vérifier si le texte correspond selon le mode de recherche
                if (exact_match and cell_text == search_text) or \
                   (not exact_match and search_text in cell_text):
                    match_found = True
                    
                    # Mettre en évidence le texte trouvé
                    item.setBackground(QColor(CYBERPUNK_COLORS['highlight']))
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)
            
            # Afficher ou masquer la ligne selon le résultat
            self.db_table.setRowHidden(row, not match_found)
            if match_found:
                visible_count += 1
        
        # Mettre à jour le compteur de résultats
        self.results_count.setText(f"{visible_count} résultat{'s' if visible_count > 1 or visible_count == 0 else ''}")
    
    def update_database_from_table(self):
        """Met à jour la base de données à partir du tableau"""
        try:
            data = {}
            for row in range(self.db_table.rowCount()):
                # Récupérer le nom (clé du dictionnaire)
                nom_item = self.db_table.item(row, 1)  # La colonne 1 contient le nom
                if not nom_item or not nom_item.text().strip():
                    continue  # Ignorer les lignes sans nom
                    
                nom = nom_item.text().strip()
                
                # Créer l'entrée dans le dictionnaire
                data[nom] = {
                    'client_code': self.db_table.item(row, 2).text() if self.db_table.item(row, 2) else "",
                    'chorus_code': self.db_table.item(row, 3).text() if self.db_table.item(row, 3) else "",
                    'address': self.db_table.item(row, 4).text() if self.db_table.item(row, 4) else ""
                }
            
            # Mettre à jour la base de données
            self.database.data = data
            self.database.save()
            logger.info(f"Base de données mise à jour avec {len(data)} entrées")
            
        except Exception as e:
            error_msg = f"Erreur lors de la mise à jour de la base de données: {e}"
            logger.error(error_msg, exc_info=True)
            QMessageBox.critical(self, "Erreur", error_msg)
    
    def setup_database_interface(self):
        """Configure l'interface de la base de données"""
        # Barre de recherche
        search_layout = QHBoxLayout()
        self.db_search_edit = QLineEdit()
        self.db_search_edit.setPlaceholderText("Rechercher dans la base de données...")
        self.db_search_edit.textChanged.connect(self.filter_database)
        search_layout.addWidget(self.db_search_edit)
        
        # Bouton d'ajout
        add_btn = QPushButton("+")
        add_btn.setFixedSize(30, 30)
        add_btn.setToolTip("Ajouter une nouvelle entrée")
        add_btn.clicked.connect(self.add_database_entry)
        search_layout.addWidget(add_btn)
        
        self.database_layout.addLayout(search_layout)
        
    def on_db_cell_changed(self, row, column):
        """Appelé lorsqu'une cellule du tableau est modifiée"""
        if self._updating_table or not hasattr(self, 'db_table'):
            return
            
        try:
            # Récupérer le nom du client (colonne 1)
            name_item = self.db_table.item(row, 1)
            if not name_item or not name_item.text().strip():
                return
                
            name = name_item.text().strip()
            
            # Mettre à jour les modifications en attente
            self.pending_changes[name] = {
                'client_code': self.db_table.item(row, 2).text() if self.db_table.item(row, 2) else "",
                'chorus_code': self.db_table.item(row, 3).text() if self.db_table.item(row, 3) else "",
                'address': self.db_table.item(row, 4).text() if self.db_table.item(row, 4) else ""
            }
            
            # Démarrer/redémarrer le minuteur de sauvegarde (500ms)
            if hasattr(self, 'save_timer'):
                self.save_timer.start(500)
                
            logger.debug(f"Modification enregistrée pour {name}")
            
        except Exception as e:
            logger.error(f"Erreur lors de l'enregistrement de la modification: {e}")
    
    def save_pending_changes(self):
        """Sauvegarde les modifications en attente"""
        if not self.pending_changes:
            return
            
        try:
            # Mettre à jour uniquement les entrées modifiées
            for name, data in self.pending_changes.items():
                self.database.data[name] = data
                
            # Sauvegarder dans le fichier
            self.database.save()
            logger.info(f"Sauvegarde automatique de {len(self.pending_changes)} modification(s)")
            self.pending_changes.clear()
            
            # Mettre à jour l'interface utilisateur si nécessaire
            self.statusBar().showMessage("Modifications sauvegardées", 3000)
            
        except Exception as e:
            logger.error(f"Erreur lors de la sauvegarde automatique: {e}")
            QMessageBox.warning(self, "Erreur", 
                "Impossible de sauvegarder les modifications. Veuillez vérifier les logs.")
    
    def force_save_database(self):
        """Force la sauvegarde de la base de données"""
        if self.pending_changes:
            self.save_pending_changes()
        elif hasattr(self, 'database') and self.database:
            try:
                self.database.save()
                logger.info("Sauvegarde périodique de la base de données")
            except Exception as e:
                logger.error(f"Erreur lors de la sauvegarde périodique: {e}")
    
    def load_database_into_table(self):
        """Charge les données de la base de données dans le tableau"""
        try:
            # Désactiver temporairement les signaux pour éviter les mises à jour multiples
            self.db_table.blockSignals(True)
            self._updating_table = True
            self.pending_changes.clear()  # Vider les modifications en attente
            
            # Vider le tableau
            self.db_table.setRowCount(0)
            
            # Charger les données
            # Vérifier si self.database.data est un dictionnaire (format attendu)
            if isinstance(self.database.data, dict):
                logger.info(f"Chargement de {len(self.database.data)} entrées depuis le dictionnaire")
                for name, data in self.database.data.items():
                    row_position = self.db_table.rowCount()
                    self.db_table.insertRow(row_position)
                    
                    # Ajouter le numéro de ligne dans la première colonne (#)
                    self.db_table.setItem(row_position, 0, QTableWidgetItem(str(row_position + 1)))
                    
                    # Ajouter les données dans les autres colonnes
                    self.db_table.setItem(row_position, 1, QTableWidgetItem(name))
                    
                    # Vérifier si data est un dictionnaire
                    if isinstance(data, dict):
                        self.db_table.setItem(row_position, 2, QTableWidgetItem(str(data.get('client_code', ''))))
                        self.db_table.setItem(row_position, 3, QTableWidgetItem(str(data.get('chorus_code', ''))))
                        self.db_table.setItem(row_position, 4, QTableWidgetItem(str(data.get('address', ''))))
                    else:
                        # Si data n'est pas un dictionnaire, afficher une valeur vide
                        logger.warning(f"Données invalides pour {name}: {type(data)}")
                        self.db_table.setItem(row_position, 2, QTableWidgetItem(''))
                        self.db_table.setItem(row_position, 3, QTableWidgetItem(''))
                        self.db_table.setItem(row_position, 4, QTableWidgetItem(''))
            else:
                # Ancien format (liste d'objets)
                logger.warning(f"Format de données inattendu: {type(self.database.data)}")
                try:
                    for item in self.database.data:
                        row_position = self.db_table.rowCount()
                        self.db_table.insertRow(row_position)
                        
                        # Ajouter le numéro de ligne dans la première colonne (#)
                        self.db_table.setItem(row_position, 0, QTableWidgetItem(str(row_position + 1)))
                        
                        # Ajouter les données dans chaque colonne
                        self.db_table.setItem(row_position, 1, QTableWidgetItem(str(item.get('nom', ''))))
                        self.db_table.setItem(row_position, 2, QTableWidgetItem(str(item.get('code_client', ''))))
                        self.db_table.setItem(row_position, 3, QTableWidgetItem(str(item.get('code_chorus', ''))))
                        self.db_table.setItem(row_position, 4, QTableWidgetItem(str(item.get('adresse', ''))))
                except Exception as format_error:
                    logger.error(f"Erreur lors du traitement des données: {format_error}")
                    raise
            
            logger.info("Base de données chargée dans le tableau")
            
        except Exception as e:
            logger.error(f"Erreur lors du chargement de la base de données: {e}")
            QMessageBox.critical(self, "Erreur", 
                "Une erreur est survenue lors du chargement de la base de données.")
        finally:
            # Réactiver les signaux
            self.db_table.blockSignals(False)
            self._updating_table = False
            
            # Forcer une sauvegarde après le chargement initial
            QTimer.singleShot(1000, self.force_save_database)
    
    def setup_database_interface(self):
        """Configure l'interface de la base de données"""
        # Barre de recherche
        search_layout = QHBoxLayout()
        self.db_search_edit = QLineEdit()
        self.db_search_edit.setPlaceholderText("Rechercher dans la base de données...")
        self.db_search_edit.textChanged.connect(self.filter_database)
        search_layout.addWidget(self.db_search_edit)
        
        # Bouton d'ajout
        add_btn = QPushButton("+")
        add_btn.setFixedSize(30, 30)
        add_btn.setToolTip("Ajouter une nouvelle entrée")
        add_btn.clicked.connect(self.add_database_entry)
        search_layout.addWidget(add_btn)
        
        self.database_layout.addLayout(search_layout)
        
        # Tableau de la base de données
        self.db_table = QTableWidget()
        self.db_table.setColumnCount(5)  # Ajout d'une colonne pour le numéro de ligne
        self.db_table.setHorizontalHeaderLabels(["#", "Nom", "Code Client", "Code Chorus", "Adresse"])
        
        # Configuration des largeurs des colonnes
        header = self.db_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)  # Colonne # en largeur fixe
        for i in range(1, 5):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.Interactive)
        header.setStretchLastSection(True)  # Étirer la dernière section (Adresse)
        self.db_table.setColumnWidth(0, 50)  # Largeur fixe pour la colonne #
        
        self.db_table.verticalHeader().setVisible(False)
        self.db_table.setEditTriggers(QTableWidget.EditTrigger.DoubleClicked | 
                                     QTableWidget.EditTrigger.EditKeyPressed |
                                     QTableWidget.EditTrigger.SelectedClicked)
        self.db_table.itemChanged.connect(self.on_db_cell_changed)
        
        # Initialisation du tableau complet (utilisé pour le filtrage)
        self.full_db_table = QTableWidget()
        self.full_db_table.setColumnCount(5)  # Ajout d'une colonne pour le numéro de ligne
        self.full_db_table.setHorizontalHeaderLabels(["#", "Nom", "Code Client", "Code Chorus", "Adresse"])
        self.full_db_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.full_db_table.verticalHeader().setVisible(False)
        
        # Charger les données dans les tableaux
        self.load_database_into_table()
        
        self.database_layout.addWidget(self.db_table)
        
        # Boutons d'action
        btn_layout = QHBoxLayout()
        
        import_btn = QPushButton("Importer")
        import_btn.clicked.connect(self.import_database)
        btn_layout.addWidget(import_btn)
        
        export_excel_btn = QPushButton("Exporter Excel")
        export_excel_btn.clicked.connect(self.export_database_to_excel)
        btn_layout.addWidget(export_excel_btn)
        
        export_json_btn = QPushButton("Exporter JSON")
        export_json_btn.clicked.connect(self.export_database_to_json)
        btn_layout.addWidget(export_json_btn)
        
        clear_btn = QPushButton("Vider")
        clear_btn.clicked.connect(self.clear_database)
        btn_layout.addWidget(clear_btn)
        
        self.database_layout.addLayout(btn_layout)
    
    def import_database(self):
        """Importe des données dans la base de données depuis un fichier (Excel ou JSON)"""
        try:
            logger.info("Démarrage de l'importation de la base de données (première méthode)")
            # Ouvrir une boîte de dialogue pour sélectionner le fichier
            file_path, selected_filter = QFileDialog.getOpenFileName(
                self,
                "Sélectionner un fichier à importer",
                "",
                "Tous les fichiers supportés (*.json *.xlsx *.xls);;Fichiers JSON (*.json);;Fichiers Excel (*.xlsx *.xls);;Tous les fichiers (*)"
            )
            
            if not file_path:
                logger.info("Importation annulée par l'utilisateur")
                return  # L'utilisateur a annulé
            
            logger.info(f"Fichier sélectionné pour importation: {file_path}")
            
            # Vérifier que le fichier existe
            if not os.path.exists(file_path):
                logger.error(f"Le fichier n'existe pas: {file_path}")
                QMessageBox.critical(self, "Erreur", f"Le fichier n'existe pas: {file_path}")
                return
                
            # Déterminer le type de fichier
            file_ext = os.path.splitext(file_path)[1].lower()
            logger.info(f"Extension du fichier: {file_ext}")
            
            if file_ext == '.json':
                # Importer depuis JSON
                logger.info("Importation d'un fichier JSON")
                try:
                    # Lire le fichier JSON
                    with open(file_path, 'r', encoding='utf-8') as f:
                        logger.info(f"Ouverture du fichier JSON: {file_path}")
                        try:
                            data = json.load(f)
                            logger.info(f"Fichier JSON chargé avec succès, type: {type(data)}, contenu: {data if len(str(data)) < 500 else 'trop volumineux pour le log'}")
                        except json.JSONDecodeError as json_error:
                            logger.error(f"Erreur de décodage JSON: {str(json_error)}")
                            QMessageBox.critical(
                                self,
                                "Erreur",
                                f"Le fichier JSON est mal formé. Erreur: {str(json_error)}"
                            )
                            return
                    
                    # Vérifier si c'est un objet avec une clé 'data' ou directement une liste/dictionnaire
                    if isinstance(data, dict) and 'data' in data:
                        logger.info("Format JSON avec clé 'data' détecté")
                        data = data['data']
                    
                    # Vérifier que nous avons des données valides
                    if not isinstance(data, dict):
                        logger.error(f"Format de données JSON invalide: {type(data)}")
                        QMessageBox.critical(
                            self,
                            "Erreur",
                            f"Format de données JSON invalide. Attendu: dictionnaire, reçu: {type(data)}"
                        )
                        return
                    
                    if not data:
                        logger.warning("Le fichier JSON ne contient aucune donnée")
                        QMessageBox.warning(
                            self,
                            "Attention",
                            "Le fichier JSON ne contient aucune donnée."
                        )
                        return
                    
                    # Mettre à jour la base de données
                    logger.info(f"Mise à jour de la base de données avec {len(data)} entrées")
                    self.database.data = data
                    
                    # Sauvegarder la base de données
                    logger.info("Sauvegarde de la base de données après importation JSON")
                    try:
                        save_result = self.database.save_database()
                        logger.info(f"Résultat de la sauvegarde: {save_result}")
                    except Exception as save_error:
                        logger.error(f"Erreur lors de la sauvegarde de la base de données: {str(save_error)}")
                    
                    # Mettre à jour l'affichage
                    try:
                        # Essayer différentes méthodes possibles pour mettre à jour l'affichage
                        if hasattr(self, 'load_database_into_table'):
                            logger.info("Mise à jour de l'affichage avec load_database_into_table")
                            self.load_database_into_table()
                        elif hasattr(self, 'update_database_view'):
                            logger.info("Mise à jour de l'affichage avec update_database_view")
                            self.update_database_view()
                        elif hasattr(self, 'refresh_database_table'):
                            logger.info("Mise à jour de l'affichage avec refresh_database_table")
                            self.refresh_database_table()
                        elif hasattr(self, 'populate_database_table'):
                            logger.info("Mise à jour de l'affichage avec populate_database_table")
                            self.populate_database_table()
                        else:
                            logger.warning("Aucune méthode de mise à jour de l'affichage trouvée")
                    except Exception as update_error:
                        logger.error(f"Erreur lors de la mise à jour de l'affichage: {str(update_error)}")
                        # Continuer malgré l'erreur, l'importation a réussi
                    
                    # Afficher un message de succès avec des détails sur les données importées
                    entries_count = len(self.database.data)
                    sample_keys = list(self.database.data.keys())[:3] if self.database.data else []
                    sample_text = "\nExemples d'entrées importées:\n" + "\n".join(sample_keys) if sample_keys else ""
                    
                    QMessageBox.information(
                        self,
                        "Importation réussie",
                        f"La base de données a été importée avec succès depuis:\n{os.path.basename(file_path)}\n\n{entries_count} entrées ont été chargées.{sample_text}"
                    )
                    
                    # Journaliser l'importation avec des détails
                    logger.info(f"Importation JSON réussie: {entries_count} entrées chargées depuis {file_path}")
                    if sample_keys:
                        logger.info(f"Exemples d'entrées importées: {', '.join(sample_keys)}")
                    
                    # Mettre à jour le répertoire de travail pour les futures importations
                    if hasattr(self, 'last_directory'):
                        self.last_directory = os.path.dirname(file_path)
                        logger.info(f"Répertoire de travail mis à jour: {self.last_directory}")
                except Exception as json_error:
                    logger.error(f"Erreur lors de l'importation JSON: {str(json_error)}")
                    QMessageBox.critical(
                        self,
                        "Erreur",
                        f"Une erreur est survenue lors de l'importation du fichier JSON : {str(json_error)}"
                    )
            elif file_ext in ['.xlsx', '.xls']:
                # Importer depuis Excel
                logger.info("Importation d'un fichier Excel")
                try:
                    # Charger le fichier Excel
                    logger.info("Lecture du fichier Excel avec pandas")
                    try:
                        df = pd.read_excel(file_path)
                        logger.info(f"Fichier Excel lu avec succès: {len(df)} lignes, colonnes: {df.columns.tolist()}")
                    except Exception as excel_read_error:
                        logger.error(f"Erreur lors de la lecture du fichier Excel: {str(excel_read_error)}")
                        QMessageBox.critical(
                            self,
                            "Erreur",
                            f"Impossible de lire le fichier Excel. Vérifiez qu'il n'est pas corrompu ou ouvert dans une autre application.\n\nErreur: {str(excel_read_error)}"
                        )
                        return
                    
                    # Vérifier que le DataFrame n'est pas vide
                    if df.empty:
                        logger.warning("Le fichier Excel est vide")
                        QMessageBox.warning(self, "Attention", "Le fichier Excel est vide. Aucune donnée à importer.")
                        return
                    
                    # Afficher les colonnes disponibles
                    column_mapping_dialog = QDialog(self)
                    column_mapping_dialog.setWindowTitle("Mapping des colonnes")
                    column_mapping_dialog.setMinimumWidth(600)
                    
                    layout = QVBoxLayout()
                    
                    # Ajouter un label explicatif
                    layout.addWidget(QLabel("Sélectionnez les colonnes correspondantes dans votre fichier Excel:"))
                    
                    # Créer les combobox pour chaque type de données
                    mapping_layout = QGridLayout()
                    mapping_layout.addWidget(QLabel("Nom (obligatoire):"), 0, 0)
                    mapping_layout.addWidget(QLabel("Code Client (optionnel):"), 1, 0)
                    mapping_layout.addWidget(QLabel("Code Chorus (optionnel):"), 2, 0)
                    mapping_layout.addWidget(QLabel("Adresse (optionnel):"), 3, 0)
                    
                    name_combo = QComboBox()
                    client_code_combo = QComboBox()
                    chorus_code_combo = QComboBox()
                    address_combo = QComboBox()
                    
                    # Ajouter les colonnes disponibles
                    columns = ["-- Sélectionner --"] + df.columns.tolist()
                    for combo in [name_combo, client_code_combo, chorus_code_combo, address_combo]:
                        combo.addItems(columns)
                    
                    # Essayer de trouver automatiquement les colonnes pertinentes
                    for i, col in enumerate(df.columns):
                        col_lower = col.lower()
                        if "nom" in col_lower or "name" in col_lower or "client" in col_lower:
                            name_combo.setCurrentIndex(i + 1)  # +1 car le premier item est "-- Sélectionner --"
                        elif "code" in col_lower and ("client" in col_lower):
                            client_code_combo.setCurrentIndex(i + 1)
                        elif "chorus" in col_lower or "code chorus" in col_lower:
                            chorus_code_combo.setCurrentIndex(i + 1)
                        elif "adresse" in col_lower or "address" in col_lower:
                            address_combo.setCurrentIndex(i + 1)
                    
                    mapping_layout.addWidget(name_combo, 0, 1)
                    mapping_layout.addWidget(client_code_combo, 1, 1)
                    mapping_layout.addWidget(chorus_code_combo, 2, 1)
                    mapping_layout.addWidget(address_combo, 3, 1)
                    
                    layout.addLayout(mapping_layout)
                    
                    # Boutons
                    button_layout = QHBoxLayout()
                    ok_button = QPushButton("Importer")
                    cancel_button = QPushButton("Annuler")
                    
                    ok_button.clicked.connect(column_mapping_dialog.accept)
                    cancel_button.clicked.connect(column_mapping_dialog.reject)
                    
                    button_layout.addWidget(ok_button)
                    button_layout.addWidget(cancel_button)
                    
                    layout.addLayout(button_layout)
                    
                    column_mapping_dialog.setLayout(layout)
                    
                    logger.info("Affichage de la boîte de dialogue de mapping des colonnes")
                    if column_mapping_dialog.exec() == QDialog.DialogCode.Accepted:
                        # Récupérer le mapping des colonnes
                        mapping = {
                            'name': name_combo.currentText() if name_combo.currentIndex() > 0 else None,
                            'client_code': client_code_combo.currentText() if client_code_combo.currentIndex() > 0 else None,
                            'chorus_code': chorus_code_combo.currentText() if chorus_code_combo.currentIndex() > 0 else None,
                            'address': address_combo.currentText() if address_combo.currentIndex() > 0 else None
                        }
                        
                        logger.info(f"Mapping des colonnes sélectionné: {mapping}")
                        
                        # Vérifier qu'au moins le nom est mappé
                        if not mapping['name']:
                            logger.warning("Aucune colonne sélectionnée pour le nom")
                            QMessageBox.warning(self, "Erreur", "Vous devez au moins sélectionner une colonne pour le nom.")
                            return
                        
                        try:
                            # Charger les données dans la base de données
                            logger.info("Chargement des données dans la base de données")
                            entries_added = self.database.load_from_dataframe(df, mapping)
                            logger.info(f"Données chargées avec succès: {entries_added} entrées ajoutées")
                            
                            # Mettre à jour l'affichage
                            logger.info("Mise à jour de l'affichage après importation Excel")
                            self.update_database_view()
                            
                            QMessageBox.information(self, "Succès", f"{entries_added} entrées ajoutées à la base de données.")
                        except Exception as load_error:
                            logger.error(f"Erreur lors du chargement des données dans la base: {str(load_error)}")
                            QMessageBox.critical(
                                self,
                                "Erreur",
                                f"Une erreur est survenue lors du chargement des données dans la base de données.\n\nErreur: {str(load_error)}"
                            )
                    else:
                        logger.info("Importation Excel annulée par l'utilisateur")
                except Exception as excel_error:
                    logger.error(f"Erreur lors de l'importation Excel: {str(excel_error)}")
                    QMessageBox.critical(
                        self,
                        "Erreur",
                        f"Une erreur est survenue lors de l'importation du fichier Excel.\n\nErreur: {str(excel_error)}"
                    )
            else:
                logger.warning(f"Format de fichier non supporté: {file_ext}")
                QMessageBox.warning(
                    self,
                    "Attention",
                    "Format de fichier non supporté. Veuillez sélectionner un fichier JSON ou Excel."
                )
        except Exception as e:
            error_msg = f"Erreur lors de l'importation du fichier : {str(e)}"
            QMessageBox.critical(
                self,
                "Erreur d'importation",
                error_msg
            )
            logger.error(error_msg, exc_info=True)
    
    def export_database_to_json(self):
        """Exporte la base de données vers un fichier JSON"""
        try:
            # Sauvegarder d'abord les modifications en attente
            if hasattr(self, 'pending_changes') and self.pending_changes:
                logger.info(f"Sauvegarde des modifications en attente avant exportation: {len(self.pending_changes)} entrées")
                self.save_pending_changes()
            
            # Demander à l'utilisateur où sauvegarder le fichier
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Exporter la base de données",
                "",
                "Fichiers JSON (*.json);;Tous les fichiers (*)"
            )
            
            if not file_path:
                return  # L'utilisateur a annulé
                
            # S'assurer que l'extension est .json
            if not file_path.lower().endswith('.json'):
                file_path += '.json'
            
            # Mettre à jour la base de données depuis le tableau pour s'assurer que toutes les modifications sont prises en compte
            self.update_database_from_table()
            
            # Créer un dictionnaire avec les données à exporter
            export_data = {}
            for name, data in self.database.data.items():
                export_data[name] = {
                    'name': name,
                    'client_code': data.get('client_code', ''),
                    'chorus_code': data.get('chorus_code', ''),
                    'address': data.get('address', '')
                }
            
            # Écrire les données dans le fichier JSON avec une indentation pour une meilleure lisibilité
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, ensure_ascii=False, indent=4)
            
            logger.info(f"Base de données exportée avec succès vers {file_path} avec {len(export_data)} entrées")
            QMessageBox.information(
                self,
                "Exportation réussie",
                f"La base de données a été exportée avec succès vers :\n{file_path}\n\n{len(export_data)} entrées exportées.",
                QMessageBox.Ok
            )
            
        except Exception as e:
            logger.error(f"Erreur lors de l'exportation de la base de données : {str(e)}")
            QMessageBox.critical(
                self,
                "Erreur d'exportation",
                f"Une erreur est survenue lors de l'exportation de la base de données :\n{str(e)}",
                QMessageBox.Ok
            )
    
    def export_database_to_excel(self):
        """Exporte la base de données vers un fichier Excel"""
        try:
            # Demander à l'utilisateur où sauvegarder le fichier
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Exporter la base de données",
                "",
                "Fichiers Excel (*.xlsx);;Tous les fichiers (*)"
            )
            
            if not file_path:
                return  # L'utilisateur a annulé
                
            # S'assurer que l'extension est .xlsx
            if not file_path.lower().endswith('.xlsx'):
                file_path += '.xlsx'
            
            # Créer un nouveau classeur Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Base de données"
            
            # Écrire les en-têtes
            headers = ["Nom", "Code Client", "Code Chorus", "Adresse"]
            ws.append(headers)
            
            # Écrire les données
            for name, data in self.database.data.items():
                ws.append([
                    name,
                    data.get('client_code', ''),
                    data.get('chorus_code', ''),
                    data.get('address', '')
                ])
            
            # Sauvegarder le fichier
            wb.save(file_path)
            
            QMessageBox.information(
                self,
                "Exportation réussie",
                f"La base de données a été exportée avec succès vers :\n{file_path}"
            )
            logger.info(f"Base de données exportée vers {file_path}")
            
            # Mettre à jour le répertoire de travail
            self.last_directory = os.path.dirname(file_path)
            
        except Exception as e:
            error_msg = f"Erreur lors de l'exportation de la base de données : {str(e)}"
            QMessageBox.critical(
                self,
                "Erreur d'exportation",
                error_msg
            )
            logger.error(error_msg, exc_info=True)
    
    def clear_database(self):
        """Vide complètement la base de données après confirmation"""
        # Demander confirmation à l'utilisateur
        reply = QMessageBox.question(
            self,
            'Confirmation',
            'Êtes-vous sûr de vouloir vider complètement la base de données ?\nCette action est irréversible !',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                # Vider la base de données
                self.database.data = {}
                self.database.save_database()
                
                # Mettre à jour l'affichage
                self.load_database_into_table()
                
                QMessageBox.information(
                    self,
                    "Base de données vidée",
                    "La base de données a été vidée avec succès."
                )
                
            except Exception as e:
                QMessageBox.critical(
                    self,
                    "Erreur",
                    f"Une erreur est survenue lors de la suppression de la base de données :\n{str(e)}"
                )
                logger.error(f"Erreur lors de la suppression de la base de données: {str(e)}")
    
    def load_invoice_file(self):
        """Ouvre une boîte de dialogue pour charger un fichier de facture"""
        try:
            # Ouvrir la boîte de dialogue pour sélectionner un fichier
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "Ouvrir une facture",
                "",
                "Fichiers Excel (*.xlsx *.xls);;Tous les fichiers (*)"
            )
            
            if not file_path:
                return  # L'utilisateur a annulé
            
            # Enregistrer le chemin du fichier actuel
            self.current_invoice_path = file_path
            
            # Mettre à jour la barre de statut
            self.statusBar().showMessage(f"Chargement de la facture : {os.path.basename(file_path)}")
            
            # Traiter le fichier de facture
            self.process_invoice_file(file_path)
            
            # Mettre à jour le titre de la fenêtre avec le nom du fichier
            self.setWindowTitle(f"Gestionnaire de Factures - {os.path.basename(file_path)}")
            
            # Afficher un message de succès
            self.statusBar().showMessage(f"Facture chargée : {os.path.basename(file_path)}", 5000)
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Erreur lors du chargement de la facture",
                f"Une erreur est survenue lors du chargement du fichier :\n{str(e)}"
            )
            logger.error(f"Erreur lors du chargement de la facture: {str(e)}")
            self.statusBar().showMessage("Erreur lors du chargement de la facture", 5000)
    
    def process_invoice_file(self, file_path):
        """Traite un fichier Excel de facturation pour en extraire les informations
        
        Args:
            file_path (str): Chemin vers le fichier Excel à traiter
        """
        try:
            # Afficher un message de progression
            progress_dialog = QProgressDialog("Traitement du fichier de facturation...", "Annuler", 0, 100, self)
            progress_dialog.setWindowTitle("Importation des factures")
            progress_dialog.setWindowModality(Qt.WindowModality.WindowModal)
            progress_dialog.setMinimumDuration(0)
            progress_dialog.show()
            QApplication.processEvents()
            
            # Charger le fichier Excel avec openpyxl
            workbook = load_workbook(file_path, data_only=True)
            
            # Initialiser la liste des factures
            self.invoices = []
            
            # Nombre total de feuilles pour le calcul de progression
            total_sheets = len(workbook.sheetnames)
            processed_sheets = 0
            
            # Parcourir chaque feuille (UH)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Mettre à jour la progression
                processed_sheets += 1
                progress_percent = int((processed_sheets / total_sheets) * 100)
                progress_dialog.setValue(progress_percent)
                progress_dialog.setLabelText(f"Traitement de l'UH: {sheet_name}")
                QApplication.processEvents()
                
                if progress_dialog.wasCanceled():
                    break
                
                # Rechercher les cellules contenant "Intitulé"
                for row_idx in range(1, sheet.max_row + 1):
                    for col_idx in range(1, sheet.max_column + 1):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        
                        # Si la cellule contient "Intitulé", c'est le début d'une facture
                        if cell.value and "Intitulé" in str(cell.value):
                            try:
                                # Extraire le nom de la facture (6 cellules fusionnées à droite)
                                invoice_name_cell = sheet.cell(row=row_idx, column=col_idx + 1)
                                invoice_name = invoice_name_cell.value if invoice_name_cell.value else ""
                                
                                # Extraire le numéro de facture (7 cellules à droite)
                                invoice_number_cell = sheet.cell(row=row_idx, column=col_idx + 7)
                                invoice_number = invoice_number_cell.value if invoice_number_cell.value else ""
                                
                                # Extraire l'adresse (2 cellules en bas et 5 cellules à droite)
                                address_cell = sheet.cell(row=row_idx + 2, column=col_idx + 5)
                                address = address_cell.value if address_cell.value else ""
                                
                                # Vérifier si les cellules sont fusionnées et extraire toutes les valeurs
                                # Pour le nom de la facture (qui peut être sur plusieurs cellules)
                                for merged_cell in sheet.merged_cells.ranges:
                                    if invoice_name_cell.coordinate in merged_cell:
                                        # Extraire toutes les valeurs des cellules fusionnées
                                        for row in range(merged_cell.min_row, merged_cell.max_row + 1):
                                            for col in range(merged_cell.min_col, merged_cell.max_col + 1):
                                                cell_value = sheet.cell(row=row, column=col).value
                                                if cell_value and not invoice_name:
                                                    invoice_name = cell_value
                                
                                # Pour l'adresse (qui peut être sur plusieurs cellules)
                                for merged_cell in sheet.merged_cells.ranges:
                                    if address_cell.coordinate in merged_cell:
                                        # Extraire toutes les valeurs des cellules fusionnées
                                        for row in range(merged_cell.min_row, merged_cell.max_row + 1):
                                            for col in range(merged_cell.min_col, merged_cell.max_col + 1):
                                                cell_value = sheet.cell(row=row, column=col).value
                                                if cell_value and not address:
                                                    address = cell_value
                                
                                # Créer un dictionnaire pour la facture
                                invoice = {
                                    "uh": sheet_name,
                                    "numero": invoice_number,  # Numéro de facture
                                    "client": invoice_name,    # Nom de la facture
                                    "adresse": address,
                                    "nom_bdd": "",           # Nouvelle colonne Nom BDD
                                    "date": datetime.now().strftime("%Y-%m-%d"),  # Date par défaut
                                    "montant": 0.0,  # Montant par défaut
                                    "statut": "Importée",  # Statut par défaut
                                    "code_client": "",      # Colonne vide
                                    "code_chorus": "",      # Colonne vide
                                    "ligne_bdd": ""         # Colonne vide
                                }
                                
                                # Ajouter la facture à la liste
                                self.invoices.append(invoice)
                                
                                # Log pour débogage
                                logger.debug(f"Facture trouvée: {invoice}")
                            except Exception as e:
                                logger.error(f"Erreur lors de l'extraction d'une facture: {str(e)}")
            
            # Fermer la boîte de dialogue de progression
            progress_dialog.close()
            
            # Mettre à jour le tableau des factures
            self.update_invoice_table()
            
            # Afficher un message de succès
            QMessageBox.information(
                self,
                "Importation terminée",
                f"{len(self.invoices)} factures ont été importées avec succès."
            )
            
        except Exception as e:
            logger.error(f"Erreur lors du traitement du fichier de facturation: {str(e)}", exc_info=True)
            QMessageBox.critical(
                self,
                "Erreur de traitement",
                f"Une erreur est survenue lors du traitement du fichier: {str(e)}"
            )
    
    def update_invoice_table(self):
        """Met à jour le tableau des factures avec les données importées"""
        try:
            # Vérifier si le tableau des factures existe
            if not hasattr(self, 'invoice_table'):
                logger.warning("Le tableau des factures n'existe pas encore, création...")
                # Créer le tableau des factures s'il n'existe pas
                self.invoice_table = QTableWidget()
                self.invoice_table.setColumnCount(8)  # 8 colonnes pour les données de facture
                self.invoice_table.setHorizontalHeaderLabels(["UH", "N° Facture", "Nom facture", "Adresse facture", "Nom BDD", "Date", "Montant", "Statut"])
                self.invoice_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
                
                # Ajouter le tableau à l'interface
                # Vérifier si l'onglet des factures existe
                if hasattr(self, 'invoice_tab'):
                    # Ajouter le tableau à l'onglet des factures
                    layout = self.invoice_tab.layout()
                    if layout:
                        layout.addWidget(self.invoice_table)
                    else:
                        # Créer un layout pour l'onglet des factures
                        layout = QVBoxLayout(self.invoice_tab)
                        layout.addWidget(self.invoice_table)
                else:
                    # Ajouter le tableau directement à la fenêtre principale
                    central_widget = self.centralWidget()
                    if central_widget:
                        layout = central_widget.layout()
                        if layout:
                            layout.addWidget(self.invoice_table)
                        else:
                            # Créer un layout pour le widget central
                            layout = QVBoxLayout(central_widget)
                            layout.addWidget(self.invoice_table)
            
            # Effacer le tableau
            self.invoice_table.setRowCount(0)
            
            # Vérifier si la liste des factures existe
            if not hasattr(self, 'invoices') or not self.invoices:
                logger.warning("Aucune facture à afficher")
                return
            
            # Ajouter les factures au tableau
            for i, invoice in enumerate(self.invoices):
                self.invoice_table.insertRow(i)
                
                # Ajouter les données de la facture au tableau avec les nouvelles colonnes
                self.invoice_table.setItem(i, 0, QTableWidgetItem(invoice.get("uh", "")))
                self.invoice_table.setItem(i, 1, QTableWidgetItem(str(invoice.get("numero", ""))))  # N° Facture
                self.invoice_table.setItem(i, 2, QTableWidgetItem(invoice.get("client", "")))      # Nom facture
                self.invoice_table.setItem(i, 3, QTableWidgetItem(invoice.get("adresse", "")))     # Adresse facture
                # Remplir les colonnes avec des valeurs vides
                self.invoice_table.setItem(i, 4, QTableWidgetItem(""))                           # Nom BDD (vide)
                self.invoice_table.setItem(i, 5, QTableWidgetItem(""))                           # Code client (vide)
                self.invoice_table.setItem(i, 6, QTableWidgetItem(""))                           # Code chorus (vide)
                
                # Créer un item pour la colonne Ligne BDD qui sera éditable
                ligne_bdd_item = QTableWidgetItem("")
                ligne_bdd_item.setFlags(ligne_bdd_item.flags() | Qt.ItemFlag.ItemIsEditable)
                self.invoice_table.setItem(i, 7, ligne_bdd_item)                               # Ligne BDD (éditable)
                
                # Créer un bouton "Valider" pour la colonne Statut
                validate_btn = QPushButton("Valider")
                validate_btn.setStyleSheet("background-color: #4CAF50; color: white;")
                validate_btn.clicked.connect(lambda checked, row=i: self.validate_invoice_row(row))
                self.invoice_table.setCellWidget(i, 8, validate_btn)                          # Statut (bouton Valider)
            
            # Ajuster les colonnes pour qu'elles s'adaptent au contenu
            self.invoice_table.resizeColumnsToContents()
            
            logger.info(f"{len(self.invoices)} factures affichées dans le tableau")
            
            # Mettre à jour les statistiques
            self.update_statistics()
            
        except Exception as e:
            logger.error(f"Erreur lors de la mise à jour du tableau des factures: {str(e)}", exc_info=True)
    
    def save_invoice_file_old(self):
        """Ancienne version de la fonction de saisie des codes (conservée pour référence)"""
        # Cette fonction est conservée pour référence mais n'est plus utilisée
        pass
        
    def est_facture_correspondante(self, cell_text, facture_num):
        """Vérifie si une cellule contient un numéro de facture donné, en tenant compte de différents formats possibles"""
        if not cell_text or not facture_num:
            return False
            
        # Nettoyer les valeurs
        cell_text = str(cell_text).lower().strip()
        facture_num_clean = str(facture_num).lower().strip()
        
        # Vérifier plusieurs formats possibles
        return (facture_num_clean == cell_text or
                cell_text.endswith(facture_num_clean) or
                cell_text.startswith(facture_num_clean) or
                f" {facture_num_clean}" in cell_text or
                f"facture {facture_num_clean}" in cell_text or
                f"facture n°{facture_num_clean}" in cell_text or
                f"facture n° {facture_num_clean}" in cell_text or
                f"facture n {facture_num_clean}" in cell_text or
                f"fact {facture_num_clean}" in cell_text or
                f"fact. {facture_num_clean}" in cell_text)
    
    def save_invoice_file(self):
        """Saisie des codes dans le fichier de facturation pour les lignes validées (en bleu)"""
        try:
            # Vérifier si un fichier Excel est chargé
            if not hasattr(self, 'current_excel_file') or not self.current_excel_file:
                # Si aucun fichier n'est chargé, utiliser le dernier fichier ouvert
                if hasattr(self, 'current_invoice_path') and self.current_invoice_path and os.path.exists(self.current_invoice_path):
                    self.current_excel_file = self.current_invoice_path
                    logger.info(f"Utilisation du dernier fichier ouvert: {self.current_excel_file}")
                else:
                    QMessageBox.warning(self, "Attention", "Aucun fichier de facturation n'est chargé.")
                    return False
            
            # Collecter les lignes validées (en bleu) du tableau
            validated_rows = []
            for row in range(self.invoice_table.rowCount()):
                # Vérifier si la ligne est validée (en bleu)
                item = self.invoice_table.item(row, 0)  # Vérifier la première cellule
                if item and item.background().color().rgb() == QColor(173, 216, 230).rgb():
                    # Récupérer les informations de la ligne
                    row_data = {
                        'uh': self.invoice_table.item(row, 0).text() if self.invoice_table.item(row, 0) else "",
                        'facture_num': self.invoice_table.item(row, 1).text() if self.invoice_table.item(row, 1) else "",
                        'nom_facture': self.invoice_table.item(row, 2).text() if self.invoice_table.item(row, 2) else "",
                        'adresse_facture': self.invoice_table.item(row, 3).text() if self.invoice_table.item(row, 3) else "",
                        'nom_bdd': self.invoice_table.item(row, 4).text() if self.invoice_table.item(row, 4) else "",
                        'code_client': self.invoice_table.item(row, 5).text() if self.invoice_table.item(row, 5) else "",
                        'code_chorus': self.invoice_table.item(row, 6).text() if self.invoice_table.item(row, 6) else "",
                        'ligne_bdd': self.invoice_table.item(row, 7).text() if self.invoice_table.item(row, 7) else ""
                    }
                    validated_rows.append(row_data)
            
            if not validated_rows:
                QMessageBox.warning(self, "Attention", "Aucune ligne validée (en bleu) n'a été trouvée.")
                return False
            
            # Créer un nouveau nom de fichier
            file_name, file_ext = os.path.splitext(self.current_excel_file)
            new_file_path = f"{file_name}_updated{file_ext}"
            
            # Copier le fichier original en utilisant shutil.copy2 qui préserve les métadonnées
            try:
                shutil.copy2(self.current_excel_file, new_file_path)
                logger.info(f"Fichier copié avec succès: {new_file_path}")
            except Exception as e:
                logger.error(f"Erreur lors de la copie du fichier: {str(e)}")
                raise Exception(f"Impossible de copier le fichier Excel: {str(e)}")
            
            # Utiliser win32com pour manipuler Excel directement (plus fiable pour les fichiers complexes)
            try:
                import win32com.client
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False  # Ne pas afficher Excel
                excel.DisplayAlerts = False  # Désactiver les alertes
                
                # Ouvrir le fichier copié
                workbook = excel.Workbooks.Open(os.path.abspath(new_file_path))
                
                # Compteurs pour le rapport
                factures_traitees = 0
                
                # Pour chaque ligne validée, chercher la feuille correspondant à l'UH
                for row_data in validated_rows:
                    uh = row_data['uh']
                    facture_num = row_data['facture_num'].strip()
                    code_client = row_data['code_client']
                    code_chorus = row_data['code_chorus']
                    
                    # Extraire le numéro pur de la facture (sans préfixe)
                    facture_num_pur = None
                    if "facture n°" in facture_num.lower():
                        facture_num_pur = facture_num.lower().replace("facture n°", "").strip()
                    elif "facture n" in facture_num.lower():
                        facture_num_pur = facture_num.lower().replace("facture n", "").strip()
                    elif "facture" in facture_num.lower():
                        facture_num_pur = facture_num.lower().replace("facture", "").strip()
                    else:
                        facture_num_pur = facture_num.lower().strip()
                    
                    # Log pour débogage
                    logger.info(f"Traitement de la facture {facture_num} (UH: {uh}, numéro pur: {facture_num_pur})")
                    
                    # Facture trouvée et traitée pour cette ligne?
                    facture_traitee = False
                    
                    # Chercher une feuille qui correspond exactement à l'UH
                    feuille_uh_trouvee = False
                    for i in range(1, workbook.Sheets.Count + 1):
                        sheet = workbook.Sheets(i)
                        sheet_name = sheet.Name
                        
                        # Vérifier si cette feuille correspond à l'UH
                        if uh.lower() in sheet_name.lower():
                            feuille_uh_trouvee = True
                            logger.info(f"Feuille correspondant à l'UH {uh} trouvée: {sheet_name}")
                            
                            # Chercher le numéro de facture exact dans cette feuille spécifique
                            facture_trouvee = False
                            facture_row = None
                            facture_col = None
                            
                            # Rechercher dans toutes les cellules de la feuille
                            for row in range(1, 100):  # Limiter la recherche aux 100 premières lignes
                                for col in range(1, 20):  # Limiter aux 20 premières colonnes
                                    try:
                                        cell_value = sheet.Cells(row, col).Value
                                        if cell_value is not None:  # Vérifier que la cellule n'est pas vide
                                            # Convertir en string, quelle que soit la valeur
                                            cell_text = str(cell_value).lower().strip()
                                            
                                            # Extraire le numéro pur de la cellule
                                            cell_num_pur = None
                                            if "facture n°" in cell_text:
                                                cell_num_pur = cell_text.replace("facture n°", "").strip()
                                            elif "facture n" in cell_text:
                                                cell_num_pur = cell_text.replace("facture n", "").strip()
                                            elif "facture" in cell_text:
                                                cell_num_pur = cell_text.replace("facture", "").strip()
                                            else:
                                                cell_num_pur = cell_text
                                            
                                            # Vérification STRICTE: le numéro pur doit être EXACTEMENT égal
                                            # Utiliser une comparaison stricte avec == et non 'in'
                                            if cell_num_pur == facture_num_pur:
                                                facture_trouvee = True
                                                facture_row = row
                                                facture_col = col
                                                logger.info(f"Numéro de facture {facture_num} trouvé dans la cellule ({row}, {col}): '{cell_text}'")
                                                break
                                    except Exception as e:
                                        logger.debug(f"Erreur lors de la lecture de la cellule ({row}, {col}): {str(e)}")
                                        continue
                                
                                # Si on a trouvé la facture, on peut sortir de la boucle des lignes
                                if facture_trouvee:
                                    break
                            
                            # Si on a trouvé le numéro de facture exact, insérer les codes
                            if facture_trouvee:
                                # Insérer le code client: 6 cellules à gauche et 2 cellules en dessous du numéro de facture
                                if code_client:
                                    client_cell_row = facture_row + 2
                                    client_cell_col = facture_col - 6
                                    if client_cell_col < 1:  # Vérifier que la colonne est valide
                                        client_cell_col = 1
                                    
                                    # Log avant insertion pour vérification
                                    logger.info(f"Position du numéro de facture: ({facture_row}, {facture_col})")
                                    logger.info(f"Position calculée pour le code client: ({client_cell_row}, {client_cell_col})")
                                    
                                    sheet.Cells(client_cell_row, client_cell_col).Value = code_client
                                    logger.info(f"Code client {code_client} inséré dans la cellule ({client_cell_row}, {client_cell_col})")
                                
                                # Insérer le code chorus: 6 cellules à gauche et 1 cellule en dessous du numéro de facture
                                if code_chorus:
                                    chorus_cell_row = facture_row + 1
                                    chorus_cell_col = facture_col - 6
                                    if chorus_cell_col < 1:  # Vérifier que la colonne est valide
                                        chorus_cell_col = 1
                                    
                                    # Log avant insertion pour vérification
                                    logger.info(f"Position calculée pour le code chorus: ({chorus_cell_row}, {chorus_cell_col})")
                                    
                                    sheet.Cells(chorus_cell_row, chorus_cell_col).Value = code_chorus
                                    logger.info(f"Code chorus {code_chorus} inséré dans la cellule ({chorus_cell_row}, {chorus_cell_col})")
                                
                                factures_traitees += 1
                                facture_traitee = True
                                logger.info(f"Codes insérés pour la facture {facture_num} dans la feuille {sheet_name}")
                                break  # Sortir de la boucle des feuilles pour cette ligne
                    
                    if not feuille_uh_trouvee:
                        logger.warning(f"Aucune feuille correspondant à l'UH {uh} n'a été trouvée")
                    elif not facture_traitee:
                        logger.warning(f"Numéro de facture {facture_num} non trouvé dans la feuille correspondant à l'UH {uh}")
                
                # Sauvegarder et fermer le fichier
                workbook.Save()
                workbook.Close(True)
                excel.Quit()
                
                # Libérer les ressources COM
                del workbook
                del excel
                
            except ImportError:
                # Si win32com n'est pas disponible, utiliser openpyxl comme fallback
                logger.warning("win32com n'est pas disponible, utilisation d'openpyxl comme alternative")
                
                # Ouvrir le fichier avec openpyxl
                workbook = load_workbook(new_file_path)
                
                # Compteurs pour le rapport
                factures_traitees = 0
                
                # Pour chaque ligne validée, chercher la feuille correspondant à l'UH
                for row_data in validated_rows:
                    uh = row_data['uh']
                    facture_num = row_data['facture_num'].strip()
                    code_client = row_data['code_client']
                    code_chorus = row_data['code_chorus']
                    
                    # Extraire le numéro pur de la facture (sans préfixe)
                    facture_num_pur = None
                    if "facture n°" in facture_num.lower():
                        facture_num_pur = facture_num.lower().replace("facture n°", "").strip()
                    elif "facture n" in facture_num.lower():
                        facture_num_pur = facture_num.lower().replace("facture n", "").strip()
                    elif "facture" in facture_num.lower():
                        facture_num_pur = facture_num.lower().replace("facture", "").strip()
                    else:
                        facture_num_pur = facture_num.lower().strip()
                    
                    # Log pour débogage
                    logger.info(f"Traitement de la facture {facture_num} (UH: {uh}, numéro pur: {facture_num_pur})")
                    
                    # Facture trouvée et traitée pour cette ligne?
                    facture_traitee = False
                    
                    # Chercher une feuille qui correspond exactement à l'UH
                    feuille_uh_trouvee = False
                    for sheet_name in workbook.sheetnames:
                        # Vérifier si cette feuille correspond à l'UH
                        if uh.lower() in sheet_name.lower():
                            feuille_uh_trouvee = True
                            sheet = workbook[sheet_name]
                            logger.info(f"Feuille correspondant à l'UH {uh} trouvée: {sheet_name}")
                            
                            # Chercher le numéro de facture exact dans cette feuille spécifique
                            facture_trouvee = False
                            facture_row = None
                            facture_col = None
                            
                            # Rechercher dans toutes les cellules de la feuille
                            for row in range(1, min(100, sheet.max_row + 1)):  # Limiter la recherche aux 100 premières lignes
                                for col in range(1, min(20, sheet.max_column + 1)):  # Limiter aux 20 premières colonnes
                                    try:
                                        cell_value = sheet.cell(row=row, column=col).value
                                        if cell_value is not None:  # Vérifier que la cellule n'est pas vide
                                            # Convertir en string, quelle que soit la valeur
                                            cell_text = str(cell_value).lower().strip()
                                            
                                            # Extraire le numéro pur de la cellule
                                            cell_num_pur = None
                                            if "facture n°" in cell_text:
                                                cell_num_pur = cell_text.replace("facture n°", "").strip()
                                            elif "facture n" in cell_text:
                                                cell_num_pur = cell_text.replace("facture n", "").strip()
                                            elif "facture" in cell_text:
                                                cell_num_pur = cell_text.replace("facture", "").strip()
                                            else:
                                                cell_num_pur = cell_text
                                            
                                            # Vérification STRICTE: le numéro pur doit être EXACTEMENT égal
                                            # Utiliser une comparaison stricte avec == et non 'in'
                                            if cell_num_pur == facture_num_pur:
                                                facture_trouvee = True
                                                facture_row = row
                                                facture_col = col
                                                logger.info(f"Numéro de facture {facture_num} trouvé dans la cellule ({row}, {col}): '{cell_text}'")
                                                break
                                    except Exception as e:
                                        logger.debug(f"Erreur lors de la lecture de la cellule ({row}, {col}): {str(e)}")
                                        continue
                                
                                # Si on a trouvé la facture, on peut sortir de la boucle des lignes
                                if facture_trouvee:
                                    break
                            
                            # Si on a trouvé le numéro de facture exact, insérer les codes
                            if facture_trouvee:
                                # Insérer le code client: 6 cellules à gauche et 2 cellules en dessous du numéro de facture
                                if code_client:
                                    client_cell_row = facture_row + 2
                                    client_cell_col = facture_col - 6
                                    if client_cell_col < 1:  # Vérifier que la colonne est valide
                                        client_cell_col = 1
                                    
                                    # Log avant insertion pour vérification
                                    logger.info(f"Position du numéro de facture: ({facture_row}, {facture_col})")
                                    logger.info(f"Position calculée pour le code client: ({client_cell_row}, {client_cell_col})")
                                    
                                    sheet.cell(row=client_cell_row, column=client_cell_col).value = code_client
                                    logger.info(f"Code client {code_client} inséré dans la cellule ({client_cell_row}, {client_cell_col})")
                                
                                # Insérer le code chorus: 6 cellules à gauche et 1 cellule en dessous du numéro de facture
                                if code_chorus:
                                    chorus_cell_row = facture_row + 1
                                    chorus_cell_col = facture_col - 6
                                    if chorus_cell_col < 1:  # Vérifier que la colonne est valide
                                        chorus_cell_col = 1
                                    
                                    # Log avant insertion pour vérification
                                    logger.info(f"Position calculée pour le code chorus: ({chorus_cell_row}, {chorus_cell_col})")
                                    
                                    sheet.cell(row=chorus_cell_row, column=chorus_cell_col).value = code_chorus
                                    logger.info(f"Code chorus {code_chorus} inséré dans la cellule ({chorus_cell_row}, {chorus_cell_col})")
                                
                                factures_traitees += 1
                                facture_traitee = True
                                logger.info(f"Codes insérés pour la facture {facture_num} dans la feuille {sheet_name}")
                                break  # Sortir de la boucle des feuilles pour cette ligne
                    
                    if not feuille_uh_trouvee:
                        logger.warning(f"Aucune feuille correspondant à l'UH {uh} n'a été trouvée")
                    elif not facture_traitee:
                        logger.warning(f"Numéro de facture {facture_num} non trouvé dans la feuille correspondant à l'UH {uh}")
                
                # Sauvegarder et fermer le fichier
                workbook.save(new_file_path)
                workbook.close()
            
            except Exception as e:
                logger.error(f"Erreur lors de la manipulation du fichier Excel: {str(e)}")
                raise Exception(f"Erreur lors de la manipulation du fichier Excel: {str(e)}")
            
            # Vérifier que le fichier a été correctement créé
            if not os.path.exists(new_file_path) or os.path.getsize(new_file_path) == 0:
                raise Exception("Erreur lors de la création du fichier Excel. Le fichier est vide ou n'a pas été créé.")
            
            if factures_traitees > 0:
                # Informer l'utilisateur du succès
                QMessageBox.information(
                    self,
                    "Succès",
                    f"Les codes ont été insérés pour {factures_traitees} facture(s) dans le fichier:\n{new_file_path}"
                )
                
                # Proposer d'ouvrir le fichier Excel
                reply = QMessageBox.question(
                    self,
                    "Ouvrir le fichier",
                    "Voulez-vous ouvrir le fichier Excel généré ?",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.Yes
                )
                
                if reply == QMessageBox.Yes:
                    try:
                        # Ouvrir le fichier avec l'application par défaut
                        os.startfile(os.path.abspath(new_file_path))
                        logger.info(f"Fichier ouvert: {new_file_path}")
                    except Exception as e:
                        logger.error(f"Erreur lors de l'ouverture du fichier: {str(e)}")
                        QMessageBox.warning(self, "Attention", f"Impossible d'ouvrir le fichier: {str(e)}")
            else:
                QMessageBox.warning(
                    self,
                    "Attention",
                    "Aucune facture correspondante n'a été trouvée dans le fichier Excel."
                )
            
            # Mettre à jour le fichier courant
            self.current_excel_file = new_file_path
            self.setWindowTitle(f"Gestionnaire de Factures - {os.path.basename(new_file_path)}")
            
            # Mettre à jour la barre de statut
            self.statusBar().showMessage(f"Saisie des codes terminée : {os.path.basename(new_file_path)}", 5000)
            
            return True
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Erreur lors de la saisie des codes",
                f"Une erreur est survenue lors de la saisie des codes dans le fichier :\n{str(e)}"
            )
            logger.error(f"Erreur lors de la saisie des codes dans le fichier: {str(e)}")
            self.statusBar().showMessage("Erreur lors de la saisie des codes", 5000)
            import traceback
            logger.error(traceback.format_exc())
            return False
            
    def open_current_file_in_excel(self):
        """Ouvre le fichier actuel dans Microsoft Excel"""
        try:
            # Vérifier si un fichier est actuellement chargé
            if not hasattr(self, 'current_invoice_path') or not self.current_invoice_path:
                QMessageBox.information(
                    self,
                    "Aucun fichier ouvert",
                    "Aucun fichier n'est actuellement chargé."
                )
                return
                
            # Vérifier si le fichier existe
            if not os.path.exists(self.current_invoice_path):
                QMessageBox.critical(
                    self,
                    "Fichier introuvable",
                    f"Le fichier n'existe plus à l'emplacement :\n{self.current_invoice_path}"
                )
                return
                
            # Ouvrir le fichier avec l'application par défaut (Excel sur Windows)
            if os.name == 'nt':  # Windows
                os.startfile(self.current_invoice_path)
            elif os.name == 'posix':  # macOS et Linux
                if sys.platform == 'darwin':  # macOS
                    subprocess.run(['open', self.current_invoice_path])
                else:  # Linux
                    subprocess.run(['xdg-open', self.current_invoice_path])
            
            self.statusBar().showMessage(f"Ouverture de {os.path.basename(self.current_invoice_path)} dans Excel...", 3000)
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Erreur lors de l'ouverture dans Excel",
                f"Impossible d'ouvrir le fichier dans Excel :\n{str(e)}"
            )
            logger.error(f"Erreur lors de l'ouverture du fichier dans Excel: {str(e)}")
            self.statusBar().showMessage("Erreur lors de l'ouverture dans Excel", 5000)
    
    def set_modified(self, modified=True):
        """Marque l'application comme ayant des modifications non sauvegardées"""
        self._modified = modified
        
        # Mettre à jour le titre de la fenêtre pour indiquer l'état de modification
        title = self.windowTitle()
        if modified and not title.endswith('*'):
            self.setWindowTitle(f"{title} *")
        elif not modified and title.endswith(' *'):
            self.setWindowTitle(title[:-2])
        
        # Activer/désactiver les boutons de sauvegarde selon l'état de modification
        if hasattr(self, 'save_action'):
            self.save_action.setEnabled(modified)
        
        logger.debug(f"Application marquée comme {'modifiée' if modified else 'non modifiée'}")
    
    def validate_invoice_row(self, row):
        """Valide une ligne de facture et récupère les informations de la base de données si un numéro de ligne est présent"""
        try:
            # Vérifier s'il y a un numéro dans la colonne "Ligne BDD"
            ligne_bdd_item = self.invoice_table.item(row, 7)
            if ligne_bdd_item and ligne_bdd_item.text().strip():
                try:
                    ligne_num = int(ligne_bdd_item.text())
                    
                    # S'assurer que la base de données est chargée
                    if not hasattr(self, 'database') or not self.database.data or len(self.database.data) == 0:
                        QMessageBox.warning(
                            self,
                            "Base de données non chargée",
                            "La base de données n'est pas encore chargée ou est vide. Veuillez patienter ou importer des données."
                        )
                        return
                    
                    # Consulter la base de données pour trouver la ligne correspondante
                    if isinstance(self.database.data, dict):
                        # Récupérer la ligne de la base de données
                        # Convertir le dictionnaire en liste pour accéder par index
                        db_items = list(self.database.data.items())
                        if ligne_num > 0 and ligne_num <= len(db_items):
                            # Les indices commencent à 1 dans l'interface, mais à 0 dans la liste
                            db_index = ligne_num - 1
                            name, data = db_items[db_index]
                            
                            # Insérer les informations dans les colonnes correspondantes
                            nom_bdd_item = QTableWidgetItem(name)  # Nom BDD
                            self.invoice_table.setItem(row, 4, nom_bdd_item)
                            
                            if isinstance(data, dict):
                                code_client_item = QTableWidgetItem(str(data.get('client_code', '')))  # Code client
                                code_chorus_item = QTableWidgetItem(str(data.get('chorus_code', '')))  # Code chorus
                                
                                self.invoice_table.setItem(row, 5, code_client_item)
                                self.invoice_table.setItem(row, 6, code_chorus_item)
                                
                                logger.info(f"Facture ligne {row} validée avec les données de la ligne {ligne_num} de la base de données")
                            else:
                                QMessageBox.warning(
                                    self,
                                    "Données invalides",
                                    f"Les données pour la ligne {ligne_num} ne sont pas au format attendu."
                                )
                                return
                        else:
                            QMessageBox.warning(
                                self,
                                "Ligne introuvable",
                                f"La ligne {ligne_num} n'existe pas dans la base de données (qui contient {len(db_items)} entrées)."
                            )
                            return
                    else:
                        QMessageBox.warning(
                            self,
                            "Base de données vide",
                            "La base de données est vide ou n'est pas au format attendu."
                        )
                        return
                except ValueError:
                    QMessageBox.warning(
                        self,
                        "Valeur invalide",
                        "Le numéro de ligne doit être un entier valide."
                    )
                    return
            
            # Colorer toute la ligne en bleu (après avoir ajouté les éléments)
            for col in range(self.invoice_table.columnCount()):
                item = self.invoice_table.item(row, col)
                if item:
                    item.setBackground(QColor(173, 216, 230))  # Bleu clair
            
            # Marquer l'application comme modifiée
            self.set_modified(True)
            
            # Conserver le bouton Valider mais changer sa couleur pour indiquer qu'il a été validé
            validate_cell = self.invoice_table.cellWidget(row, 8)
            if isinstance(validate_cell, QPushButton):
                validate_cell.setStyleSheet("background-color: #87CEFA; color: white; font-weight: bold;")  # Bleu clair plus vif
                validate_cell.setText("Revalider")
            
            logger.info(f"Facture ligne {row} validée")
            
        except Exception as e:
            logger.error(f"Erreur lors de la validation de la facture: {str(e)}")
            QMessageBox.critical(
                self,
                "Erreur",
                f"Une erreur est survenue lors de la validation de la facture:\n{str(e)}"
            )
    
    def on_invoice_item_changed(self, item):
        """Gère les modifications dans le tableau des factures"""
        try:
            # Vérifier si la modification provient d'un utilisateur (et non du programme)
            if not hasattr(self, '_updating_invoice'):
                self._updating_invoice = False
                
            if self._updating_invoice:
                return
                
            # Marquer que nous sommes en train de mettre à jour le tableau
            self._updating_invoice = True
            
            # Récupérer la ligne et la colonne modifiées
            row = item.row()
            col = item.column()
            new_value = item.text()
            
            # Vérifier si c'est la colonne Ligne BDD
            ligne_bdd_col = self.invoice_columns.index("Ligne BDD")
            if col == ligne_bdd_col:
                # Valider que la valeur est un nombre
                try:
                    if new_value.strip() != "":
                        # Essayer de convertir en entier
                        ligne_num = int(new_value)
                        # Mettre à jour avec la valeur validée
                        item.setText(str(ligne_num))
                        
                        # Mettre à jour l'état de modification
                        self.set_modified(True)
                        logger.info(f"Ligne BDD mise à jour pour la ligne {row}: {ligne_num}")
                    else:
                        # Valeur vide autorisée
                        pass
                except ValueError:
                    # Si ce n'est pas un nombre valide, réinitialiser à vide
                    QMessageBox.warning(
                        self,
                        "Valeur invalide",
                        "Veuillez entrer un nombre entier pour la ligne de la base de données."
                    )
                    item.setText("")
            
        except Exception as e:
            logger.error(f"Erreur lors de la modification de l'élément de facture: {str(e)}")
            QMessageBox.critical(
                self,
                "Erreur",
                f"Une erreur est survenue lors de la modification de l'élément :\n{str(e)}"
            )
        finally:
            # Réactiver les mises à jour
            self._updating_invoice = False
            
    def load_state(self):
        """Charge l'état précédent de l'application"""
        try:
            # Chemin du fichier de configuration pour sauvegarder l'état
            config_dir = os.path.join(os.path.expanduser('~'), '.factures_manager')
            os.makedirs(config_dir, exist_ok=True)
            state_file = os.path.join(config_dir, 'app_state.json')
            
            # Vérifier si le fichier d'état existe
            if not os.path.exists(state_file):
                return
                
            # Charger l'état depuis le fichier
            with open(state_file, 'r', encoding='utf-8') as f:
                state = json.load(f)
                
            # Restaurer la géométrie de la fenêtre
            if 'geometry' in state:
                self.restoreGeometry(QByteArray.fromHex(state['geometry'].encode()))
                
            # Restaurer l'état de la fenêtre (maximisé, etc.)
            if 'window_state' in state:
                self.restoreState(QByteArray.fromHex(state['window_state'].encode()))
                
            # Restaurer le dossier de travail précédent
            if 'last_directory' in state and os.path.isdir(state['last_directory']):
                self.last_directory = state['last_directory']
                
            # Restaurer le chemin du fichier de facture récent s'il existe, mais ne pas le charger automatiquement
            if 'recent_invoice' in state and os.path.exists(state['recent_invoice']):
                self.current_invoice_path = state['recent_invoice']
                # Ne pas charger automatiquement le fichier
                # self.load_invoice_file()
                
            logger.info("État de l'application chargé avec succès")
            
        except Exception as e:
            logger.error(f"Erreur lors du chargement de l'état de l'application: {str(e)}")
            # Ne pas afficher de message d'erreur à l'utilisateur pour cette opération
            
    def save_state(self):
        """Sauvegarde l'état actuel de l'application"""
        try:
            # Préparer les données à sauvegarder
            state = {
                'geometry': self.saveGeometry().toHex().data().decode(),
                'window_state': self.saveState().toHex().data().decode(),
                'last_directory': getattr(self, 'last_directory', os.path.expanduser('~')),
            }
            
            # Ajouter le fichier de facture actuel s'il existe
            if hasattr(self, 'current_invoice_path') and self.current_invoice_path:
                state['recent_invoice'] = self.current_invoice_path
            
            # Créer le dossier de configuration s'il n'existe pas
            config_dir = os.path.join(os.path.expanduser('~'), '.factures_manager')
            os.makedirs(config_dir, exist_ok=True)
            state_file = os.path.join(config_dir, 'app_state.json')
            
            # Sauvegarder dans le fichier
            with open(state_file, 'w', encoding='utf-8') as f:
                json.dump(state, f, indent=4, ensure_ascii=False)
                
            logger.info("État de l'application sauvegardé avec succès")
            
        except Exception as e:
            logger.error(f"Erreur lors de la sauvegarde de l'état de l'application: {str(e)}")
            # Ne pas afficher de message d'erreur à l'utilisateur pour cette opération
            
    def closeEvent(self, event):
        """Gère l'événement de fermeture de la fenêtre"""
        try:
            # Vérifier s'il y a des modifications non enregistrées
            if hasattr(self, 'is_modified') and self.is_modified:
                reply = QMessageBox.question(
                    self,
                    'Confirmation',
                    'Voulez-vous enregistrer les modifications avant de quitter ?',
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel,
                    QMessageBox.StandardButton.Yes
                )
                
                if reply == QMessageBox.StandardButton.Yes:
                    # Essayer d'enregistrer
                    if hasattr(self, 'save_invoice_file'):
                        if not self.save_invoice_file():
                            # L'utilisateur a annulé l'enregistrement
                            event.ignore()
                            return
                elif reply == QMessageBox.StandardButton.Cancel:
                    # Annuler la fermeture
                    event.ignore()
                    return
            
            # Sauvegarder les modifications en attente de la base de données
            if hasattr(self, 'pending_changes') and self.pending_changes:
                try:
                    self.save_pending_changes()
                except Exception as e:
                    logger.error(f"Erreur lors de la sauvegarde finale: {e}")
            
            # Arrêter les minuteurs
            if hasattr(self, 'save_timer') and self.save_timer.isActive():
                self.save_timer.stop()
            if hasattr(self, 'auto_save_timer') and self.auto_save_timer.isActive():
                self.auto_save_timer.stop()
            
            # Fermer correctement toutes les ressources
            if hasattr(self, 'cleanup_resources'):
                self.cleanup_resources()
            
            # Sauvegarder l'état avant de quitter
            self.save_state()
            
            # Accepter l'événement de fermeture
            event.accept()
            
        except Exception as e:
            logger.error(f"Erreur lors de la fermeture de l'application: {str(e)}")
            # En cas d'erreur, forcer la fermeture
            event.accept()
            
    def load_state_async(self):
        """Charge l'état de l'application de manière asynchrone"""
        try:
            # Cette méthode est appelée de manière asynchrone après le chargement de l'interface
            # pour éviter de bloquer l'interface utilisateur
            
            # Ne pas charger automatiquement le dernier fichier Excel ouvert au démarrage
            # Cela empêche l'ouverture automatique de l'explorateur Windows
            if hasattr(self, 'current_invoice_path') and self.current_invoice_path:
                if os.path.exists(self.current_invoice_path):
                    # Mettre à jour le titre de la fenêtre avec le nom du fichier
                    self.setWindowTitle(f"Gestionnaire de Factures - {os.path.basename(self.current_invoice_path)}")
                    # Activer le bouton pour ouvrir avec Excel
                    if hasattr(self, 'open_in_excel_btn'):
                        self.open_in_excel_btn.setEnabled(True)
            
            # Autres opérations de chargement asynchrone peuvent être ajoutées ici
            
            logger.info("Chargement asynchrone de l'état terminé")
            
        except Exception as e:
            logger.error(f"Erreur lors du chargement asynchrone de l'état: {str(e)}")
    
    def setup_invoice_interface(self):
        """Configure l'interface des factures"""
        # Initialisation du tableau des factures
        self.invoice_table = QTableWidget()
        
        # Définition des colonnes
        self.invoice_columns = [
            "UH",                    # 0
            "N°Facture",            # 1
            "Nom facture",          # 2
            "Adresse facture",      # 3
            "Nom BDD",              # 4
            "Code client",          # 5
            "Code chorus",          # 6
            "Ligne BDD",            # 7
            "Statut"                # 8
        ]
        
        self.invoice_table.setColumnCount(len(self.invoice_columns))
        self.invoice_table.setHorizontalHeaderLabels(self.invoice_columns)
        
        # Configuration de l'en-tête
        header = self.invoice_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)  # Permet le redimensionnement manuel
        header.setStretchLastSection(False)  # Ne pas étirer la dernière section
        
        # Définition des largeurs par défaut pour certaines colonnes
        column_widths = {
            "UH": 80,               # 0
            "N°Facture": 100,       # 1
            "Nom facture": 200,     # 2
            "Adresse facture": 250, # 3
            "Nom BDD": 150,         # 4
            "Code client": 100,     # 5
            "Code chorus": 100,     # 6
            "Ligne BDD": 100,       # 7
            "Statut": 100           # 8
        }
        
        # Application des largeurs personnalisées
        for col, width in column_widths.items():
            if col in self.invoice_columns:
                idx = self.invoice_columns.index(col)
                self.invoice_table.setColumnWidth(idx, width)
        
        # Rendre certaines colonnes non modifiables
        for col in [0, 1, 2, 3, 4, 5, 6, 8]:  # Toutes les colonnes sauf Ligne BDD
            for row in range(self.invoice_table.rowCount()):
                item = self.invoice_table.item(row, col)
                if item:
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        
        # S'assurer que la colonne Ligne BDD est éditable
        ligne_bdd_col = self.invoice_columns.index("Ligne BDD")
        for row in range(self.invoice_table.rowCount()):
            item = self.invoice_table.item(row, ligne_bdd_col)
            if item:
                item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
        
        # Barre d'outils
        toolbar = QHBoxLayout()
        
        open_btn = QPushButton("Ouvrir")
        open_btn.clicked.connect(self.load_invoice_file)
        toolbar.addWidget(open_btn)
        
        save_btn = QPushButton("Saisie des codes")
        save_btn.clicked.connect(self.save_invoice_file)
        toolbar.addWidget(save_btn)
        
        save_as_btn = QPushButton("Traitement")
        save_as_btn.clicked.connect(self.save_invoice_as)
        toolbar.addWidget(save_as_btn)
        
        self.invoice_layout.addLayout(toolbar)
        
        # Barre de recherche
        search_layout = QHBoxLayout()
        self.invoice_search_edit = QLineEdit()
        self.invoice_search_edit.setPlaceholderText("Rechercher dans les factures...")
        self.invoice_search_edit.textChanged.connect(self.filter_invoices)
        search_layout.addWidget(self.invoice_search_edit)
        
        self.invoice_layout.addLayout(search_layout)
        # Configuration du tableau
        self.invoice_table.verticalHeader().setVisible(False)
        self.invoice_table.setEditTriggers(QTableWidget.EditTrigger.DoubleClicked)
        self.invoice_table.itemChanged.connect(self.on_invoice_item_changed)
        
        # Connecter le changement de sélection pour mettre à jour les statistiques
        self.invoice_table.itemSelectionChanged.connect(self.update_statistics)
        
        # Activer le tri
        self.invoice_table.setSortingEnabled(True)
        
        # Configurer le tableau
        self.db_table.itemChanged.connect(self.on_db_cell_changed)
        
        # Configurer le minuteur de sauvegarde
        self.save_timer = QTimer(self)
        self.save_timer.setSingleShot(True)
        self.save_timer.timeout.connect(self.save_pending_changes)
        
        # Sauvegarde périodique toutes les 5 minutes
        self.auto_save_timer = QTimer(self)
        self.auto_save_timer.timeout.connect(self.force_save_database)
        self.auto_save_timer.start(5 * 60 * 1000)  # 5 minutes
        
        self.invoice_layout.addWidget(self.invoice_table)
        
        # Zone de statistiques
        self.stats_label = QLabel("Statistiques")
        self.stats_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # Création du conteneur pour les barres de progression
        self.stats_container = QWidget()
        self.stats_layout = QVBoxLayout(self.stats_container)
        
        # Barre de progression pour les lignes bleues (Validé)
        self.blue_progress_layout = QHBoxLayout()
        self.blue_progress_label = QLabel("Validé:")
        self.blue_progress = QProgressBar()
        self.blue_progress.setStyleSheet("QProgressBar { background-color: #f0f0f0; border: 1px solid #ccc; border-radius: 5px; text-align: center; } "
                                         "QProgressBar::chunk { background-color: #0078d7; border-radius: 5px; }")
        self.blue_progress.setTextVisible(True)
        self.blue_progress.setFormat("%p%")
        self.blue_progress_layout.addWidget(self.blue_progress_label)
        self.blue_progress_layout.addWidget(self.blue_progress)
        self.stats_layout.addLayout(self.blue_progress_layout)
        
        # Barre de progression pour les lignes vertes (Concordance exacte)
        self.green_progress_layout = QHBoxLayout()
        self.green_progress_label = QLabel("Concordance exacte:")
        self.green_progress = QProgressBar()
        self.green_progress.setStyleSheet("QProgressBar { background-color: #f0f0f0; border: 1px solid #ccc; border-radius: 5px; text-align: center; } "
                                          "QProgressBar::chunk { background-color: #2ecc71; border-radius: 5px; }")
        self.green_progress.setTextVisible(True)
        self.green_progress.setFormat("%p%")
        self.green_progress_layout.addWidget(self.green_progress_label)
        self.green_progress_layout.addWidget(self.green_progress)
        self.stats_layout.addLayout(self.green_progress_layout)
        
        # Barre de progression pour les lignes orange (Concordance partielle)
        self.orange_progress_layout = QHBoxLayout()
        self.orange_progress_label = QLabel("Concordance partielle:")
        self.orange_progress = QProgressBar()
        self.orange_progress.setStyleSheet("QProgressBar { background-color: #f0f0f0; border: 1px solid #ccc; border-radius: 5px; text-align: center; } "
                                           "QProgressBar::chunk { background-color: #f39c12; border-radius: 5px; }")
        self.orange_progress.setTextVisible(True)
        self.orange_progress.setFormat("%p%")
        self.orange_progress_layout.addWidget(self.orange_progress_label)
        self.orange_progress_layout.addWidget(self.orange_progress)
        self.stats_layout.addLayout(self.orange_progress_layout)
        
        # Barre de progression pour les lignes sans couleur (Pas de concordance)
        self.no_color_progress_layout = QHBoxLayout()
        self.no_color_progress_label = QLabel("Pas de concordance:")
        self.no_color_progress = QProgressBar()
        self.no_color_progress.setStyleSheet("QProgressBar { background-color: #f0f0f0; border: 1px solid #ccc; border-radius: 5px; text-align: center; } "
                                             "QProgressBar::chunk { background-color: #bdc3c7; border-radius: 5px; }")
        self.no_color_progress.setTextVisible(True)
        self.no_color_progress.setFormat("%p%")
        self.no_color_progress_layout.addWidget(self.no_color_progress_label)
        self.no_color_progress_layout.addWidget(self.no_color_progress)
        self.stats_layout.addLayout(self.no_color_progress_layout)
        
        self.invoice_layout.addWidget(self.stats_label)
        self.invoice_layout.addWidget(self.stats_container)
        
        # Initialiser les barres de progression à zéro
        self.update_statistics()
    
    def on_invoice_item_changed(self, item):
        """
        Appelé lorsqu'une cellule du tableau des factures est modifiée.
        Met à jour les statistiques après la modification.
        """
        # Mettre à jour les statistiques
        self.update_statistics()
        
        # Journaliser la modification
        row = item.row()
        col = item.column()
        logger.debug(f"Cellule modifiée dans le tableau des factures: ligne {row}, colonne {col}")
    
    def validate_invoice_row(self, row):
        """
        Valide une ligne du tableau des factures et met à jour sa couleur en fonction de la correspondance
        avec la base de données.
        
        Args:
            row (int): Index de la ligne à valider
        """
        try:
            # Vérifier si la ligne existe
            if row < 0 or row >= self.invoice_table.rowCount():
                logger.warning(f"Tentative de validation d'une ligne invalide: {row}")
                return
            
            # Récupérer les informations de la facture
            uh_item = self.invoice_table.item(row, 0)
            invoice_num_item = self.invoice_table.item(row, 1)
            invoice_name_item = self.invoice_table.item(row, 2)
            invoice_address_item = self.invoice_table.item(row, 3)
            db_name_item = self.invoice_table.item(row, 4)
            db_line_item = self.invoice_table.item(row, 7)  # Ligne BDD
            
            if not uh_item or not invoice_num_item or not invoice_name_item or not invoice_address_item:
                logger.warning(f"Données de facture incomplètes pour la ligne {row}")
                return
            
            uh = uh_item.text()
            invoice_num = invoice_num_item.text()
            invoice_name = invoice_name_item.text()
            invoice_address = invoice_address_item.text()
            
            # Vérifier si l'utilisateur a saisi un numéro de ligne BDD
            db_line = None
            if db_line_item and db_line_item.text().strip():
                try:
                    # Convertir en entier et soustraire 1 car l'interface commence à 1 mais les indices à 0
                    db_line = int(db_line_item.text().strip()) - 1
                except ValueError:
                    logger.warning(f"Numéro de ligne BDD invalide: {db_line_item.text()}")
            
            # Si un numéro de ligne BDD valide a été saisi, rechercher la correspondance dans la base de données
            match_found = False
            exact_match = False
            partial_match = False
            db_name = ""
            client_code = ""
            chorus_code = ""
            
            if hasattr(self, 'database') and self.database.data and db_line is not None:
                # Convertir le dictionnaire en liste pour accéder par index
                db_items = list(self.database.data.items())
                
                # Vérifier si l'index est valide
                if 0 <= db_line < len(db_items):
                    name, data = db_items[db_line]
                    db_name = name
                    
                    # Mettre à jour les cellules avec les données de la base
                    self.invoice_table.setItem(row, 4, QTableWidgetItem(db_name))  # Nom BDD
                    
                    if 'client_code' in data and data['client_code']:
                        client_code = str(data['client_code'])
                        self.invoice_table.setItem(row, 5, QTableWidgetItem(client_code))  # Code client
                    
                    if 'chorus_code' in data and data['chorus_code']:
                        chorus_code = str(data['chorus_code'])
                        self.invoice_table.setItem(row, 6, QTableWidgetItem(chorus_code))  # Code chorus
                    
                    # Déterminer le type de correspondance
                    match_found = True
                    if db_name.lower() == invoice_name.lower():
                        exact_match = True
                    else:
                        partial_match = True
                    
                    logger.info(f"Correspondance trouvée à la ligne {db_line+1} de la base de données: {db_name}")
                else:
                    logger.warning(f"Ligne BDD hors limites: {db_line+1}. La base contient {len(db_items)} entrées.")
                    QMessageBox.warning(
                        self,
                        "Ligne BDD invalide",
                        f"La ligne {db_line+1} n'existe pas dans la base de données qui contient {len(db_items)} entrées."
                    )
                    return
            
            # Toujours appliquer la couleur bleue pour la validation manuelle, quelle que soit la couleur précédente
            # Cela permet au bouton Validé de fonctionner même après l'utilisation du bouton Traitement
            color = QColor("#0078d7")  # Bleu
            
            # Appliquer la couleur à toutes les cellules de la ligne
            for col in range(self.invoice_table.columnCount()):
                item = self.invoice_table.item(row, col)
                if item:
                    item.setBackground(color)
            
            # Mettre à jour le statut dans la colonne Statut (8)
            statut = "Validé"
            if match_found:
                if exact_match:
                    statut = "Validé (Exact)"
                else:
                    statut = "Validé (Partiel)"
            
            self.invoice_table.setItem(row, 8, QTableWidgetItem(statut))
            
            # Mettre à jour les statistiques
            self.update_statistics()
            
            # Journaliser l'action
            if match_found:
                logger.info(f"Ligne {row} validée avec succès. Correspondance avec {db_name} (ligne BDD: {db_line+1})")
            else:
                logger.info(f"Ligne {row} validée manuellement sans correspondance en base de données.")
            
        except Exception as e:
            logger.error(f"Erreur lors de la validation de la ligne {row}: {str(e)}", exc_info=True)
            QMessageBox.critical(self, "Erreur", f"Une erreur est survenue lors de la validation: {str(e)}")
    
    def update_statistics(self):
        """
        Met à jour les barres de progression des statistiques en fonction des couleurs des lignes
        dans le tableau des factures.
        - Bleu = Validé
        - Vert = Concordance exacte
        - Orange = Concordance partielle
        - Sans couleur = Pas de concordance
        """
        # Compter le nombre total de lignes
        total_rows = self.invoice_table.rowCount()
        if total_rows == 0:
            # Si aucune ligne, mettre toutes les barres à zéro
            self.blue_progress.setValue(0)
            self.green_progress.setValue(0)
            self.orange_progress.setValue(0)
            self.no_color_progress.setValue(0)
            return
        
        # Initialiser les compteurs
        blue_count = 0
        green_count = 0
        orange_count = 0
        no_color_count = 0
        
        # Parcourir toutes les lignes du tableau
        for row in range(total_rows):
            # Vérifier la couleur de fond de la première cellule de chaque ligne
            item = self.invoice_table.item(row, 0)
            if item is None:
                no_color_count += 1
                continue
                
            background_color = item.background().color()
            
            # Comparer la couleur avec les couleurs attendues
            if background_color.name() == "#0078d7":  # Bleu
                blue_count += 1
            elif background_color.name() == "#2ecc71":  # Vert
                green_count += 1
            elif background_color.name() == "#f39c12":  # Orange
                orange_count += 1
            else:  # Sans couleur ou autre
                no_color_count += 1
        
        # Calculer les pourcentages
        blue_percent = (blue_count / total_rows) * 100
        green_percent = (green_count / total_rows) * 100
        orange_percent = (orange_count / total_rows) * 100
        no_color_percent = (no_color_count / total_rows) * 100
        
        # Mettre à jour les barres de progression
        self.blue_progress.setValue(int(blue_percent))
        self.green_progress.setValue(int(green_percent))
        self.orange_progress.setValue(int(orange_percent))
        self.no_color_progress.setValue(int(no_color_percent))
        
        # Journaliser les statistiques
        logger.debug(f"Statistiques mises à jour: Bleu={blue_percent:.1f}%, Vert={green_percent:.1f}%, "
                    f"Orange={orange_percent:.1f}%, Sans couleur={no_color_percent:.1f}%")
    
    def show_about_dialog(self):
        """Affiche la boîte de dialogue À propos"""
        about_text = """
        <h2>Gestionnaire de Factures</h2>
        <p>Version 1.0.0</p>
        <p>Application de gestion de factures avec extraction automatique des données</p>
        <p> 2025 Tous droits réservés</p>
        """
        QMessageBox.about(self, "À propos", about_text)
    
    def show_documentation(self):
        """Ouvre la documentation de l'application"""
        QMessageBox.information(
            self,
            "Documentation",
            "La documentation sera disponible dans une future version.",
            QMessageBox.StandardButton.Ok
        )
    
    def setup_menu(self):
        """Configure la barre de menu de l'application"""
        menubar = self.menuBar()
        
        # Menu Fichier
        file_menu = menubar.addMenu("&Fichier")
        
        # Actions Fichier
        new_action = QAction("&Nouvelle facture", self)
        new_action.setShortcut("Ctrl+N")
        new_action.triggered.connect(self.new_invoice)
        
        open_action = QAction("&Ouvrir une facture", self)
        open_action.setShortcut("Ctrl+O")
        open_action.triggered.connect(self.load_invoice_file)
        
        save_action = QAction("&Enregistrer", self)
        save_action.setShortcut("Ctrl+S")
        save_action.triggered.connect(self.save_invoice_file)
        
        save_as_action = QAction("Enregistrer &sous...", self)
        save_as_action.setShortcut("Ctrl+Shift+S")
        save_as_action.triggered.connect(self.save_invoice_as)
        
        export_pdf_action = QAction("Exporter en &PDF", self)
        export_pdf_action.triggered.connect(self.export_to_pdf)
        
        exit_action = QAction("&Quitter", self)
        exit_action.setShortcut("Alt+F4")
        exit_action.triggered.connect(self.close)
        
        # Ajouter les actions au menu Fichier
        file_menu.addAction(new_action)
        file_menu.addAction(open_action)
        file_menu.addSeparator()
        file_menu.addAction(save_action)
        file_menu.addAction(save_as_action)
        file_menu.addSeparator()
        file_menu.addAction(export_pdf_action)
        file_menu.addSeparator()
        file_menu.addAction(exit_action)
        
        # Menu Base de données
        db_menu = menubar.addMenu("&Base de données")
        
        # Actions Base de données
        load_db_action = QAction("&Importer une base de données...", self)
        load_db_action.triggered.connect(self.import_database)
        
        export_db_action = QAction("&Exporter la base de données...", self)
        export_db_action.triggered.connect(self.export_database_to_excel)
        
        clear_db_action = QAction("&Vider la base de données", self)
        clear_db_action.triggered.connect(self.clear_database)
        
        # Ajouter les actions au menu Base de données
        db_menu.addAction(load_db_action)
        db_menu.addAction(export_db_action)
        db_menu.addSeparator()
        db_menu.addAction(clear_db_action)
        
        # Menu Affichage
        view_menu = menubar.addMenu("&Affichage")
        
        # Actions Affichage
        toggle_db_action = QAction("Afficher/Masquer la base de données", self, checkable=True, checked=True)
        toggle_db_action.triggered.connect(lambda checked: self.database_dock.setVisible(checked))
        
        toggle_invoice_action = QAction("Afficher/Masquer les factures", self, checkable=True, checked=True)
        toggle_invoice_action.triggered.connect(lambda checked: self.invoice_dock.setVisible(checked))
        
        fullscreen_action = QAction("Plein écran", self, checkable=True)
        fullscreen_action.setShortcut("F11")
        fullscreen_action.triggered.connect(self.toggle_fullscreen)
        
        # Ajouter les actions au menu Affichage
        view_menu.addAction(toggle_db_action)
        view_menu.addAction(toggle_invoice_action)
        view_menu.addSeparator()
        view_menu.addAction(fullscreen_action)
        
        # Menu Aide
        help_menu = menubar.addMenu("&Aide")
        
        # Actions Aide
        about_action = QAction("À &propos...", self)
        about_action.triggered.connect(self.show_about_dialog)
        
        docs_action = QAction("&Documentation...", self)
        docs_action.triggered.connect(self.show_documentation)
        
        # Ajouter les actions au menu Aide
        help_menu.addAction(about_action)
        help_menu.addAction(docs_action)
        
    def toggle_fullscreen(self, checked):
        """Bascule le mode plein écran"""
        if checked:
            self.showFullScreen()
        else:
            self.showNormal()
    
    def on_dock_visibility_changed(self, dock_widget, visible):
        """Gère les changements de visibilité des fenêtres dockées"""
        try:
            print(f"\n=== on_dock_visibility_changed: {dock_widget.windowTitle()} - Visible: {visible}")
            
            # Vérifier si le dock est toujours valide
            if sip.isdeleted(dock_widget):
                print("Le dock n'existe plus, sortie")
                return
                
            # Ne rien faire si le dock n'est pas visible ou n'est pas flottant
            if not visible or not dock_widget.isFloating():
                print("Dock non visible ou non flottant, sortie")
                dock_widget.blockSignals(False)
                return
                
            # Vérifier si la fenêtre flottante est déjà configurée
            if hasattr(dock_widget, '_floating_window'):
                print("Fenêtre flottante déjà configurée, sortie")
                return
                
            print("Configuration de la fenêtre flottante en cours...")
            
            # Bloquer temporairement les signaux pour éviter les boucles
            dock_widget.blockSignals(True)
            
            # Utiliser un timer pour différer la configuration avec un délai plus long
            QTimer.singleShot(100, lambda: self._delayed_float_config(dock_widget))
            
        except Exception as e:
            logging.error(f"Erreur dans on_dock_visibility_changed: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def _delayed_float_config(self, dock_widget):
        """Configure la fenêtre flottante après un délai"""
        try:
            print(f"\n=== _delayed_float_config: {dock_widget.windowTitle()}")
            
            # Vérifier si le dock est toujours valide
            if sip.isdeleted(dock_widget):
                print("Le dock n'existe plus, sortie")
                return
                
            # Vérifier à nouveau les conditions
            if not dock_widget.isVisible() or not dock_widget.isFloating():
                print("Dock non visible ou non flottant, sortie")
                dock_widget.blockSignals(False)
                return
            
            # Récupérer la fenêtre flottante (top-level window)
            floating_window = dock_widget.window()
            if floating_window == self:
                print("La fenêtre flottante est la fenêtre principale, sortie")
                return
                
            print("1. Configuration de la fenêtre flottante...")
            
            # Stocker la référence à la fenêtre flottante et au dock
            dock_widget._floating_window = floating_window
            floating_window._dock_widget = dock_widget
            
            # Activer le drag & drop
            floating_window.setAcceptDrops(True)
            floating_window.setMouseTracking(True)
            
            # Variables pour le drag & drop
            floating_window.mouse_pos = None
            floating_window.dragging = False
            floating_window.drop_indicator = None
            
            # Créer un indicateur de zone de dépôt
            def create_drop_indicator():
                indicator = QWidget(floating_window)
                indicator.setStyleSheet("background-color: rgba(100, 180, 255, 100);")
                indicator.hide()
                return indicator
                
            floating_window.drop_indicator = create_drop_indicator()
            
            # Surcharger les événements de souris pour le drag
            def mousePressEvent(event):
                pass
                
            # Nous n'utilisons plus cette fonction car elle est maintenant gérée dans la classe FloatingWindow
            def mouseMoveEvent(event):
                # Nous gardons uniquement la logique de gestion des zones de dépôt
                if hasattr(floating_window, 'dragging') and floating_window.dragging and hasattr(floating_window, 'mouse_pos') and floating_window.mouse_pos is not None:
                    # Vérifier si on est au-dessus de la fenêtre principale
                    main_window_pos = self.mapFromGlobal(floating_window.pos())
                    if self.rect().contains(main_window_pos):
                        # Afficher les zones de dépôt avec surbrillance
                        self.show_drop_zones(floating_window, event.globalPos())
                    else:
                        self.hide_drop_zones()
                    
                    # Mettre à jour la surbrillance en fonction de la position de la souris
                    self.update_drop_highlight(event.globalPos())
                
            # Nous modifions cette fonction pour qu'elle fonctionne avec notre nouvelle implémentation
            def mouseReleaseEvent(event):
                if hasattr(floating_window, 'dragging') and floating_window.dragging:
                    # La gestion du relâchement du bouton est maintenant dans FloatingWindow
                    # Nous gardons uniquement la logique de gestion des zones de dépôt
                    self.hide_drop_zones()
                    
                    # Vérifier si on relâche au-dessus de la fenêtre principale
                    main_window_pos = self.mapFromGlobal(floating_window.pos())
                    if self.rect().contains(main_window_pos):
                        # Déterminer la zone de dépôt
                        drop_area = None
                        local_pos = self.mapFromGlobal(event.globalPos())
                        
                        if hasattr(self, 'drop_zones'):
                            for area, zone in self.drop_zones.items():
                                if zone.geometry().contains(local_pos):
                                    drop_area = area
                                    break
                        
                        # Utiliser la zone de dépôt ou la zone par défaut
                        target_area = drop_area if drop_area is not None else Qt.DockWidgetArea.LeftDockWidgetArea
                        
                        # Réinsérer le dock dans la fenêtre principale
                        floating_window.hide()
                        dock_widget.setFloating(False)
                        self.addDockWidget(target_area, dock_widget)
                        
                        # Nettoyer les références
                        if hasattr(dock_widget, '_floating_window'):
                            delattr(dock_widget, '_floating_window')
                        if hasattr(floating_window, '_dock_widget'):
                            delattr(floating_window, '_dock_widget')
            
            # Ne pas remplacer les gestionnaires d'événements de souris car ils sont maintenant gérés dans la classe FloatingWindow
            # Nous utilisons les méthodes natives de la classe FloatingWindow
            
            # Remplacer le gestionnaire de fermeture
            def close_event(event):
                print("\n=== Bouton de fermeture cliqué ===")
                # Empêcher la fermeture réelle
                if event is not None:
                    event.ignore()
                
                # Cacher la fenêtre flottante
                floating_window.hide()
                
                # Réinsérer le widget dans le dock
                dock_widget.setFloating(False)
                
                # Réactiver les fonctionnalités du dock
                dock_widget.setFeatures(
                    QDockWidget.DockWidgetFeature.DockWidgetMovable |
                    QDockWidget.DockWidgetFeature.DockWidgetFloatable |
                    QDockWidget.DockWidgetFeature.DockWidgetClosable
                )
                
                # S'assurer que le dock est visible
                dock_widget.show()
                
                # Nettoyer les références
                if hasattr(dock_widget, '_floating_window'):
                    delattr(dock_widget, '_floating_window')
                if hasattr(floating_window, '_dock_widget'):
                    delattr(floating_window, '_dock_widget')
                
                print("Fenêtre flottante réintégrée avec succès")
            
            # Remplacer le gestionnaire de fermeture
            floating_window.closeEvent = close_event
            
            # Remplacer également l'action du bouton de fermeture
            close_button = None
            # Chercher le bouton de fermeture parmi tous les enfants
            for child in floating_window.findChildren(QPushButton):
                if "close" in child.objectName().lower() or "fermer" in child.text().lower():
                    close_button = child
                    break
                    
            if close_button:
                try:
                    close_button.clicked.disconnect()
                except:
                    pass
                close_button.clicked.connect(lambda: close_event(None))
            
            print("2. Configuration du bouton de fermeture personnalisé")
            
            # Réactiver les signaux
            dock_widget.blockSignals(False)
            print("3. Configuration terminée avec succès")
            
        except Exception as e:
            print(f"\n=== ERREUR DANS _delayed_float_config ===")
            print(f"Type d'erreur: {type(e).__name__}")
            print(f"Message d'erreur: {str(e)}")
            print("\nTraceback complet:")
            import traceback
            traceback.print_exc()
            print("=== FIN ERREUR ===\n")
            
            # Réactiver les signaux en cas d'erreur
            if not sip.isdeleted(dock_widget):
                dock_widget.blockSignals(False)
            
    def show_drop_zones(self, floating_window, mouse_pos):
        """Affiche les zones de dépôt potentielles avec une surbrillance améliorée"""
        if not hasattr(self, 'drop_zones'):
            self.drop_zones = {}
            
            # Définir les zones avec leurs libellés
            zones = [
                (Qt.DockWidgetArea.LeftDockWidgetArea, "Gauche"),
                (Qt.DockWidgetArea.RightDockWidgetArea, "Droite"),
                (Qt.DockWidgetArea.TopDockWidgetArea, "Haut"),
                (Qt.DockWidgetArea.BottomDockWidgetArea, "Bas")
            ]
            
            # Créer des indicateurs pour chaque zone de dépôt avec un style amélioré
            for area, label_text in zones:
                indicator = QWidget(self)
                indicator.setStyleSheet(f"""
                    background-color: {CYBERPUNK_COLORS['drop_zone']};
                    border: 3px solid rgba(255, 255, 255, 150);
                    border-radius: 8px;
                """)
                
                # Ajouter un label pour le nom de la zone avec un style amélioré
                label = QLabel(label_text, indicator)
                label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                label.setStyleSheet(f"""
                    color: white;
                    font-weight: bold;
                    font-size: 18px;
                    background: rgba(0, 0, 0, 180);
                    border-radius: 6px;
                    padding: 10px 16px;
                    border: 1px solid {CYBERPUNK_COLORS['accent1']};
                """)
                
                # Positionner le label au centre
                layout = QVBoxLayout(indicator)
                layout.addWidget(label, 0, Qt.AlignmentFlag.AlignCenter)
                layout.setContentsMargins(10, 10, 10, 10)
                
                indicator.hide()
                self.drop_zones[area] = indicator
                
                # Stocker la zone de dépôt
                indicator.drop_area = area
        
        # Calculer les dimensions avec des marges réduites pour une meilleure visibilité
        rect = self.rect()
        margin = 8
        
        # Taille des zones (plus grandes et plus visibles)
        side_width = max(220, int(rect.width() * 0.28))  # 28% de la largeur
        side_height = max(180, int(rect.height() * 0.38))  # 38% de la hauteur
        
        # Définir les zones avec des tailles plus généreuses et des positions optimisées
        zones = {
            # Gauche - prend toute la hauteur
            Qt.DockWidgetArea.LeftDockWidgetArea: QRect(
                rect.left() + margin, 
                rect.top() + margin, 
                side_width, 
                rect.height() - 2 * margin
            ),
            # Droite - prend toute la hauteur
            Qt.DockWidgetArea.RightDockWidgetArea: QRect(
                rect.right() - side_width - margin, 
                rect.top() + margin, 
                side_width, 
                rect.height() - 2 * margin
            ),
            # Haut - occupe tout l'espace en haut
            Qt.DockWidgetArea.TopDockWidgetArea: QRect(
                rect.left() + margin, 
                rect.top() + margin, 
                rect.width() - 2 * margin, 
                side_height
            ),
            # Bas - occupe tout l'espace en bas
            Qt.DockWidgetArea.BottomDockWidgetArea: QRect(
                rect.left() + margin, 
                rect.bottom() - side_height - margin, 
                rect.width() - 2 * margin, 
                side_height
            )
        }
        
        # Mettre à jour la géométrie des zones et les afficher
        for area, area_rect in zones.items():
            self.drop_zones[area].setGeometry(area_rect)
            self.drop_zones[area].show()
            self.drop_zones[area].raise_()
        
        # Mettre à jour la surbrillance initiale
        self.update_drop_highlight(mouse_pos)
    
    def update_drop_highlight(self, mouse_pos):
        """Met à jour la surbrillance de la zone survolée avec des effets visuels améliorés"""
        if not hasattr(self, 'drop_zones'):
            return
            
        # Convertir la position en coordonnées locales
        local_pos = self.mapFromGlobal(mouse_pos)
        
        # Vérifier quelle zone est survolée
        hovered_area = None
        for area, zone in self.drop_zones.items():
            if zone.geometry().contains(local_pos):
                hovered_area = area
                break
        
        # Mettre à jour le style des zones avec des animations plus visibles
        for area, zone in self.drop_zones.items():
            if area == hovered_area:
                # Style pour la zone survolée - plus vif et plus visible
                zone.setStyleSheet(f"""
                    background-color: {CYBERPUNK_COLORS['drop_zone_highlight']};
                    border: 4px solid {CYBERPUNK_COLORS['accent1']};
                    border-radius: 10px;
                    box-shadow: 0 0 15px rgba(255, 77, 77, 0.7);
                """)
                
                # Mettre en évidence le label avec un style plus accrocheur
                label = zone.findChild(QLabel)
                if label:
                    label.setStyleSheet(f"""
                        color: white;
                        font-weight: bold;
                        font-size: 20px;
                        background: {CYBERPUNK_COLORS['accent1']};
                        border-radius: 8px;
                        padding: 12px 20px;
                        box-shadow: 0 0 10px rgba(0, 0, 0, 0.5);
                    """)
                    
                # Mettre la zone en avant-plan
                zone.raise_()
            else:
                # Style pour les zones non survolées - plus discrètes mais toujours visibles
                zone.setStyleSheet(f"""
                    background-color: {CYBERPUNK_COLORS['drop_zone']};
                    border: 2px solid rgba(255, 255, 255, 180);
                    border-radius: 8px;
                """)
                
                # Style normal pour les labels non survolés
                label = zone.findChild(QLabel)
                if label:
                    label.setStyleSheet(f"""
                        color: white;
                        font-weight: bold;
                        font-size: 16px;
                        background: rgba(0, 0, 0, 150);
                        border-radius: 6px;
                        padding: 8px 14px;
                        border: 1px solid {CYBERPUNK_COLORS['accent1']};
                    """)
    
    def hide_drop_zones(self):
        """Cache toutes les zones de dépôt"""
        if hasattr(self, 'drop_zones'):
            for zone in self.drop_zones.values():
                zone.hide()
    
    def on_floating_window_close(self, event, dock_widget):
        """Ancienne méthode de gestion de la fermeture - conservée pour compatibilité"""
        pass
    
    def _safe_close_floating_window(self, window):
        """Ferme de manière sécurisée une fenêtre flottante"""
        try:
            if not sip.isdeleted(window):
                window.close()
                window.deleteLater()
        except Exception as e:
            print(f"Erreur lors de la fermeture sécurisée de la fenêtre: {str(e)}")
            
            print("Fenêtre rattachée avec succès")
            
        except Exception as e:
            print(f"Erreur lors du rattachement: {str(e)}")
            import traceback
            traceback.print_exc()
            logging.error(traceback.format_exc())
    
    def closeEvent(self, event):
        """Gère la fermeture de l'application"""
        try:
            # Vérifier si la méthode save_state accepte le paramètre show_message
            import inspect
            sig = inspect.signature(self.save_state)
            if 'show_message' in sig.parameters:
                self.save_state(show_message=False)
            else:
                self.save_state()
            
            # Fermer toutes les fenêtres flottantes
            for dock in [self.database_dock, self.invoice_dock]:
                if dock.isFloating():
                    dock.close()
        except Exception as e:
            logger.error(f"Erreur lors de la fermeture de l'application: {str(e)}")
        
        event.accept()
    
    def restore_dock_widgets(self):
        """Restaure les widgets dock et force leur affichage"""
    
    def configure_floating_window(self, dock_widget, floating):
        """Configure les fenêtres détachées avec une vraie fenêtre indépendante et tous les boutons système."""
        print("\n=== DÉBUT configure_floating_window ===")
        print(f"Configuration de la fenêtre détachée: {dock_widget.windowTitle()}")
        print(f"Floating: {floating}")

        # Vérifier si le dock existe encore
        if sip.isdeleted(dock_widget):
            print("Le dock n'existe plus, sortie")
            return

        # Cas 1: Création d'une fenêtre flottante
        if floating:
            print("1. Configuration d'une nouvelle fenêtre flottante...")
            # Cacher temporairement le dock pendant la création de la fenêtre flottante
            dock_widget.hide()
            
            # Vérifier si une fenêtre flottante existe déjà
            if hasattr(dock_widget, '_floating_window') and dock_widget._floating_window is not None:
                print("Une fenêtre flottante existe déjà, sortie")
                return
                
            # Créer la fenêtre flottante
            try:
                # Vérifier si le dock a un widget
                if dock_widget.widget() is None:
                    print("Le dock n'a pas de widget, impossible de créer une fenêtre flottante")
                    dock_widget.show()
                    return
                
                print(f"Widget du dock avant création de la fenêtre flottante: {dock_widget.widget()}")
                
                # Créer la fenêtre flottante avec le widget du dock
                floating_window = FloatingWindow(dock_widget)
                floating_window.setMinimumSize(800, 600)
                
                # Positionnement intelligent
                screen_geometry = QApplication.primaryScreen().availableGeometry()
                main_window_rect = self.geometry()
                window_width = min(int(screen_geometry.width() * 0.4), 1200)
                window_height = min(int(screen_geometry.height() * 0.8), 900)
                x = main_window_rect.right() + 10
                y = main_window_rect.top()
                if x + window_width > screen_geometry.right():
                    x = max(0, screen_geometry.right() - window_width - 10)
                if y + window_height > screen_geometry.bottom():
                    y = max(0, screen_geometry.bottom() - window_height - 10)
                floating_window.setGeometry(x, y, window_width, window_height)
                
                # Définir les flags de fenêtre appropriés
                flags = Qt.WindowFlags()
                flags |= Qt.Window
                flags |= Qt.WindowMinimizeButtonHint
                flags |= Qt.WindowMaximizeButtonHint
                flags |= Qt.WindowCloseButtonHint
                flags |= Qt.WindowSystemMenuHint
                flags |= Qt.WindowTitleHint
                
                # Appliquer les flags
                floating_window.setWindowFlags(flags)
                
                # Stocker la référence à la fenêtre flottante
                dock_widget._floating_window = floating_window
                
                # Connecter le signal destroyed à notre méthode de gestion
                floating_window.destroyed.connect(
                    lambda: self.on_floating_window_closed(dock_widget))
                    
                # Afficher la fenêtre flottante
                floating_window.show()
                floating_window.raise_()
                
                # Traiter les événements en attente pour éviter le blocage
                QApplication.processEvents()
                
                print("Fenêtre flottante créée et affichée avec succès")
                
            except Exception as e:
                print(f"Erreur lors de la création de la fenêtre flottante: {e}")
                # En cas d'erreur, réafficher le dock
                dock_widget.show()
            
            return
            
        # Cas 2: Retour du dock à la fenêtre principale
        if hasattr(dock_widget, '_floating_window') and dock_widget._floating_window is not None:
            # Ne pas supprimer la référence ici, cela sera fait dans on_floating_window_closed
            pass
            
        # Réafficher le dock
        dock_widget.show()
        
    def on_floating_window_closed(self, dock_widget):
        """Gère la fermeture d'une fenêtre flottante."""
        try:
            print(f"Fermeture de la fenêtre flottante pour {dock_widget.windowTitle()}")
            
            # Vérifier si le dock existe encore
            if sip.isdeleted(dock_widget):
                print("Le dock n'existe plus, impossible de le restaurer")
                return
                
            # Nettoyer la référence à la fenêtre flottante
            if hasattr(dock_widget, '_floating_window'):
                dock_widget._floating_window = None
                
            # Vérifier si le dock a un widget
            if dock_widget.widget() is None:
                print("Le dock n'a pas de widget après la fermeture de la fenêtre flottante")
                # Cela ne devrait pas arriver car le widget devrait être réattaché dans closeEvent
                # de la fenêtre flottante, mais au cas où...
            else:
                print(f"Le dock a un widget après la fermeture: {dock_widget.widget()}")
                
            # Réafficher le dock dans la fenêtre principale
            dock_widget.setFloating(False)
            dock_widget.show()
            
            # Traiter les événements en attente pour éviter le blocage
            QApplication.processEvents()
            
            print("Dock restauré avec succès")
            
        except Exception as e:
            print(f"Erreur lors de la fermeture de la fenêtre flottante: {e}")
            # En cas d'erreur, on essaie quand même de nettoyer
            if hasattr(dock_widget, '_floating_window') and not sip.isdeleted(dock_widget):
                dock_widget._floating_window = None

    def new_invoice(self):
        """Crée une nouvelle facture vide"""
        # Demander confirmation si des modifications non enregistrées
        if hasattr(self, 'current_invoice_path') and hasattr(self, 'is_modified') and self.is_modified():
            reply = QMessageBox.question(
                self, 
                'Nouvelle facture', 
                'Voulez-vous enregistrer les modifications apportées à la facture actuelle ?',
                QMessageBox.StandardButton.Save | QMessageBox.StandardButton.Discard | QMessageBox.StandardButton.Cancel
            )
            
            if reply == QMessageBox.StandardButton.Save:
                self.save_invoice_file()
            elif reply == QMessageBox.StandardButton.Cancel:
                return
        
        # Réinitialiser l'interface pour une nouvelle facture
        self.current_invoice_path = None
        self.setWindowTitle("Nouvelle facture - Gestionnaire de Factures")
        
        # Réinitialiser les champs du formulaire
        if hasattr(self, 'invoice_table'):
            self.invoice_table.setRowCount(0)
            
        # Réinitialiser les statistiques
        if hasattr(self, 'invoice_stats'):
            self.invoice_stats.update_stats({})
            
        self.statusBar().showMessage("Nouvelle facture créée", 3000)
        
    def save_invoice_as(self):
        """Traitement des factures avec recherche de concordances dans la base de données"""
        try:
            
            # S'assurer que la base de données est chargée
            if not hasattr(self, 'database') or not self.database.data or len(self.database.data) == 0:
                QMessageBox.warning(
                    self,
                    "Base de données non chargée",
                    "La base de données n'est pas encore chargée ou est vide. Veuillez patienter ou importer des données."
                )
                return False
            
            # Créer une boîte de dialogue de progression
            progress_dialog = QProgressDialog("Recherche de concordances dans la base de données...", "Annuler", 0, self.invoice_table.rowCount(), self)
            progress_dialog.setWindowTitle("Traitement en cours")
            progress_dialog.setWindowModality(Qt.WindowModality.WindowModal)
            progress_dialog.setMinimumDuration(0)
            progress_dialog.show()
            
            # Compteurs pour le rapport
            concordances_parfaites = 0
            concordances_partielles = 0
            sans_concordance = 0
            
            # Traiter chaque ligne du tableau des factures
            for row in range(self.invoice_table.rowCount()):
                # Vérifier si l'utilisateur a annulé
                if progress_dialog.wasCanceled():
                    break
                    
                # Mettre à jour la progression
                progress_dialog.setValue(row)
                progress_dialog.setLabelText(f"Traitement de la ligne {row+1}/{self.invoice_table.rowCount()}...")
                
                # Récupérer les informations de la ligne
                uh_item = self.invoice_table.item(row, 0)
                facture_num_item = self.invoice_table.item(row, 1)
                nom_facture_item = self.invoice_table.item(row, 2)
                adresse_facture_item = self.invoice_table.item(row, 3)
                
                if not uh_item or not facture_num_item or not nom_facture_item:
                    continue  # Ignorer les lignes incomplètes
                
                uh = uh_item.text().strip()
                facture_num = facture_num_item.text().strip()
                nom_facture = nom_facture_item.text().strip()
                adresse_facture = adresse_facture_item.text().strip() if adresse_facture_item else ""
                
                # Nettoyer le numéro de facture pour la recherche
                facture_num_clean = facture_num.lower()
                for prefix in ["facture n°", "facture n", "facture"]:
                    if facture_num_clean.startswith(prefix):
                        facture_num_clean = facture_num_clean.replace(prefix, "").strip()
                        break
                
                # Rechercher dans la base de données
                meilleure_correspondance = None
                meilleur_score = 0
                correspondance_exacte = False
                
                # Convertir le dictionnaire en liste pour itération
                db_items = list(self.database.data.items())
                
                for idx, (name, data) in enumerate(db_items):
                    # Calculer un score de similarité basé sur le nom et l'adresse
                    score = 0
                    
                    # Vérifier si le nom de la facture correspond exactement au nom dans la base de données
                    if nom_facture.lower() == name.lower():
                        score += 100  # Score élevé pour une correspondance exacte du nom
                        correspondance_exacte = True
                    elif nom_facture.lower() in name.lower() or name.lower() in nom_facture.lower():
                        # Score partiel pour une correspondance partielle
                        score += 50
                    
                    # Vérifier l'adresse si elle est disponible
                    if adresse_facture and 'address' in data and data['address']:
                        if adresse_facture.lower() == data['address'].lower():
                            score += 50  # Score pour une correspondance exacte de l'adresse
                        elif adresse_facture.lower() in data['address'].lower() or data['address'].lower() in adresse_facture.lower():
                            score += 25  # Score partiel pour une correspondance partielle de l'adresse
                    
                    # Si le score est meilleur que le précédent, mettre à jour la meilleure correspondance
                    if score > meilleur_score:
                        meilleur_score = score
                        meilleure_correspondance = (idx, name, data)
                
                # Traiter la meilleure correspondance trouvée
                if meilleure_correspondance:
                    idx, name, data = meilleure_correspondance
                    
                    # Mettre à jour les cellules du tableau
                    self.invoice_table.setItem(row, 4, QTableWidgetItem(name))  # Nom BDD
                    
                    if 'client_code' in data and data['client_code']:
                        self.invoice_table.setItem(row, 5, QTableWidgetItem(str(data['client_code'])))  # Code client
                    
                    if 'chorus_code' in data and data['chorus_code']:
                        self.invoice_table.setItem(row, 6, QTableWidgetItem(str(data['chorus_code'])))  # Code chorus
                    
                    # Mettre à jour la ligne BDD (index + 1 car l'interface commence à 1)
                    self.invoice_table.setItem(row, 7, QTableWidgetItem(str(idx + 1)))  # Ligne BDD
                    
                    # Appliquer la couleur selon le type de correspondance
                    couleur = None
                    if correspondance_exacte and meilleur_score >= 100:
                        # Concordance parfaite - vert
                        couleur = QColor(CYBERPUNK_COLORS['success'])  # Vert
                        concordances_parfaites += 1
                        statut = "Parfaite"
                    elif meilleur_score >= 50:
                        # Concordance partielle - orange
                        couleur = QColor(CYBERPUNK_COLORS['warning'])  # Orange
                        concordances_partielles += 1
                        statut = "Partielle"
                    else:
                        # Pas de concordance suffisante
                        sans_concordance += 1
                        statut = "Aucune"
                        continue  # Passer à la ligne suivante sans changer la couleur
                    
                    # Appliquer la couleur à toutes les cellules de la ligne
                    for col in range(self.invoice_table.columnCount()):
                        item = self.invoice_table.item(row, col)
                        if item:
                            item.setBackground(couleur)
                    
                    # Mettre à jour le statut
                    self.invoice_table.setItem(row, 8, QTableWidgetItem(statut))
                    
                    logger.info(f"Ligne {row+1}: Concordance {statut} trouvée avec {name} (score: {meilleur_score})")
                else:
                    # Aucune correspondance trouvée
                    sans_concordance += 1
                    logger.info(f"Ligne {row+1}: Aucune concordance trouvée pour {nom_facture}")
            
            # Fermer la boîte de dialogue de progression
            progress_dialog.setValue(self.invoice_table.rowCount())
            
            # Afficher un résumé des résultats
            QMessageBox.information(
                self,
                "Traitement terminé",
                f"Résultats du traitement:\n\n"
                f"- Concordances parfaites: {concordances_parfaites}\n"
                f"- Concordances partielles: {concordances_partielles}\n"
                f"- Sans concordance: {sans_concordance}"
            )
            
            # Mettre à jour la barre de statut
            self.statusBar().showMessage(f"Traitement terminé: {concordances_parfaites} parfaites, {concordances_partielles} partielles, {sans_concordance} sans concordance", 5000)
            
            return True
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Erreur lors du traitement",
                f"Une erreur est survenue lors du traitement des factures:\n{str(e)}"
            )
            logger.error(f"Erreur lors du traitement des factures: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            self.statusBar().showMessage("Erreur lors du traitement", 5000)
            return False
            
    def export_to_pdf(self):
        """Exporte la facture actuelle en PDF"""
        if not hasattr(self, 'invoice_table') or self.invoice_table.rowCount() == 0:
            QMessageBox.warning(self, "Export PDF", "Aucune donnée de facture à exporter.")
            return
            
        # Demander où enregistrer le PDF
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Exporter en PDF",
            "",
            "Fichiers PDF (*.pdf)"
        )
        
        if not file_path:
            return
            
        # Ajouter l'extension si elle n'est pas présente
        if not file_path.endswith('.pdf'):
            file_path += '.pdf'
            
        try:
            # Ici, vous devrez implémenter la logique d'export PDF
            # Ceci est un exemple basique
            from PyQt5.QtPrintSupport import QPrinter, QPrintDialog
            
            printer = QPrinter(QPrinter.HighResolution)
            printer.setOutputFormat(QPrinter.PdfFormat)
            printer.setOutputFileName(file_path)
            
            # Afficher la boîte de dialogue d'impression
            print_dialog = QPrintDialog(printer, self)
            if print_dialog.exec_() == QPrintDialog.Accepted:
                # Ici, vous devrez implémenter la logique pour dessiner le PDF
                # en fonction du contenu de votre facture
                self.statusBar().showMessage(f"Facture exportée en PDF : {file_path}", 5000)
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Impossible d'exporter en PDF : {str(e)}")
    
    def is_modified(self):
        """Vérifie si la facture a été modifiée"""
        # Implémentez la logique pour vérifier si la facture a été modifiée
        # Par exemple, en comparant avec la dernière version enregistrée
        return False  # À implémenter
        
    def setup_invoice_tab(self):
        """Configure l'onglet Gestion des factures"""
        # Création du widget principal
        invoice_tab = QWidget()
        layout = QVBoxLayout(invoice_tab)
        
        # Barre d'outils supérieure
        toolbar = QHBoxLayout()
        
        # Bouton Charger une facture
        self.load_invoice_btn = CustomButton("Charger une facture")
        self.load_invoice_btn.clicked.connect(self.load_invoice_file)
        toolbar.addWidget(self.load_invoice_btn)
        
        # Bouton Sauvegarder les modifications
        self.save_btn = CustomButton("Sauvegarder les modifications")
        self.save_btn.clicked.connect(self.save_invoice_file)
        toolbar.addWidget(self.save_btn)
        
        toolbar.addStretch()
        
        # Barre de recherche
        search_layout = QHBoxLayout()
        self.invoice_search_input = QLineEdit()
        self.invoice_search_input.setPlaceholderText("Rechercher dans les factures...")
        self.invoice_search_input.textChanged.connect(self.filter_invoices)
        
        search_layout.addWidget(self.invoice_search_input)
        
        # Tableau des factures
        self.invoice_table = QTableWidget()
        self.invoice_table.setColumnCount(5)
        self.invoice_table.setHorizontalHeaderLabels(["Client", "N° Facture", "Date", "Montant", "Statut"])
        self.invoice_table.setEditTriggers(QTableWidget.EditTrigger.DoubleClicked)
        self.invoice_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.invoice_table.setAlternatingRowColors(True)
        self.invoice_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        
        # Style du tableau des factures
        self.invoice_table.horizontalHeader().setStyleSheet(f"""
            QHeaderView::section {{ 
                background-color: {CYBERPUNK_COLORS['accent1']};
                color: white;
                font-weight: bold;
                padding: 6px;
                border: none;
                border-radius: 0px;
            }}
        """)
        
        self.invoice_table.setStyleSheet(f"""
            QTableWidget {{ 
                gridline-color: {CYBERPUNK_COLORS['border']};
                background-color: {CYBERPUNK_COLORS['background']};
                alternate-background-color: {CYBERPUNK_COLORS['secondary_bg']};
                selection-background-color: {CYBERPUNK_COLORS['highlight']};
                selection-color: {CYBERPUNK_COLORS['text']};
            }}
        """)
        
        # Ajouter les éléments au layout
        layout.addLayout(toolbar)
        layout.addLayout(search_layout)
        layout.addWidget(self.invoice_table)
        
        # Ajouter un groupe pour les statistiques
        stats_group = QGroupBox("Statistiques")
        stats_layout = QHBoxLayout()
        
        self.invoice_stats = StatsWidget()
        stats_layout.addWidget(self.invoice_stats)
        stats_group.setLayout(stats_layout)
        
        layout.addWidget(stats_group)
        
        # Définir le layout de l'onglet
        self.tab2.setLayout(layout)

    def setup_database_tab(self):
        """Configure l'onglet Base de données"""
        # Créer le contenu de l'onglet principal de la base de données
        layout = QVBoxLayout(self.tab1)
        
        # Barre d'outils supérieure
        toolbar = QHBoxLayout()
        
        # Boutons de gestion de la base de données
        buttons = [
            ("Charger la base de données", self.load_database_from_excel, "load_db_btn"),
            ("Exporter la base de données", self.export_database_to_excel, "export_db_btn"),
            ("Vider la base de données", self.clear_database, "clear_db_btn")
        ]
        
        for text, callback, attr_name in buttons:
            btn = CustomButton(text)
            btn.clicked.connect(callback)
            toolbar.addWidget(btn)
            setattr(self, attr_name, btn)
        
        toolbar.addStretch()
        
        # Barre de recherche
        search_layout = QHBoxLayout()
        search_label = QLabel("Recherche:")
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Rechercher dans la base de données...")
        self.search_input.textChanged.connect(self.filter_database)
        
        search_layout.addWidget(search_label)
        search_layout.addWidget(self.search_input)
        
        # Configuration du tableau principal
        self.db_table = self._create_database_table()
        
        # Ajouter les éléments au layout
        layout.addLayout(toolbar)
        layout.addLayout(search_layout)
        layout.addWidget(self.db_table)
        
        # Ajouter un groupe pour les statistiques
        stats_group = QGroupBox("Statistiques")
        stats_layout = QHBoxLayout()
        
        self.stats_widget = StatsWidget()
        stats_layout.addWidget(self.stats_widget)
        stats_group.setLayout(stats_layout)
        
        layout.addWidget(stats_group)
        
        # Configurer le widget détachable pour les factures
        self._setup_invoice_dock()
        
        # Configurer la barre d'outils et charger les données
        self.setup_toolbar()
        self.database.ensure_loaded(self.update_database_view)
    
    def _create_database_table(self):
        """Crée et configure le tableau de la base de données"""
        table = QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(["Nom", "Code Client", "Code Chorus", "Adresse"])
        table.setEditTriggers(QTableWidget.EditTrigger.DoubleClicked)
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        table.setAlternatingRowColors(True)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        
        # Style du tableau
        table.horizontalHeader().setStyleSheet(f"""
            QHeaderView::section {{ 
                background-color: {CYBERPUNK_COLORS['accent1']};
                color: white;
                font-weight: bold;
                padding: 6px;
                border: none;
                border-radius: 0px;
            }}
        """)
        
        table.setStyleSheet(f"""
            QTableWidget {{ 
                font-size: 12px;
                gridline-color: {CYBERPUNK_COLORS['border']};
                background-color: {CYBERPUNK_COLORS['background']};
                alternate-background-color: {CYBERPUNK_COLORS['secondary_bg']};
                selection-background-color: {CYBERPUNK_COLORS['highlight']};
                selection-color: {CYBERPUNK_COLORS['text']};
            }}
            QTableWidget::item {{
                padding: 4px;
            }}
        """)
        
        # Connecter le signal de modification d'élément
        table.itemChanged.connect(self.on_db_item_changed)
        
        return table
    
    def _setup_invoice_dock(self):
        """Configure le widget détachable pour les factures"""
        self.invoice_dock = QDockWidget("Factures", self)
        self.invoice_dock.setWidget(self.tab2)  # Utiliser l'onglet des factures existant
        self.invoice_dock.setAllowedAreas(Qt.DockWidgetArea.AllDockWidgetAreas)
        self.invoice_dock.setFeatures(
            QDockWidget.DockWidgetFeature.DockWidgetMovable | 
            QDockWidget.DockWidgetFeature.DockWidgetFloatable |
            QDockWidget.DockWidgetFeature.DockWidgetClosable
        )
        
        # Connecter le signal de changement d'état flottant
        self.invoice_dock.topLevelChanged.connect(
            lambda: self.configure_floating_window(self.invoice_dock, self.invoice_dock.isFloating()))
    
    def _setup_full_database_tab(self):
        """Méthode conservée pour la compatibilité, la logique a été déplacée dans setup_database_tab"""
        pass
        
    def setup_toolbar(self):
        """Configure la barre d'outils de l'application"""
        # Créer la barre d'outils
        toolbar = self.addToolBar("Outils")
        toolbar.setMovable(False)
        toolbar.setIconSize(QSize(24, 24))
        
        # Style de la barre d'outils
        toolbar.setStyleSheet(f"""
            QToolBar {{
                background-color: {CYBERPUNK_COLORS['secondary_bg']};
                border: none;
                padding: 4px;
                spacing: 4px;
            }}
            QToolButton {{
                background-color: {CYBERPUNK_COLORS['accent1']};
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 12px;
                margin: 0 2px;
            }}
            QToolButton:hover {{
                background-color: {CYBERPUNK_COLORS['accent2']};
            }}
            QToolButton:pressed {{
                background-color: {CYBERPUNK_COLORS['accent3']};
            }}
        """)
        
        # Actions principales
        load_invoice_action = QAction("Ouvrir une facture", self)
        load_invoice_action.triggered.connect(self.load_invoice_file)
        load_invoice_action.setIcon(self.style().standardIcon(getattr(QStyle.StandardPixmap, 'SP_DialogOpenButton')))
        toolbar.addAction(load_invoice_action)
        
        save_invoice_action = QAction("Enregistrer les modifications", self)
        save_invoice_action.triggered.connect(self.save_invoice_file)
        save_invoice_action.setIcon(self.style().standardIcon(getattr(QStyle.StandardPixmap, 'SP_DialogSaveButton')))
        toolbar.addAction(save_invoice_action)
        
        toolbar.addSeparator()
        
        # Bouton pour afficher/masquer le dock des factures
        toggle_dock_action = QAction("Afficher/Masquer les factures", self)
        toggle_dock_action.triggered.connect(self.toggle_invoice_dock)
        toggle_dock_action.setCheckable(True)
        toggle_dock_action.setChecked(True)
        toolbar.addAction(toggle_dock_action)
        
        # Ajouter un séparateur
        toolbar.addSeparator()
        
        # Indicateur de chargement
        toolbar.addWidget(QLabel("Statut: "))
        self.status_indicator = QLabel("Prêt")
        toolbar.addWidget(self.status_indicator)
        toolbar.addWidget(self.loading_indicator)
        
    def toggle_invoice_dock(self, visible):
        """Affiche ou masque le dock des factures"""
        if hasattr(self, 'invoice_dock'):
            self.invoice_dock.setVisible(visible)
            self.status_indicator.setText("Factures affichées" if visible else "Factures masquées")

    def update_database_view(self):
        """Met à jour l'affichage de la base de données"""
        try:
            self.database.ensure_loaded()
            
            # Désactiver temporairement les signaux
            self.db_table.blockSignals(True)
            if hasattr(self, 'full_db_table'):
                self.full_db_table.blockSignals(True)
            
            # Mettre à jour le tableau de la base de données principale
            self.db_table.setRowCount(0)
            if hasattr(self, 'full_db_table'):
                self.full_db_table.setRowCount(0)
            
            row = 0
            for key, entry in self.database.data.items():
                # Tableau principal
                self.db_table.insertRow(row)
                self.db_table.setItem(row, 0, QTableWidgetItem(entry.get('name', '')))
                self.db_table.setItem(row, 1, QTableWidgetItem(entry.get('client_code', '')))
                self.db_table.setItem(row, 2, QTableWidgetItem(entry.get('chorus_code', '')))
                self.db_table.setItem(row, 3, QTableWidgetItem(entry.get('address', '')))
                
                # Tableau complet (si initialisé)
                if hasattr(self, 'full_db_table'):
                    self.full_db_table.insertRow(row)
                    self.full_db_table.setItem(row, 0, QTableWidgetItem(entry.get('name', '')))
                    self.full_db_table.setItem(row, 1, QTableWidgetItem(entry.get('client_code', '')))
                    self.full_db_table.setItem(row, 2, QTableWidgetItem(entry.get('chorus_code', '')))
                    self.full_db_table.setItem(row, 3, QTableWidgetItem(entry.get('address', '')))
                
                row += 1
                
        except Exception as e:
            logger.error(f"Erreur lors de la mise à jour de la vue de la base de données: {e}")
        finally:
            # Réactiver les signaux
            self.db_table.blockSignals(False)
            if hasattr(self, 'full_db_table'):
                self.full_db_table.blockSignals(False)
    
    def filter_database(self):
        """Filtre la base de données en fonction de la recherche"""
        search_text = self.search_input.text().lower()
        
        for row in range(self.db_table.rowCount()):
            hide_row = True
            for col in range(self.db_table.columnCount()):
                item = self.db_table.item(row, col)
                if item and search_text in item.text().lower():
                    hide_row = False
                    break
            
            self.db_table.setRowHidden(row, hide_row)
    
    def filter_invoices(self):
        """Filtre les factures en fonction de la recherche"""
        search_text = self.invoice_search_input.text().lower()
        
        for row in range(self.invoice_table.rowCount()):
            hide_row = True
            for col in range(self.invoice_table.columnCount()):
                item = self.invoice_table.item(row, col)
                if item and search_text in item.text().lower():
                    hide_row = False
                    break
            
            self.invoice_table.setRowHidden(row, hide_row)
    
    def on_full_db_item_changed(self, item):
        """Gère les modifications dans le tableau complet de la base de données"""
        row = item.row()
        col = item.column()
        value = item.text()
        
        # Récupérer la clé de l'entrée dans la base de données
        key = list(self.database.data.keys())[row]
        
        # Mettre à jour la valeur dans la base de données
        if col == 0:  # Nom
            self.database.data[key]['name'] = value
        elif col == 1:  # Code Client
            self.database.data[key]['client_code'] = value
        elif col == 2:  # Code Chorus
            self.database.data[key]['chorus_code'] = value
        elif col == 3:  # Adresse
            self.database.data[key]['address'] = value
        
        # Sauvegarder la base de données
        self.database.save_database()
        
        # Mettre à jour le tableau principal
        self.db_table.item(row, col).setText(value)

def __init__(self):
    super().__init__()
    # ... (autres initialisations existantes)
    self._updating_invoice_item = False  # Ajout d'un indicateur de mise à jour
        
def on_invoice_item_changed(self, item):
    """Gère les modifications dans le tableau des factures"""
    if not hasattr(self, '_updating_invoice_item'):
        self._updating_invoice_item = False
        
    if not item or self._updating_invoice_item:  # Ne rien faire si déjà en cours de mise à jour
        return
    
    try:
        self._updating_invoice_item = True
        
        row = item.row()
        col = item.column()
        value = item.text()
        
        # Seule la colonne "Ligne BDD" est éditable
        if col == 3:  # Ligne BDD
            self.update_database_line(row, value)
            # La mise à jour de l'interface est gérée par update_database_line si nécessaire
    finally:
        self._updating_invoice_item = False
    
def update_database_line(self, row, value):
    """Met à jour la ligne de base de données pour une entrée"""
    try:
        # Convertir la valeur en entier si possible
        int_value = int(value) if value and str(value).strip() else 0
        
        # Récupérer l'entrée correspondante
        current_row = 0
        
        for table_name, entries in self.current_tables:
            for entry in entries:
                if current_row == row:
                    # Mettre à jour la valeur dans l'entrée
                    entry['database_line'] = int_value if int_value > 0 else None
                    return
                current_row += 1
        
        logger.warning(f"Ligne {row} non trouvée pour la mise à jour de la ligne de base de données")
    except ValueError as e:
        logger.warning(f"Valeur invalide pour la ligne {row}: {value}")
        logger.debug(f"Détails de l'erreur: {str(e)}")
    
    def load_database_from_excel(self):
        """Charge la base de données depuis un fichier Excel"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Sélectionner un fichier Excel", "", "Fichiers Excel (*.xlsx *.xls)"
        )
        
        if not file_path:
            return
        
        try:
            # Charger le fichier Excel
            df = pd.read_excel(file_path)
            
            # Afficher les colonnes disponibles
            column_mapping_dialog = QDialog(self)
            column_mapping_dialog.setWindowTitle("Mapping des colonnes")
            column_mapping_dialog.setMinimumWidth(600)
            
            layout = QVBoxLayout()
            
            # Créer les combobox pour chaque type de données
            mapping_layout = QGridLayout()
            mapping_layout.addWidget(QLabel("Nom:"), 0, 0)
            mapping_layout.addWidget(QLabel("Code Client:"), 1, 0)
            mapping_layout.addWidget(QLabel("Code Chorus:"), 2, 0)
            mapping_layout.addWidget(QLabel("Adresse:"), 3, 0)
            
            name_combo = QComboBox()
            client_code_combo = QComboBox()
            chorus_code_combo = QComboBox()
            address_combo = QComboBox()
            
            # Ajouter les colonnes disponibles
            columns = ["-- Sélectionner --"] + df.columns.tolist()
            for combo in [name_combo, client_code_combo, chorus_code_combo, address_combo]:
                combo.addItems(columns)
            
            mapping_layout.addWidget(name_combo, 0, 1)
            mapping_layout.addWidget(client_code_combo, 1, 1)
            mapping_layout.addWidget(chorus_code_combo, 2, 1)
            mapping_layout.addWidget(address_combo, 3, 1)
            
            layout.addLayout(mapping_layout)
            
            # Boutons
            button_layout = QHBoxLayout()
            ok_button = QPushButton("OK")
            cancel_button = QPushButton("Annuler")
            
            ok_button.clicked.connect(column_mapping_dialog.accept)
            cancel_button.clicked.connect(column_mapping_dialog.reject)
            
            button_layout.addWidget(ok_button)
            button_layout.addWidget(cancel_button)
            
            layout.addLayout(button_layout)
            
            column_mapping_dialog.setLayout(layout)
            
            if column_mapping_dialog.exec() == QDialog.DialogCode.Accepted:
                # Récupérer le mapping des colonnes
                mapping = {
                    'name': name_combo.currentText() if name_combo.currentIndex() > 0 else None,
                    'client_code': client_code_combo.currentText() if client_code_combo.currentIndex() > 0 else None,
                    'chorus_code': chorus_code_combo.currentText() if chorus_code_combo.currentIndex() > 0 else None,
                    'address': address_combo.currentText() if address_combo.currentIndex() > 0 else None
                }
                
                # Vérifier qu'au moins le nom est mappé
                if not mapping['name']:
                    QMessageBox.warning(self, "Erreur", "Vous devez au moins sélectionner une colonne pour le nom.")
                    return
                
                # Charger les données dans la base de données
                entries_added = self.database.load_from_dataframe(df, mapping)
                
                # Mettre à jour l'affichage
                self.update_database_view()
                
                QMessageBox.information(self, "Succès", f"{entries_added} entrées ajoutées à la base de données.")
        
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors du chargement du fichier Excel:\n{str(e)}")
            logger.error(f"Erreur lors du chargement du fichier Excel: {str(e)}")
    
    def load_invoice_file(self):
        """Charge un fichier de facturation"""
        try:
            # Afficher la boîte de dialogue pour sélectionner le fichier
            file_path, _ = QFileDialog.getOpenFileName(
                self, 
                "Sélectionner un fichier Excel", 
                "", 
                "Fichiers Excel (*.xlsx *.xls *.xlsm);;Tous les fichiers (*)"
            )
            
            if not file_path:
                return
                
            # Afficher un indicateur de chargement
            self.statusBar().showMessage("Chargement du fichier en cours...")
            QApplication.processEvents()  # Mettre à jour l'interface
            
            # Vérifier que le fichier existe et est accessible
            if not os.path.isfile(file_path):
                raise FileNotFoundError(f"Le fichier spécifié n'existe pas: {file_path}")
                
            if not os.access(file_path, os.R_OK):
                raise PermissionError(f"Impossible de lire le fichier. Vérifiez les permissions: {file_path}")
            
            # Sauvegarder le chemin du fichier
            self.current_excel_file = file_path
            
            # Traiter le fichier
            self.process_invoice_file(file_path)
            self.update_preview_table()
            self.save_state(show_message=False)  # Ne pas afficher de message de confirmation
            
            # Activer le bouton d'ouverture avec Excel si nécessaire
            if hasattr(self, 'open_in_excel_btn'):
                self.open_in_excel_btn.setEnabled(True)
            
            # Mettre à jour le titre de la fenêtre avec le nom du fichier
            self.setWindowTitle(f"Gestionnaire de Factures - {os.path.basename(file_path)}")
            self.statusBar().showMessage(f"Fichier chargé avec succès: {os.path.basename(file_path)}", 5000)
            
        except FileNotFoundError as e:
            error_msg = f"Fichier introuvable : {str(e)}"
            logger.error(error_msg)
            QMessageBox.critical(self, "Erreur de fichier", error_msg)
            
        except PermissionError as e:
            error_msg = f"Erreur de permission : {str(e)}"
            logger.error(error_msg)
            QMessageBox.critical(self, "Erreur d'accès", error_msg)
            
        except Exception as e:
            error_msg = f"Erreur lors du chargement du fichier : {str(e)}"
            logger.error(error_msg, exc_info=True)
            QMessageBox.critical(self, "Erreur", error_msg)
            
        finally:
            # S'assurer que la barre d'état est réinitialisée
            self.statusBar().clearMessage()
    
    def open_current_file_in_excel(self):
        """Ouvre le fichier actuel avec l'application par défaut (Excel)"""
        if hasattr(self, 'current_excel_file') and self.current_excel_file:
            try:
                os.startfile(self.current_excel_file)
            except Exception as e:
                QMessageBox.critical(self, "Erreur", f"Impossible d'ouvrir le fichier avec Excel:\n{str(e)}")
                logger.error(f"Erreur lors de l'ouverture du fichier avec Excel: {str(e)}")
        else:
            QMessageBox.information(self, "Information", "Aucun fichier n'est actuellement chargé.")
    
    def _extract_uh_from_sheet_name(self, sheet_name):
        """Extrait le numéro UH du nom de la feuille"""
        import re
        # Exemple: "04 - UH478 PHARMACO 2" -> "04 - UH 478"
        match = re.search(r'(\d+)\s*-\s*UH(\d+)', sheet_name, re.IGNORECASE)
        if match:
            return f"{match.group(1)} - UH {match.group(2)}"
        return ""

    def _extract_invoice_info(self, sheet, start_row, start_col):
        """Extrait les informations d'une facture à partir d'une cellule 'intitulé'"""
        info = {}
        
        # Numéro de facture (7 cellules à droite de 'intitulé')
        num_facture_cell = sheet.cell(row=start_row, column=start_col + 7)
        if num_facture_cell.value and 'facture' in str(num_facture_cell.value).lower():
            # Extraire le numéro de la facture (ex: "Facture N° 1" -> "1")
            match = re.search(r'\d+', str(num_facture_cell.value))
            if match:
                info['num_facture'] = match.group(0)
        
        # Nom de la facture (6 cellules fusionnées à droite de 'intitulé')
        nom_facture = ""
        for col in range(start_col, start_col + 6):
            cell = sheet.cell(row=start_row, column=col)
            if cell.value:
                nom_facture += f"{cell.value} ".strip()
        info['nom_facture'] = nom_facture.strip()
        
        # Adresse de la facture (2 cellules fusionnées en bas à droite de 'intitulé')
        addr_row = start_row + 1
        addr_col = start_col + 5
        if addr_row <= sheet.max_row and addr_col <= sheet.max_column:
            addr_cell = sheet.cell(row=addr_row, column=addr_col)
            if addr_cell.value:
                info['adresse_facture'] = str(addr_cell.value).strip()
        
        return info

    def process_invoice_file(self, file_path):
        """Traite un fichier de facturation selon les spécifications"""
        from openpyxl.utils import get_column_letter
        
        # Charger le fichier Excel en mode lecture seule
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        
        # Initialiser les tables
        self.current_tables = []
        
        # Parcourir toutes les feuilles
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            entries = []
            
            # Extraire le numéro UH du nom de la feuille
            uh_value = self._extract_uh_from_sheet_name(sheet_name)
            
            # Parcourir toutes les cellules pour trouver "intitulé"
            for row_idx, row in enumerate(sheet.iter_rows(), 1):
                for col_idx, cell in enumerate(row, 1):
                    if cell.value and 'intitulé' in str(cell.value).lower():
                        # Extraire les informations de la facture
                        info = self._extract_invoice_info(sheet, row_idx, col_idx)
                        
                        # Créer une entrée pour cette facture
                        entry = {
                            'original_name': info.get('nom_facture', ''),
                            'normalized_name': normalize_text(info.get('nom_facture', '')),
                            'row': row_idx,
                            'uh': uh_value,
                            'num_facture': info.get('num_facture', ''),
                            'nom_facture': info.get('nom_facture', ''),
                            'adresse_facture': info.get('adresse_facture', ''),
                            'client_code': "",
                            'chorus_code': "",
                            'database_line': None,
                            'confidence': 0
                        }
                        
                        # Si les codes ne sont pas dans le fichier, essayer de les trouver dans la base de données
                        if not entry['client_code'] or not entry['chorus_code']:
                            normalized_name = entry['normalized_name']
                            best_match = None
                            best_score = 0
                            
                            for key, db_entry in self.database.data.items():
                                if normalized_name == key:
                                    best_match = db_entry
                                    best_score = 100
                                    break
                                
                                # Recherche floue
                                if normalized_name in key or key in normalized_name:
                                    score = len(key) / max(len(normalized_name), 1) * 100
                                    if score > best_score:
                                        best_match = db_entry
                                        best_score = score
                            
                            # Mettre à jour avec les données de la base de données
                            if best_match:
                                if not entry['client_code']:
                                    entry['client_code'] = best_match.get('client_code', '')
                                if not entry['chorus_code']:
                                    entry['chorus_code'] = best_match.get('chorus_code', '')
                                entry['confidence'] = best_score
                        
                        entries.append(entry)
            
            if entries:
                self.current_tables.append((sheet_name, entries))
        
        # Fermer le workbook
        workbook.close()
    
    def update_preview_table(self):
        """Met à jour l'affichage du tableau des factures avec les nouvelles colonnes"""
        # Désactiver le tri pendant la mise à jour pour éviter les problèmes
        self.invoice_table.setSortingEnabled(False)
        
        # Sauvegarder l'état de la sélection et du défilement
        current_row = self.invoice_table.currentRow()
        scroll_pos = self.invoice_table.verticalScrollBar().value()
        
        # Réinitialiser le tableau
        self.invoice_table.setRowCount(0)
        
        if not self.current_tables:
            return
        
        # Statistiques
        found = 0
        not_found = 0
        
        # Remplir le tableau
        row = 0
        for table_name, entries in self.current_tables:
            for entry in entries:
                self.invoice_table.insertRow(row)
                
                # Récupérer les valeurs avec des valeurs par défaut si non présentes
                uh = entry.get('uh', '')
                num_facture = entry.get('num_facture', '')
                nom_facture = entry.get('nom_facture', entry.get('original_name', ''))
                adresse_facture = entry.get('adresse_facture', '')
                code_client = entry.get('client_code', '')
                code_chorus = entry.get('chorus_code', '')
                ligne_bdd = str(entry.get('database_line', '')) if entry.get('database_line') else ""
                validation = "Validé" if entry.get('confidence', 0) >= 90 else "À valider"
                
                # Remplir chaque colonne
                # Colonnes non modifiables
                self.invoice_table.setItem(row, 0, QTableWidgetItem(str(uh)))  # UH (0)
                self.invoice_table.setItem(row, 1, QTableWidgetItem(str(num_facture)))  # N°Facture (1)
                self.invoice_table.setItem(row, 2, QTableWidgetItem(nom_facture))  # Nom facture (2)
                self.invoice_table.setItem(row, 3, QTableWidgetItem(adresse_facture))  # Adresse facture (3)
                
                # Colonnes modifiables
                # Code client (4)
                client_item = QTableWidgetItem(str(code_client))
                self.invoice_table.setItem(row, 4, client_item)
                
                # Code Chorus (5)
                chorus_item = QTableWidgetItem(str(code_chorus))
                self.invoice_table.setItem(row, 5, chorus_item)
                
                # Ligne BDD (6)
                ligne_bdd_item = QTableWidgetItem(ligne_bdd)
                self.invoice_table.setItem(row, 6, ligne_bdd_item)
                
                # Colonne Validation (7) - Bouton
                btn_validate = QPushButton("Valider")
                btn_validate.setProperty("row", row)
                btn_validate.setStyleSheet("""
                    QPushButton {
                        background-color: #4CAF50;
                        color: white;
                        border: none;
                        padding: 5px 10px;
                        border-radius: 4px;
                    }
                    QPushButton:hover {
                        background-color: #45a049;
                    }
                    QPushButton:pressed {
                        background-color: #3e8e41;
                    }
                """)
                btn_validate.clicked.connect(self.validate_invoice)
                
                # Créer un widget pour contenir le bouton
                widget = QWidget()
                layout = QHBoxLayout(widget)
                layout.addWidget(btn_validate)
                layout.setAlignment(Qt.AlignCenter)
                layout.setContentsMargins(0, 0, 0, 0)
                widget.setLayout(layout)
                
                # Ajouter le widget à la cellule
                self.invoice_table.setCellWidget(row, 7, widget)
                
                # Mettre à jour les statistiques
                if validation == "Validé":
                    found += 1
                else:
                    not_found += 1
                
                row += 1
        
        # Restaurer la position de défilement et la sélection
        if 0 <= current_row < self.invoice_table.rowCount():
            self.invoice_table.scrollToItem(self.invoice_table.item(current_row, 0))
            self.invoice_table.selectRow(current_row)
        
        # Réactiver le tri
        self.invoice_table.setSortingEnabled(True)
        
        # Mettre à jour les statistiques
        self.stats_widget.update_stats(found, not_found)
        
        # Connecter le signal itemChanged pour la validation des entrées
        self.invoice_table.itemChanged.connect(self.validate_inputs)
    
    def validate_inputs(self, item):
        """Valide les entrées dans les cellules"""
        if item.column() == 6:  # Colonne Ligne BDD (index 6)
            try:
                value = item.text().strip()
                if value:  # Si le champ n'est pas vide
                    num = int(value)
                    if num < 0 or num > 99999:
                        QMessageBox.warning(None, "Valeur invalide", "La ligne BDD doit être un nombre entre 0 et 99999")
                        item.setText("")
                    else:
                        # Mettre à jour la valeur dans les données
                        for table_name, entries in self.current_tables:
                            if 0 <= item.row() < len(entries):
                                entries[item.row()]['database_line'] = num
                                break
            except ValueError:
                QMessageBox.warning(None, "Valeur invalide", "Veuillez entrer un nombre valide")
                item.setText("")
    
    def validate_invoice(self):
        """Valide une facture"""
        button = self.sender()
        if button:
            row = button.property("row")
            if row is not None:
                # Mettre à jour l'état de validation dans les données
                for table_name, entries in self.current_tables:
                    if 0 <= row < len(entries):
                        entries[row]['confidence'] = 100  # Marquer comme validé
                        
                # Mettre à jour l'affichage
                self.update_preview_table()
                
                # Sauvegarder les modifications
                self.save_state()
    
def update_preview_table(self):
    """Met à jour l'affichage du tableau des factures avec les nouvelles colonnes"""
    # Désactiver le tri pendant la mise à jour pour éviter les problèmes
    self.invoice_table.setSortingEnabled(False)
    
    # Sauvegarder l'état de la sélection et du défilement
    current_row = self.invoice_table.currentRow()
    scroll_pos = self.invoice_table.verticalScrollBar().value()
    
    # Réinitialiser le tableau
    self.invoice_table.setRowCount(0)
    
    if not self.current_tables:
        return
    
    # Statistiques
    found = 0
    not_found = 0
    
    # Remplir le tableau
    row = 0
    for table_name, entries in self.current_tables:
        for entry in entries:
            self.invoice_table.insertRow(row)
            
            # Récupérer les valeurs avec des valeurs par défaut si non présentes
            uh = entry.get('uh', '')
            num_facture = entry.get('num_facture', '')
            nom_facture = entry.get('nom_facture', entry.get('original_name', ''))
            adresse_facture = entry.get('adresse_facture', '')
            code_client = entry.get('client_code', '')
            code_chorus = entry.get('chorus_code', '')
            ligne_bdd = str(entry.get('database_line', '')) if entry.get('database_line') else ""
            validation = "Validé" if entry.get('confidence', 0) >= 90 else "À valider"
            
            # Remplir chaque colonne
            # Colonnes non modifiables
            self.invoice_table.setItem(row, 0, QTableWidgetItem(str(uh)))  # UH (0)
            self.invoice_table.setItem(row, 1, QTableWidgetItem(str(num_facture)))  # N°Facture (1)
            self.invoice_table.setItem(row, 2, QTableWidgetItem(nom_facture))  # Nom facture (2)
            self.invoice_table.setItem(row, 3, QTableWidgetItem(adresse_facture))  # Adresse facture (3)
            
            # Colonnes modifiables
            # Code client (4)
            client_item = QTableWidgetItem(str(code_client))
            self.invoice_table.setItem(row, 4, client_item)
            
            # Code Chorus (5)
            chorus_item = QTableWidgetItem(str(code_chorus))
            self.invoice_table.setItem(row, 5, chorus_item)
            
            # Ligne BDD (6)
            ligne_bdd_item = QTableWidgetItem(ligne_bdd)
            self.invoice_table.setItem(row, 6, ligne_bdd_item)
            
            # Colonne Validation (7) - Bouton
            btn_validate = QPushButton("Valider")
            btn_validate.setProperty("row", row)
            btn_validate.setStyleSheet("""
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    border: none;
                    padding: 5px 10px;
                    border-radius: 4px;
                }
                QPushButton:hover {
                    background-color: #45a049;
                }
                QPushButton:pressed {
                    background-color: #3e8e41;
                }
            """)
            btn_validate.clicked.connect(self.validate_invoice)
            
            # Créer un widget pour contenir le bouton
            widget = QWidget()
            layout = QHBoxLayout(widget)
            layout.addWidget(btn_validate)
            layout.setAlignment(Qt.AlignCenter)
            layout.setContentsMargins(0, 0, 0, 0)
            widget.setLayout(layout)
            
            # Ajouter le widget à la cellule
            self.invoice_table.setCellWidget(row, 7, widget)
            
            # Mettre à jour les statistiques
            if validation == "Validé":
                found += 1
            else:
                not_found += 1
            
            row += 1
    
    # Restaurer la position de défilement et la sélection
    if 0 <= current_row < self.invoice_table.rowCount():
        self.invoice_table.scrollToItem(self.invoice_table.item(current_row, 0))
        self.invoice_table.selectRow(current_row)
    
    # Réactiver le tri
    self.invoice_table.setSortingEnabled(True)
    
    # Mettre à jour les statistiques
    self.stats_widget.update_stats(found, not_found)
    
    # Connecter le signal itemChanged pour la validation des entrées
    self.invoice_table.itemChanged.connect(self.validate_inputs)

def validate_inputs(self, item):
    """Valide les entrées dans les cellules"""
    if item.column() == 6:  # Colonne Ligne BDD (index 6)
        try:
            value = item.text().strip()
            if value:  # Si le champ n'est pas vide
                num = int(value)
                if num < 0 or num > 99999:
                    QMessageBox.warning(None, "Valeur invalide", "La ligne BDD doit être un nombre entre 0 et 99999")
                    item.setText("")
                else:
                    # Mettre à jour la valeur dans les données
                    for table_name, entries in self.current_tables:
                        if 0 <= item.row() < len(entries):
                            entries[item.row()]['database_line'] = num
                            break
        except ValueError:
            QMessageBox.warning(None, "Valeur invalide", "Veuillez entrer un nombre valide")
            item.setText("")

def validate_invoice(self):
    """Valide une facture"""
    button = self.sender()
    if button:
        row = button.property("row")
        if row is not None:
            # Mettre à jour l'état de validation dans les données
            for table_name, entries in self.current_tables:
                if 0 <= row < len(entries):
                    entries[row]['confidence'] = 100  # Marquer comme validé
                    
            # Mettre à jour l'affichage
            self.update_preview_table()
            
            # Sauvegarder les modifications
            self.save_state()

def save_invoice_file(self):
    """Saisie des codes dans le fichier de facturation pour les lignes validées (en bleu)"""
    if not self.current_excel_file or not self.current_tables:
        QMessageBox.warning(self, "Attention", "Aucun fichier de facturation n'est chargé.")
        return
    
    try:
        # Créer un nouveau nom de fichier
        file_name, file_ext = os.path.splitext(self.current_excel_file)
        new_file_path = f"{file_name}_updated{file_ext}"
        
        # Créer un mapping des entrées par leur identifiant unique (UH + N°Facture)
        entries_map = {}
        for table_name, entries in self.current_tables:
            for entry in entries:
                uh = entry.get('uh', '')
                num_facture = entry.get('num_facture', '')
                if uh and num_facture:
                    entries_map[f"{uh}_{num_facture}"] = entry
        
        # Parcourir les lignes du tableau
        for row in range(self.invoice_table.rowCount()):
            uh_item = self.invoice_table.item(row, 0)  # Colonne UH
            num_facture_item = self.invoice_table.item(row, 1)  # Colonne N°Facture
            
            if not uh_item or not num_facture_item:
                continue
                
            uh = uh_item.text().strip()
            num_facture = num_facture_item.text().strip()
            
            if not uh or not num_facture:
                continue
            
            entry_key = f"{uh}_{num_facture}"
            if entry_key not in entries_map:
                continue
                
            entry = entries_map[entry_key]
            
            # Mettre à jour les champs modifiables
            for col, field in [(4, 'client_code'), (5, 'chorus_code'), (6, 'database_line')]:
                item = self.invoice_table.item(row, col)
                if item:
                    value = item.text().strip()
                    if col == 6:  # Ligne BDD - convertir en entier si possible
                        try:
                            value = int(value) if value else None
                        except ValueError:
                            value = None
                    entry[field] = value
    except Exception as e:
        QMessageBox.critical(
            self,
            "Erreur lors de la mise à jour",
            f"Une erreur est survenue lors de la mise à jour des données :\n{str(e)}"
        )
        logger.error(f"Erreur lors de la mise à jour des données: {str(e)}")
        return False
    
    return True
    
    def export_database_to_excel(self):
        """Exporte la base de données vers un fichier Excel"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Exporter la base de données", "", "Fichiers Excel (*.xlsx)"
        )
        
        if not file_path:
            return
        
        if not file_path.endswith('.xlsx'):
            file_path += '.xlsx'
        
        try:
            # Créer un DataFrame à partir de la base de données
            data = []
            for key, entry in self.database.data.items():
                data.append({
                    'Nom': entry.get('name', ''),
                    'Code Client': entry.get('client_code', ''),
                    'Code Chorus': entry.get('chorus_code', ''),
                    'Adresse': entry.get('address', '')
                })
            
            df = pd.DataFrame(data)
            
            # Exporter vers Excel
            df.to_excel(file_path, index=False)
            
            QMessageBox.information(
                self,
                "Succès",
                f"La base de données a été exportée vers:\n{file_path}"
            )
        
        except Exception as e:
            QMessageBox.critical(
                self,
                "Erreur",
                f"Erreur lors de l'exportation de la base de données:\n{str(e)}"
            )
            logger.error(f"Erreur lors de l'exportation de la base de données: {str(e)}")
    
    def import_database(self):
        """Importe une base de données depuis un fichier JSON ou Excel"""
        try:
            logger.info("Démarrage de l'importation de la base de données")
            file_path, _ = QFileDialog.getOpenFileName(
                self, "Importer une base de données", "", "Tous les fichiers supportés (*.json *.xlsx *.xls);;Fichiers JSON (*.json);;Fichiers Excel (*.xlsx *.xls)"
            )
            
            if not file_path:
                logger.info("Importation annulée par l'utilisateur")
                return
            
            logger.info(f"Fichier sélectionné pour importation: {file_path}")
            
            # Vérifier que le fichier existe
            if not os.path.exists(file_path):
                logger.error(f"Le fichier n'existe pas: {file_path}")
                QMessageBox.critical(self, "Erreur", f"Le fichier n'existe pas: {file_path}")
                return
                
            # Déterminer le type de fichier
            file_ext = os.path.splitext(file_path)[1].lower()
            logger.info(f"Extension du fichier: {file_ext}")
            
            if file_ext == '.json':
                # Importer depuis JSON
                logger.info("Importation d'un fichier JSON")
                try:
                    if self.database.load_file(file_path):
                        # Mettre à jour l'affichage
                        logger.info("Mise à jour de l'affichage après importation JSON")
                        self.update_database_view()
                        
                        # Sauvegarder la base de données
                        logger.info("Sauvegarde de la base de données après importation JSON")
                        self.database.save_database()
                        
                        QMessageBox.information(
                            self,
                            "Succès",
                            f"La base de données a été importée avec succès.\n{len(self.database.data)} entrées chargées."
                        )
                        logger.info(f"Importation JSON réussie: {len(self.database.data)} entrées chargées")
                    else:
                        logger.warning("Échec de l'importation du fichier JSON")
                        QMessageBox.warning(
                            self,
                            "Attention",
                            "Impossible de charger le fichier JSON."
                        )
                except Exception as json_error:
                    logger.error(f"Erreur lors de l'importation JSON: {str(json_error)}")
                    QMessageBox.critical(
                        self,
                        "Erreur",
                        f"Une erreur est survenue lors de l'importation du fichier JSON : {str(json_error)}"
                    )
            elif file_ext in ['.xlsx', '.xls']:
                # Importer depuis Excel
                logger.info("Importation d'un fichier Excel")
                try:
                    # Charger le fichier Excel
                    logger.info("Lecture du fichier Excel avec pandas")
                    try:
                        df = pd.read_excel(file_path)
                        logger.info(f"Fichier Excel lu avec succès: {len(df)} lignes, colonnes: {df.columns.tolist()}")
                    except Exception as excel_read_error:
                        logger.error(f"Erreur lors de la lecture du fichier Excel: {str(excel_read_error)}")
                        QMessageBox.critical(
                            self,
                            "Erreur",
                            f"Impossible de lire le fichier Excel. Vérifiez qu'il n'est pas corrompu ou ouvert dans une autre application.\n\nErreur: {str(excel_read_error)}"
                        )
                        return
                    
                    # Vérifier que le DataFrame n'est pas vide
                    if df.empty:
                        logger.warning("Le fichier Excel est vide")
                        QMessageBox.warning(self, "Attention", "Le fichier Excel est vide. Aucune donnée à importer.")
                        return
                    
                    # Afficher les colonnes disponibles
                    column_mapping_dialog = QDialog(self)
                    column_mapping_dialog.setWindowTitle("Mapping des colonnes")
                    column_mapping_dialog.setMinimumWidth(600)
                    
                    layout = QVBoxLayout()
                    
                    # Ajouter un label explicatif
                    layout.addWidget(QLabel("Sélectionnez les colonnes correspondantes dans votre fichier Excel:"))
                    
                    # Créer les combobox pour chaque type de données
                    mapping_layout = QGridLayout()
                    mapping_layout.addWidget(QLabel("Nom (obligatoire):"), 0, 0)
                    mapping_layout.addWidget(QLabel("Code Client (optionnel):"), 1, 0)
                    mapping_layout.addWidget(QLabel("Code Chorus (optionnel):"), 2, 0)
                    mapping_layout.addWidget(QLabel("Adresse (optionnel):"), 3, 0)
                    
                    name_combo = QComboBox()
                    client_code_combo = QComboBox()
                    chorus_code_combo = QComboBox()
                    address_combo = QComboBox()
                    
                    # Ajouter les colonnes disponibles
                    columns = ["-- Sélectionner --"] + df.columns.tolist()
                    for combo in [name_combo, client_code_combo, chorus_code_combo, address_combo]:
                        combo.addItems(columns)
                    
                    # Essayer de trouver automatiquement les colonnes pertinentes
                    for i, col in enumerate(df.columns):
                        col_lower = col.lower()
                        if "nom" in col_lower or "name" in col_lower or "client" in col_lower:
                            name_combo.setCurrentIndex(i + 1)  # +1 car le premier item est "-- Sélectionner --"
                        elif "code" in col_lower and ("client" in col_lower):
                            client_code_combo.setCurrentIndex(i + 1)
                        elif "chorus" in col_lower or "code chorus" in col_lower:
                            chorus_code_combo.setCurrentIndex(i + 1)
                        elif "adresse" in col_lower or "address" in col_lower:
                            address_combo.setCurrentIndex(i + 1)
                    
                    mapping_layout.addWidget(name_combo, 0, 1)
                    mapping_layout.addWidget(client_code_combo, 1, 1)
                    mapping_layout.addWidget(chorus_code_combo, 2, 1)
                    mapping_layout.addWidget(address_combo, 3, 1)
                    
                    layout.addLayout(mapping_layout)
                    
                    # Boutons
                    button_layout = QHBoxLayout()
                    ok_button = QPushButton("Importer")
                    cancel_button = QPushButton("Annuler")
                    
                    ok_button.clicked.connect(column_mapping_dialog.accept)
                    cancel_button.clicked.connect(column_mapping_dialog.reject)
                    
                    button_layout.addWidget(ok_button)
                    button_layout.addWidget(cancel_button)
                    
                    layout.addLayout(button_layout)
                    
                    column_mapping_dialog.setLayout(layout)
                    
                    logger.info("Affichage de la boîte de dialogue de mapping des colonnes")
                    if column_mapping_dialog.exec() == QDialog.DialogCode.Accepted:
                        # Récupérer le mapping des colonnes
                        mapping = {
                            'name': name_combo.currentText() if name_combo.currentIndex() > 0 else None,
                            'client_code': client_code_combo.currentText() if client_code_combo.currentIndex() > 0 else None,
                            'chorus_code': chorus_code_combo.currentText() if chorus_code_combo.currentIndex() > 0 else None,
                            'address': address_combo.currentText() if address_combo.currentIndex() > 0 else None
                        }
                        
                        logger.info(f"Mapping des colonnes sélectionné: {mapping}")
                        
                        # Vérifier qu'au moins le nom est mappé
                        if not mapping['name']:
                            logger.warning("Aucune colonne sélectionnée pour le nom")
                            QMessageBox.warning(self, "Erreur", "Vous devez au moins sélectionner une colonne pour le nom.")
                            return
                        
                        try:
                            # Charger les données dans la base de données
                            logger.info("Chargement des données dans la base de données")
                            entries_added = self.database.load_from_dataframe(df, mapping)
                            logger.info(f"Données chargées avec succès: {entries_added} entrées ajoutées")
                            
                            # Mettre à jour l'affichage
                            logger.info("Mise à jour de l'affichage après importation Excel")
                            self.update_database_view()
                            
                            QMessageBox.information(self, "Succès", f"{entries_added} entrées ajoutées à la base de données.")
                        except Exception as load_error:
                            logger.error(f"Erreur lors du chargement des données dans la base: {str(load_error)}")
                            QMessageBox.critical(
                                self,
                                "Erreur",
                                f"Une erreur est survenue lors du chargement des données dans la base de données.\n\nErreur: {str(load_error)}"
                            )
                    else:
                        logger.info("Importation Excel annulée par l'utilisateur")
                except Exception as excel_error:
                    logger.error(f"Erreur lors de l'importation Excel: {str(excel_error)}")
                    QMessageBox.critical(
                        self,
                        "Erreur",
                        f"Une erreur est survenue lors de l'importation du fichier Excel.\n\nErreur: {str(excel_error)}"
                    )
            else:
                logger.warning(f"Format de fichier non supporté: {file_ext}")
                QMessageBox.warning(
                    self,
                    "Attention",
                    "Format de fichier non supporté. Veuillez sélectionner un fichier JSON ou Excel."
                )
        except Exception as e:
            logger.error(f"Erreur générale lors de l'importation de la base de données: {str(e)}")
            QMessageBox.critical(
                self,
                "Erreur",
                f"Une erreur est survenue lors de l'importation de la base de données.\n\nErreur: {str(e)}"
            )
    
    def clear_database(self):
        """Vide la base de données"""
        reply = QMessageBox.question(
            self,
            "Confirmation",
            "Êtes-vous sûr de vouloir vider la base de données ?\nCette action est irréversible.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.database.data = {}
                self.database.save_database()
                self.update_database_view()
                QMessageBox.information(
                    self,
                    "Succès",
                    "La base de données a été vidée avec succès."
                )
            except Exception as e:
                QMessageBox.critical(
                    self,
                    "Erreur",
                    f"Une erreur est survenue lors de la suppression de la base de données : {str(e)}"
                )
    
    def save_state(self, show_message=True):
        """Sauvegarde l'état actuel de l'application"""
        try:
            state = {
                'current_excel_file': self.current_excel_file,
                'manual_matches': self.manual_matches,
                'current_tables': self.current_tables
            }
            
            with open(self.state_file, 'w', encoding='utf-8') as f:
                json.dump(state, f, ensure_ascii=False, default=lambda o: str(o))
            
            if show_message:
                logger.info("État sauvegardé avec succès")
        
        except Exception as e:
            logger.error(f"Erreur lors de la sauvegarde de l'état: {str(e)}")
            if show_message:
                QMessageBox.warning(
                    self,
                    "Erreur",
                    f"Impossible de sauvegarder l'état:\n{str(e)}"
                )
    
    def load_state_async(self):
        """Charge l'état précédent de l'application de manière asynchrone"""
        # Afficher l'indicateur de chargement dans la barre de statut
        self.statusBar().showMessage("Chargement de la base de données...")
        self.loading_indicator.setVisible(True)
        self.statusBar().addPermanentWidget(self.loading_indicator)
        
        # Charger la base de données de manière asynchrone
        self.database.ensure_loaded(self.on_database_loaded)
    
    def on_database_loaded(self):
        """Callback appelé lorsque la base de données est chargée"""
        # Mettre à jour l'affichage de la base de données
        self.update_database_view()
        
        # Masquer l'indicateur de chargement
        self.loading_indicator.setVisible(False)
        self.statusBar().removeWidget(self.loading_indicator)
        self.statusBar().showMessage("Prêt", 3000)  # Afficher "Prêt" pendant 3 secondes
        
        # S'assurer que les widgets dock sont visibles
        self.restore_dock_widgets()
    
    def restore_dock_widgets(self):
        """Restaure les widgets dock et force leur affichage"""
        # Vérifier si les widgets existent
        if hasattr(self, 'db_dock') and hasattr(self, 'invoice_dock'):
            # Forcer la visibilité des widgets
            self.db_dock.setVisible(True)
            self.invoice_dock.setVisible(True)
            
            # S'assurer qu'ils sont dans la fenêtre principale
            self.db_dock.setFloating(False)
            self.invoice_dock.setFloating(False)
            
            # Déterminer la zone de dock appropriée
            self.addDockWidget(Qt.DockWidgetArea.LeftDockWidgetArea, self.db_dock)
            self.addDockWidget(Qt.DockWidgetArea.RightDockWidgetArea, self.invoice_dock)
            
            # Forcer la mise à jour de l'interface
            self.db_dock.raise_()
            self.invoice_dock.raise_()
            
            # Journaliser l'action
            logger.info("Widgets dock restaurés et forcés à être visibles")
        else:
            logger.warning("Tentative de restauration des widgets dock avant leur création")
    
    def load_app_state(self):
        """Charge l'état de l'application depuis le fichier JSON"""
        try:
            if os.path.exists(self.state_file):
                with open(self.state_file, 'r', encoding='utf-8') as f:
                    state = json.load(f)
                
                self.current_excel_file = state.get('current_excel_file')
                self.manual_matches = state.get('manual_matches', {})
                self.current_tables = state.get('current_tables', [])
                
                # Si un fichier Excel était chargé, mettre à jour l'interface
                if self.current_excel_file and os.path.exists(self.current_excel_file):
                    self.update_preview_table()
                
                logger.info("État précédent chargé avec succès")
        
        except Exception as e:
            logger.error(f"Erreur lors du chargement de l'état: {str(e)}")
            # Ne pas afficher de message d'erreur à l'utilisateur au démarrage
    
    def load_state(self):
        """Charge l'état précédent de l'application (méthode synchrone pour compatibilité)"""
        try:
            # Charger d'abord la base de données
            self.database.ensure_loaded()
            self.update_database_view()
            
            # Puis charger l'état de l'application
            self.load_app_state()
        
        except Exception as e:
            logger.error(f"Erreur lors du chargement de l'état: {str(e)}")
            # Ne pas afficher de message d'erreur à l'utilisateur au démarrage
            
    def prevent_close_and_redock(self, dock_widget, event):
        """Empêche la fermeture des widgets détachables et les redock si nécessaire"""
        # Toujours ignorer l'événement de fermeture
        event.ignore()
        
        if dock_widget.isFloating():
            # Si le widget est flottant, le redocker dans la fenêtre principale
            dock_widget.setFloating(False)
            
            # Déterminer la zone de dock appropriée
            if dock_widget == self.db_dock:
                self.addDockWidget(Qt.DockWidgetArea.LeftDockWidgetArea, dock_widget)
            elif dock_widget == self.invoice_dock:
                self.addDockWidget(Qt.DockWidgetArea.RightDockWidgetArea, dock_widget)
        
        # S'assurer que le widget est visible
        dock_widget.show()
        dock_widget.raise_()
        
        logger.info(f"Widget {dock_widget.windowTitle()} réintégré dans la fenêtre principale")
    
    def ensure_maximize_button(self, window):
        """S'assure que le bouton maximiser est actif"""
        # Forcer la mise à jour des attributs de fenêtre
        window.setAttribute(Qt.WidgetAttribute.WA_WState_MaximizeButtonShown, True)
        
        # Forcer la mise à jour de la fenêtre
        window.setWindowState(window.windowState())
        window.show()
        logger.info("Bouton maximiser activé")
    
    def setup_dock_widgets(self):
        """Configuration initiale des widgets de dock"""
        # Configuration du dock de la base de données
        self.db_dock.setFeatures(
            QDockWidget.DockWidgetFeature.DockWidgetMovable |
            QDockWidget.DockWidgetFeature.DockWidgetFloatable
        )
        self.db_dock.setAllowedAreas(Qt.DockWidgetArea.AllDockWidgetAreas)
        
        # Configuration du dock des factures
        self.invoice_dock.setFeatures(
            QDockWidget.DockWidgetFeature.DockWidgetMovable |
            QDockWidget.DockWidgetFeature.DockWidgetFloatable
        )
        self.invoice_dock.setAllowedAreas(Qt.DockWidgetArea.AllDockWidgetAreas)
        
        # Connecter les signaux pour la gestion des fenêtres flottantes
        self.connect_signals()
        
    def connect_signals(self):
        """Connecte les signaux des widgets dock pour gérer les fenêtres flottantes"""
        # Connecter le signal topLevelChanged des widgets dock à la méthode configure_floating_window
        self.db_dock.topLevelChanged.connect(
            lambda floating: self.configure_floating_window(self.db_dock, floating)
        )
        
        self.invoice_dock.topLevelChanged.connect(
            lambda floating: self.configure_floating_window(self.invoice_dock, floating)
        )
            
if __name__ == '__main__':

    app = QApplication(sys.argv)
    app.setStyle(QStyleFactory.create('Fusion'))

    
    # Configuration de la palette de couleurs globale
    palette = QPalette()
    palette.setColor(QPalette.ColorRole.Window, QColor(CYBERPUNK_COLORS['background']))
    palette.setColor(QPalette.ColorRole.WindowText, QColor(CYBERPUNK_COLORS['text']))
    palette.setColor(QPalette.ColorRole.Base, QColor(CYBERPUNK_COLORS['secondary_bg']))
    palette.setColor(QPalette.ColorRole.AlternateBase, QColor(CYBERPUNK_COLORS['background']))
    palette.setColor(QPalette.ColorRole.ToolTipBase, QColor(CYBERPUNK_COLORS['text']))
    palette.setColor(QPalette.ColorRole.ToolTipText, QColor(CYBERPUNK_COLORS['background']))
    palette.setColor(QPalette.ColorRole.Text, QColor(CYBERPUNK_COLORS['text']))
    palette.setColor(QPalette.ColorRole.Button, QColor(CYBERPUNK_COLORS['secondary_bg']))
    palette.setColor(QPalette.ColorRole.ButtonText, QColor(CYBERPUNK_COLORS['text']))
    palette.setColor(QPalette.ColorRole.Highlight, QColor(CYBERPUNK_COLORS['accent1']))
    palette.setColor(QPalette.ColorRole.HighlightedText, QColor(CYBERPUNK_COLORS['text']))
    app.setPalette(palette)
    
    # Créer un écran de démarrage
    splash_pixmap = QPixmap(400, 300)
    splash_pixmap.fill(QColor(CYBERPUNK_COLORS['background']))
    splash = QSplashScreen(splash_pixmap)
    
    # Ajouter du texte à l'écran de démarrage
    splash.showMessage(
        "Chargement du Gestionnaire de Factures...",
        Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignBottom,
        QColor(CYBERPUNK_COLORS['accent1'])
    )
    splash.show()
    app.processEvents()
    
    # Créer la fenêtre principale
    window = MainWindow()
    
    # Définir l'icône de l'application
    icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "icon.png")
    if os.path.exists(icon_path):
        window.setWindowIcon(QIcon(icon_path))
        splash.setWindowIcon(QIcon(icon_path))
    
    # Fermer l'écran de démarrage et afficher la fenêtre principale
    splash.finish(window)
    window.show()
    
    sys.exit(app.exec_())
