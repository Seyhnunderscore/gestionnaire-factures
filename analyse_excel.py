#!/usr/bin/env python3
"""
Script pour analyser le fichier Excel et comprendre pourquoi certaines factures ne sont pas trouvées
"""

import os
import sys
import openpyxl
import re
from openpyxl.utils import get_column_letter
import logging

# Configuration du logging
logging.basicConfig(level=logging.INFO, 
                   format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                   datefmt='%Y-%m-%d %H:%M:%S')

logger = logging.getLogger("ExcelAnalyzer")

def clean_facture_number(facture_num):
    """
    Nettoie et normalise un numéro de facture pour faciliter les comparaisons
    """
    if facture_num is None:
        return ""
    
    # Convertir en texte et mettre en minuscule
    facture_num = str(facture_num).lower().strip()
    
    # Supprimer les préfixes communs
    prefixes = ["facture n°", "facture n ", "facture ", "fact ", "fact. ", "n°", "n ", "numéro ", "numero "]
    for prefix in prefixes:
        if facture_num.startswith(prefix):
            facture_num = facture_num.replace(prefix, "", 1)
    
    # Supprimer les caractères non alphanumériques tout en conservant les chiffres et lettres
    facture_num = re.sub(r'[^\w\d]', '', facture_num)
    
    return facture_num.strip()

def analyze_excel_file(file_path, sheet_name, facture_nums):
    """
    Analyse un fichier Excel pour trouver des numéros de facture spécifiques
    
    Args:
        file_path (str): Chemin vers le fichier Excel
        sheet_name (str): Nom de la feuille à analyser
        facture_nums (list): Liste des numéros de facture à rechercher
    """
    logger.info(f"Analyse du fichier: {file_path}")
    logger.info(f"Recherche dans la feuille: {sheet_name}")
    logger.info(f"Numéros de facture à rechercher: {facture_nums}")
    
    try:
        # Charger le fichier Excel
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # Vérifier si la feuille existe
        sheet_found = False
        target_sheet = None
        
        for sheet in wb.sheetnames:
            if sheet_name.lower() in sheet.lower():
                sheet_found = True
                target_sheet = wb[sheet]
                logger.info(f"Feuille trouvée: {sheet}")
                break
        
        if not sheet_found:
            logger.error(f"Feuille '{sheet_name}' non trouvée dans le fichier")
            logger.info(f"Feuilles disponibles: {wb.sheetnames}")
            return
        
        # Analyser la feuille pour trouver les numéros de facture
        max_row = target_sheet.max_row
        max_col = target_sheet.max_column
        
        logger.info(f"Dimensions de la feuille: {max_row} lignes x {max_col} colonnes")
        
        # Afficher les 10 premières lignes pour comprendre la structure
        logger.info("=== Aperçu des 10 premières lignes ===")
        for row in range(1, min(11, max_row + 1)):
            row_values = []
            for col in range(1, min(10, max_col + 1)):
                cell_value = target_sheet.cell(row=row, column=col).value
                row_values.append(str(cell_value) if cell_value is not None else "")
            logger.info(f"Ligne {row}: {' | '.join(row_values)}")
        
        # Rechercher chaque numéro de facture
        for facture_num in facture_nums:
            logger.info(f"\n=== Recherche de '{facture_num}' ===")
            facture_num_pur = clean_facture_number(facture_num)
            logger.info(f"Numéro nettoyé: '{facture_num_pur}'")
            
            found = False
            potential_matches = []
            
            # Recherche exhaustive dans toutes les cellules
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell_value = target_sheet.cell(row=row, column=col).value
                    if cell_value is not None:
                        cell_text = str(cell_value)
                        cell_clean = clean_facture_number(cell_text)
                        
                        # Vérifier différents types de correspondances
                        exact_match = cell_text.lower() == facture_num.lower()
                        clean_match = cell_clean == facture_num_pur
                        contains_match = facture_num_pur in cell_clean
                        
                        # Si une correspondance est trouvée ou potentielle
                        if exact_match or clean_match:
                            found = True
                            logger.info(f"TROUVÉ à la cellule {get_column_letter(col)}{row}: '{cell_text}'")
                            # Afficher les cellules environnantes pour contexte
                            logger.info("Contexte autour de la cellule:")
                            for r in range(max(1, row-2), min(max_row+1, row+3)):
                                row_values = []
                                for c in range(max(1, col-2), min(max_col+1, col+3)):
                                    cell = target_sheet.cell(row=r, column=c).value
                                    row_values.append(str(cell) if cell is not None else "")
                                logger.info(f"  Ligne {r}: {' | '.join(row_values)}")
                        elif contains_match:
                            potential_matches.append((row, col, cell_text))
            
            # Afficher les correspondances potentielles
            if potential_matches and not found:
                logger.info("Correspondances potentielles trouvées:")
                for row, col, text in potential_matches:
                    logger.info(f"  Cellule {get_column_letter(col)}{row}: '{text}'")
            
            if not found and not potential_matches:
                logger.warning(f"Numéro de facture '{facture_num}' non trouvé dans la feuille")
        
    except Exception as e:
        logger.error(f"Erreur lors de l'analyse du fichier: {e}")

if __name__ == "__main__":
    # Fichier à analyser
    file_path = r"C:\Users\james\Desktop\BUREAU TRAVAIL\02 - FACTURATION FEVRIER 2025.xlsm"
    
    # Feuille à analyser
    sheet_name = "05 - UH485 PHARMACO"
    
    # Numéros de facture à rechercher
    facture_nums = ["Facture N° 66", "Facture N° 74"]
    
    # Analyser le fichier
    analyze_excel_file(file_path, sheet_name, facture_nums)
